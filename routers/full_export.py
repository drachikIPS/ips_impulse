"""
Full project database Excel export.

One endpoint that produces a single .xlsx workbook with one sheet per
domain area (contacts, packages, areas, meeting points, tasks,
budget overview, procurement plan + register, scope changes, documents,
ITP, punch list, daily reports, workers/subcontractors, work permits,
LOTO, safety observations/incidents/toolboxes, files).

Column layouts mirror the per-module Excel exports already shipped
(routers/meeting_export.py, routers/safety_export.py,
routers/construction.py, routers/attachments.py) so users get the same
columns whether they pull a single module or the full bundle.

Permission: ADMIN or PROJECT_OWNER only — the bundle exposes data from
every module at once, even ones the caller wouldn't normally see.
"""
import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
import models
import auth


router = APIRouter(prefix="/api/projects/full-database", tags=["full-export"])


# ── Style constants (match the existing meeting_export.py palette) ──────────
_DARK_BLUE  = "1E3A5F"
_LIGHT_BLUE = "D6E4F0"
_WHITE      = "FFFFFF"


def _add_sheet(wb, title: str, headers: list[str], rows: list[list]):
    """Create a sheet with the standard header row + alternating zebra fill
    + auto-width. `title` is truncated to 31 chars (Excel limit)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    safe = title[:31]
    ws = wb.create_sheet(title=safe)
    ws.append(headers)
    fill = PatternFill("solid", fgColor=_DARK_BLUE)
    for cell in ws[1]:
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22
    zebra = PatternFill("solid", fgColor=_LIGHT_BLUE)
    for ri, row in enumerate(rows, start=2):
        ws.append(row)
        if ri % 2 == 0:
            for cell in ws[ri]:
                cell.fill = zebra
        for cell in ws[ri]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Auto-width (capped to 60)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    return ws


def _ts(dt) -> str:
    if not dt:
        return ""
    try:
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(dt)


def _user_name(u) -> str:
    if not u:
        return ""
    return getattr(u, "name", None) or getattr(u, "email", "") or ""


def _status_label(s: str) -> str:
    return {
        "DRAFT": "Draft", "SUBMITTED": "Submitted", "RECEIVED": "Received",
        "CLOSED": "Closed", "APPROVED": "Approved", "REJECTED": "Rejected",
        "PENDING": "Pending", "PENDING_REVIEW": "Pending review",
        "UNDER_INVESTIGATION": "Under investigation",
        "ACTION_IN_PROGRESS": "Action in progress",
        "OPEN": "Open", "MONITORING": "Monitoring",
        "IN_PROGRESS": "In progress", "ON_HOLD": "On hold",
        "NOT_STARTED": "Not started", "COMPETING": "Competing",
        "EXCLUDED": "Excluded", "AWAITING": "Awaiting", "AWARDED": "Awarded",
        "REQUEST": "Request", "LOCKED": "Locked", "REFUSED": "Refused",
        "CANCELLED": "Cancelled", "RELEASED": "Released",
        "TO_BE_RELEASED": "To be released",
        "RESOLVED": "Resolved",
        "IN_REVIEW": "In review",
    }.get(s, s or "")


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders — one per domain
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_contacts(wb, db: Session, pid: int):
    rows = (db.query(models.Contact)
              .filter_by(project_id=pid)
              .order_by(models.Contact.name).all())
    data = [[
        c.name or "",
        c.email or "",
        c.company or "",
        c.function or "",
        c.phone or "",
    ] for c in rows]
    _add_sheet(wb, "Contacts",
               ["Name", "Email", "Company", "Function", "Phone"],
               data)


def _sheet_packages(wb, db: Session, pid: int):
    pkgs = (db.query(models.Package)
              .filter_by(project_id=pid)
              .order_by(models.Package.tag_number).all())
    contact_by_id = {
        c.id: c for c in db.query(models.Contact).filter_by(project_id=pid).all()
    }
    def _cn(cid):
        c = contact_by_id.get(cid) if cid else None
        return c.name if c else ""
    data = []
    for p in pkgs:
        linked = ", ".join(
            (contact_by_id[pc.contact_id].name if pc.contact_id in contact_by_id else "")
            for pc in (p.package_contacts or [])
        )
        data.append([
            p.tag_number or "",
            p.name or "",
            p.company or "",
            p.address or "",
            _cn(p.package_owner_id),
            _cn(p.account_manager_id),
            _cn(p.pmc_technical_reviewer_id),
            _cn(p.pmc_commercial_reviewer_id),
            _cn(p.client_technical_reviewer_id),
            _cn(p.client_commercial_reviewer_id),
            linked,
        ])
    _add_sheet(wb, "Packages",
               ["Tag", "Name", "Company", "Address",
                "Package Owner", "Account Manager",
                "PMC Technical", "PMC Commercial",
                "Client Technical", "Client Commercial",
                "Linked Contacts"],
               data)


def _sheet_areas(wb, db: Session, pid: int):
    areas = (db.query(models.Area).filter_by(project_id=pid)
                .order_by(models.Area.tag).all())
    data = []
    for a in areas:
        supervisors = ", ".join(
            (s.contact.name if s.contact else "")
            for s in (a.site_supervisors or [])
        )
        data.append([
            a.tag or "",
            a.description or "",
            a.area_manager.name if getattr(a, "area_manager", None) else "",
            supervisors,
        ])
    _add_sheet(wb, "Areas",
               ["Tag", "Description", "Area Manager", "Site Supervisors"],
               data)


def _sheet_meeting_points(wb, db: Session, pid: int):
    rows = (db.query(models.MeetingPoint)
              .filter_by(project_id=pid)
              .order_by(models.MeetingPoint.project_seq_id).all())
    data = []
    for p in rows:
        # First linked meeting (if any) — gives us the meeting-type context
        link = next(iter(p.meeting_links or []), None)
        meeting = link.meeting if link else None
        mt_name = meeting.meeting_type.name if (meeting and meeting.meeting_type) else ""
        data.append([
            f"P-{str(p.project_seq_id or p.id).zfill(6)}",
            p.type or "",
            p.topic or "",
            p.details or "",
            p.responsible.name if p.responsible else "",
            p.due_date or "",
            _status_label(p.status),
            p.source_module or "",
            mt_name,
            _ts(p.closed_at),
            _user_name(p.created_by),
            _ts(p.created_at),
        ])
    _add_sheet(wb, "Meeting Points",
               ["ID", "Type", "Topic", "Details", "Responsible",
                "Due Date", "Status", "Source", "Meeting Type",
                "Closed At", "Created By", "Created At"],
               data)


def _sheet_tasks(wb, db: Session, pid: int):
    # Reuse schedule's _fmt_task so progress + late flag exactly match what
    # the user sees in the Schedule module.
    from routers.schedule import _fmt_task as _schedule_fmt_task

    def _status(pct):
        if pct >= 100: return "Complete"
        if pct > 0:    return "In progress"
        return "Not started"

    rows = (db.query(models.Task).filter_by(project_id=pid)
              .order_by(models.Task.project_seq_id).all())
    data = []
    for t in rows:
        info = _schedule_fmt_task(t, db)
        pct = info.get("current_progress") or 0
        data.append([
            f"T-{str(t.project_seq_id or t.id).zfill(6)}",
            f"{t.package.tag_number} - {t.package.name}" if t.package else "",
            t.description or "",
            t.details or "",
            t.start_date or "",
            t.finish_date or "",
            t.financial_weight if t.financial_weight is not None else "",
            round(pct, 1),
            _status(pct),
            "Y" if info.get("is_late") else "N",
            f"{t.area.tag} — {t.area.description}" if t.area else "",
            f"{t.unit.tag} — {t.unit.description}" if t.unit else "",
        ])
    _add_sheet(wb, "Tasks",
               ["ID", "Package", "Description", "Details",
                "Start Date", "Finish Date", "Financial Weight (%)",
                "Progress (%)", "Status", "Late",
                "Area", "Unit"],
               data)


def _sheet_budget_overview(wb, user: auth.ProjectContext, db: Session):
    # Reuse the live overview computation so the bundle is consistent with the
    # standalone export (same totals, same columns).
    from routers.budget import get_budget_overview
    rows = get_budget_overview(user=user, db=db)
    data = []
    for r in rows:
        data.append([
            r.get("tag_number", ""),
            r.get("name", ""),
            r.get("currency", "EUR"),
            r.get("baseline", 0) or 0,
            r.get("forecast", 0) or 0,
            r.get("bid_value") if r.get("bid_value") is not None else "",
            r.get("committed", 0) or 0,
            r.get("remaining", 0) or 0,
            r.get("pending_sc_cost", 0) or 0,
            r.get("remaining_incl_pending", 0) or 0,
            r.get("spend", 0) or 0,
            (r.get("bid_status") or "").title(),
            r.get("awarded_company_name") or "",
        ])
    _add_sheet(wb, "Budget Overview",
               ["Package", "Package Name", "Currency", "Baseline",
                "Actual Budget", "Bid Value", "Committed",
                "Remaining", "Pending SC", "Remaining incl. pending SC",
                "Spend", "Bid Status", "Awarded Vendor"],
               data)


def _sheet_procurement_plan(wb, db: Session, pid: int):
    plans = (db.query(models.PackagePlan).filter_by(project_id=pid).all())
    data = []
    for plan in plans:
        pkg = plan.package
        bidders = ", ".join(
            (b.company.name if b.company else "")
            for b in (plan.bidders or [])
        )
        step_dates = "; ".join(
            f"{(sd.step.step_id if sd.step else '?')}: {sd.due_date or '—'}"
            for sd in (plan.step_dates or [])
            if sd.due_date
        )
        data.append([
            pkg.tag_number if pkg else "",
            pkg.name if pkg else "",
            plan.notes or "",
            bidders,
            step_dates,
        ])
    _add_sheet(wb, "Procurement Plan",
               ["Package Tag", "Package Name", "Notes",
                "Bidders", "Step Due Dates"],
               data)


def _sheet_procurement_register(wb, db: Session, pid: int):
    entries = (db.query(models.ProcurementEntry).filter_by(project_id=pid).all())
    data = []
    for e in entries:
        pkg = e.package
        company = e.company
        step = e.current_step
        data.append([
            pkg.tag_number if pkg else "",
            pkg.name if pkg else "",
            company.name if company else "",
            _status_label(e.status),
            step.step_id if step else "",
            e.bid_value if e.bid_value is not None else "",
            e.bid_currency or "",
            e.technical_compliance or "",
            e.commercial_compliance or "",
            e.exclusion_reason or "",
            _user_name(e.created_by),
            _ts(e.created_at),
        ])
    _add_sheet(wb, "Procurement Register",
               ["Package Tag", "Package Name", "Bidder",
                "Status", "Current Step", "Bid Value", "Currency",
                "Technical Compliance", "Commercial Compliance",
                "Exclusion Reason", "Created By", "Created At"],
               data)


def _sheet_scope_changes(wb, db: Session, pid: int):
    rows = (db.query(models.ScopeChange).filter_by(project_id=pid)
              .order_by(models.ScopeChange.project_seq_id).all())
    data = []
    for sc in rows:
        data.append([
            f"SC-{str(sc.project_seq_id or sc.id).zfill(6)}",
            sc.package.tag_number if sc.package else "",
            sc.description or "",
            sc.details or "",
            sc.cost if sc.cost is not None else "",
            sc.schedule_impact_months if sc.schedule_impact_months is not None else "",
            _status_label(sc.status),
            "Y" if sc.pmc_approved else ("N" if sc.pmc_reviewed else ""),
            "Y" if sc.client_approved else ("N" if sc.client_reviewed else ""),
            _user_name(sc.created_by),
            _ts(sc.created_at),
        ])
    _add_sheet(wb, "Scope Changes",
               ["ID", "Package", "Description", "Details",
                "Cost", "Schedule Impact (months)", "Status",
                "PMC Approved", "Client Approved",
                "Created By", "Created At"],
               data)


def _sheet_documents(wb, db: Session, pid: int):
    rows = (db.query(models.Document).filter_by(project_id=pid)
              .order_by(models.Document.project_seq_id).all())
    data = []
    for d in rows:
        data.append([
            f"D-{str(d.project_seq_id or d.id).zfill(6)}",
            d.package.tag_number if d.package else "",
            d.subservice.description if getattr(d, "subservice", None) else "",
            d.document_type or "",
            d.description or "",
            _status_label(d.status),
            d.current_version if d.current_version is not None else 0,
            d.weight if d.weight is not None else "",
            d.start_date or "",
            d.first_issue_date or "",
            d.approval_due_date or "",
            d.actual_start_date or "",
            _user_name(d.created_by),
            _ts(d.created_at),
        ])
    _add_sheet(wb, "Documents",
               ["ID", "Package", "Subservice", "Type", "Description",
                "Status", "Current Version", "Weight",
                "Start Date", "First Issue Date", "Approval Due Date",
                "Actual Start Date", "Created By", "Created At"],
               data)


def _sheet_itp(wb, db: Session, pid: int):
    rows = (db.query(models.ITPRecord).filter_by(project_id=pid)
              .order_by(models.ITPRecord.project_seq_id).all())
    data = []
    for r in rows:
        data.append([
            f"ITP-{str(r.project_seq_id or r.id).zfill(6)}",
            r.package.tag_number if r.package else "",
            r.test or "",
            r.details or "",
            r.test_type.name if r.test_type else "",
            r.witness_level.name if getattr(r, "witness_level", None) else "",
            _status_label(r.status),
            r.area.tag if r.area else "",
            r.unit.tag if r.unit else "",
            r.planned_date or "",
            "Y" if r.pmc_approved else ("N" if r.pmc_reviewed else ""),
            "Y" if r.client_approved else ("N" if r.client_reviewed else ""),
            _user_name(r.created_by),
            _ts(r.created_at),
        ])
    _add_sheet(wb, "ITP Register",
               ["ID", "Package", "Test", "Details", "Test Type",
                "Witness Level", "Status", "Area", "Unit", "Planned Date",
                "PMC Approved", "Client Approved",
                "Created By", "Created At"],
               data)


def _sheet_punch(wb, db: Session, pid: int):
    rows = (db.query(models.PunchItem).filter_by(project_id=pid)
              .order_by(models.PunchItem.project_seq_id).all())
    data = []
    for p in rows:
        data.append([
            f"PCH-{str(p.project_seq_id or p.id).zfill(6)}",
            p.package.tag_number if p.package else "",
            p.topic or "",
            p.details or "",
            _status_label(p.status),
            p.obligation_time.name if getattr(p, "obligation_time", None) else "",
            p.area.tag if p.area else "",
            p.unit.tag if p.unit else "",
            _user_name(p.created_by),
            _ts(p.created_at),
        ])
    _add_sheet(wb, "Punch List",
               ["ID", "Package", "Topic", "Details", "Status",
                "Obligation Time", "Area", "Unit",
                "Created By", "Created At"],
               data)


def _sheet_daily_reports(wb, db: Session, pid: int):
    rows = (db.query(models.DailyReport).filter_by(project_id=pid)
              .order_by(models.DailyReport.report_date.desc(),
                        models.DailyReport.id.desc()).all())
    data = []
    for r in rows:
        workers = list(r.workers) if r.workers else []
        worker_names = ", ".join(drw.worker.name for drw in workers if drw.worker)
        worker_count = len(workers)
        total_hours = 0 if r.no_work else round((r.avg_hours_per_worker or 0) * worker_count, 2)
        areas = ", ".join(dra.area.tag for dra in (r.areas or []) if dra.area)
        data.append([
            r.package.tag_number if r.package else "",
            r.report_date or "",
            "Y" if r.no_work else "N",
            r.avg_hours_per_worker or 0,
            worker_count,
            total_hours,
            worker_names,
            areas,
            r.description or "",
            _user_name(r.created_by),
            _ts(r.created_at),
        ])
    _add_sheet(wb, "Daily Reports",
               ["Package", "Date", "No Work", "Avg Hours/Worker",
                "Workers (count)", "Total Hours", "Workers", "Areas",
                "Description", "Created By", "Created At"],
               data)


def _sheet_workers(wb, db: Session, pid: int):
    rows = (db.query(models.Worker).filter_by(project_id=pid)
              .order_by(models.Worker.name).all())
    data = []
    for w in rows:
        cert_names = ", ".join(
            (wc.certificate_type.name if getattr(wc, "certificate_type", None) else "")
            for wc in (w.certificates or [])
        )
        data.append([
            w.name or "",
            w.phone or "",
            w.package.tag_number if w.package else "",
            "Y" if w.is_subcontractor else "N",
            w.subcontractor.company if getattr(w, "subcontractor", None) else "",
            _status_label(w.status),
            cert_names,
            _ts(w.submitted_at),
            _ts(w.reviewed_at),
            _user_name(w.reviewed_by),
            w.rejection_comment or "",
            _user_name(w.created_by),
            _ts(w.created_at),
        ])
    _add_sheet(wb, "Workers",
               ["Name", "Phone", "Package",
                "Subcontractor?", "Subcontractor Company",
                "Status", "Certificates",
                "Submitted At", "Reviewed At", "Reviewed By",
                "Rejection Comment", "Created By", "Created At"],
               data)


def _sheet_subcontractors(wb, db: Session, pid: int):
    rows = (db.query(models.Subcontractor).filter_by(project_id=pid).all())
    data = []
    for s in rows:
        data.append([
            getattr(s, "company", "") or "",
            getattr(s, "contact_name", "") or "",
            getattr(s, "phone", "") or "",
            getattr(s, "email", "") or "",
        ])
    _add_sheet(wb, "Subcontractors",
               ["Company", "Contact", "Phone", "Email"],
               data)


def _sheet_work_permits(wb, db: Session, pid: int):
    rows = (db.query(models.WorkPermit).filter_by(project_id=pid)
              .order_by(models.WorkPermit.project_seq_id).all())
    data = []
    for wp in rows:
        types = ", ".join(
            (pt.permit_type.name if getattr(pt, "permit_type", None) else "")
            for pt in (wp.permit_types or [])
        )
        areas = ", ".join(
            (wpa.area.tag if wpa.area else "")
            for wpa in (wp.areas or [])
        )
        hazards = ", ".join((h.hazard_key or "") for h in (wp.hazards or []))
        ppe = ", ".join((p.ppe_key or "") for p in (wp.ppes or []))
        data.append([
            f"WP-{str(wp.project_seq_id or wp.id).zfill(6)}",
            wp.package.tag_number if wp.package else "",
            wp.title or "",
            wp.description or "",
            types,
            areas,
            hazards,
            ppe,
            wp.start_date or "",
            wp.end_date or "",
            _status_label(wp.status),
            _ts(wp.submitted_at),
            _user_name(wp.created_by),
            _ts(wp.created_at),
        ])
    _add_sheet(wb, "Work Permits",
               ["ID", "Package", "Title", "Description",
                "Permit Types", "Areas", "Hazards", "PPE",
                "Start Date", "End Date", "Status",
                "Submitted At", "Created By", "Created At"],
               data)


def _sheet_lotos(wb, db: Session, pid: int):
    rows = (db.query(models.LOTO).filter_by(project_id=pid)
              .order_by(models.LOTO.project_seq_id).all())
    data = []
    for lt in rows:
        wp = lt.work_permit
        data.append([
            f"L-{str(lt.project_seq_id or lt.id).zfill(6)}",
            f"WP-{str(wp.project_seq_id or wp.id).zfill(6)}" if wp else "",
            lt.tag_number or "",
            lt.description or "",
            _status_label(lt.status),
            "Y" if lt.locked_state else "N",
            _ts(lt.submitted_at),
            _ts(lt.reviewed_at),
            _user_name(lt.reviewed_by),
            lt.refusal_comment or "",
            _user_name(lt.created_by),
            _ts(lt.created_at),
        ])
    _add_sheet(wb, "LOTO",
               ["ID", "Work Permit", "Tag", "Description",
                "Status", "Locked",
                "Submitted At", "Reviewed At", "Reviewed By",
                "Refusal Comment", "Created By", "Created At"],
               data)


def _sheet_safety_observations(wb, db: Session, pid: int):
    rows = (db.query(models.SafetyObservation).filter_by(project_id=pid)
              .order_by(models.SafetyObservation.submitted_at.desc().nullslast(),
                        models.SafetyObservation.id.desc()).all())
    data = []
    for r in rows:
        cat = r.category
        data.append([
            f"SO-{str(r.project_seq_id or r.id).zfill(6)}",
            _status_label(r.status),
            cat.name if cat else "",
            cat.polarity if cat else "",
            r.package.tag_number if r.package else "",
            r.area.tag if r.area else "",
            getattr(getattr(r, "subcontractor", None), "name", "") or
                getattr(getattr(r, "subcontractor", None), "company", "") or "",
            r.worker.name if r.worker else "",
            r.details or "",
            r.remediation_request or "",
            _ts(r.submitted_at),
            _ts(r.acknowledged_at),
            _user_name(r.acknowledged_by),
            r.acknowledge_comment or "",
            _ts(r.closed_at),
            _user_name(r.closed_by),
            _user_name(r.created_by),
            _ts(r.created_at),
        ])
    _add_sheet(wb, "Safety Observations",
               ["Seq #", "Status", "Category", "Polarity",
                "Package", "Area", "Subcontractor", "Worker",
                "Details", "Remediation Request",
                "Submitted At", "Acknowledged At", "Acknowledged By",
                "Acknowledge Comment", "Closed At", "Closed By",
                "Created By", "Created At"],
               data)


def _sheet_safety_incidents(wb, db: Session, pid: int):
    rows = (db.query(models.SafetyIncident).filter_by(project_id=pid)
              .order_by(models.SafetyIncident.incident_date.desc().nullslast(),
                        models.SafetyIncident.id.desc()).all())
    data = []
    for r in rows:
        sev = r.severity_class
        cause = r.incident_cause
        worker_names = ", ".join(
            (siw.worker.name if siw.worker else "")
            for siw in (r.workers or [])
        )
        data.append([
            f"SI-{str(r.project_seq_id or r.id).zfill(6)}",
            _status_label(r.status),
            sev.level if sev else "",
            sev.name if sev else "",
            cause.name if cause else "",
            r.other_cause_text or "",
            r.package.tag_number if r.package else "",
            r.area.tag if r.area else "",
            r.incident_date or "",
            worker_names,
            r.details or "",
            r.action or "",
            _ts(r.submitted_at),
            _user_name(r.submitted_by),
            _ts(r.investigated_at),
            _user_name(r.investigated_by),
            r.investigation_comment or "",
            _ts(r.action_completed_at),
            _user_name(r.action_completed_by),
            r.action_completion_comment or "",
            _ts(r.closed_at),
            _user_name(r.closed_by),
            _user_name(r.created_by),
            _ts(r.created_at),
        ])
    _add_sheet(wb, "Safety Incidents",
               ["Seq #", "Status", "Severity Level", "Severity Class",
                "Cause", "Other Cause", "Package", "Area", "Incident Date",
                "Workers Involved", "Details", "Action",
                "Submitted At", "Submitted By",
                "Investigated At", "Investigated By", "Investigation Comment",
                "Action Completed At", "Action Completed By",
                "Completion Comment",
                "Closed At", "Closed By", "Created By", "Created At"],
               data)


def _sheet_safety_toolboxes(wb, db: Session, pid: int):
    rows = (db.query(models.SafetyToolbox).filter_by(project_id=pid)
              .order_by(models.SafetyToolbox.talk_date.desc().nullslast(),
                        models.SafetyToolbox.id.desc()).all())
    data = []
    for r in rows:
        given_by = (r.given_by_user.name if r.given_by_user else
                    r.given_by_worker.name if r.given_by_worker else "")
        packages = ", ".join(
            (tp.package.tag_number if tp.package else "") for tp in (r.packages or [])
        )
        attendees = list(r.workers or [])
        attendee_names = ", ".join(
            (tw.worker.name if tw.worker else "") for tw in attendees
        )
        obs_seqs = ", ".join(
            f"#{(to.observation.project_seq_id or to.observation.id)}"
            for to in (r.observations or []) if to.observation
        )
        inc_seqs = ", ".join(
            f"#{(ti.incident.project_seq_id or ti.incident.id)}"
            for ti in (r.incidents or []) if ti.incident
        )
        data.append([
            f"TBX-{str(r.project_seq_id or r.id).zfill(6)}",
            _status_label(r.status),
            r.category.name if r.category else "",
            r.other_category_text or "",
            r.talk_date or "",
            given_by,
            packages,
            len(attendees),
            attendee_names,
            obs_seqs,
            inc_seqs,
            r.details or "",
            _ts(r.submitted_at),
            _user_name(r.submitted_by),
            _ts(r.acknowledged_at),
            _user_name(r.acknowledged_by),
            r.acknowledge_comment or "",
            _user_name(r.created_by),
            _ts(r.created_at),
        ])
    _add_sheet(wb, "Toolbox Talks",
               ["Seq #", "Status", "Category", "Other Category",
                "Talk Date", "Given By", "Packages",
                "Attendees (count)", "Attendees",
                "Linked Observations", "Linked Incidents",
                "Details",
                "Submitted At", "Submitted By",
                "Acknowledged At", "Acknowledged By", "Acknowledge Comment",
                "Created By", "Created At"],
               data)


def _sheet_files(wb, user: auth.ProjectContext, db: Session):
    # Reuse the master-list union (file attachments + floorplans + ready
    # background reports) so the bundle matches the Files module screen.
    from routers.attachments import list_all_attachments
    rows = list_all_attachments(record_type=None, db=db, user=user)
    data = []
    for r in rows:
        uploaded_at = r.get("uploaded_at") or ""
        if isinstance(uploaded_at, str) and "T" in uploaded_at:
            uploaded_at = uploaded_at.replace("T", " ").rstrip("Z")
        data.append([
            r.get("original_filename") or "",
            r.get("record_type_label") or r.get("record_type") or "",
            (r.get("source") or "").title(),
            r.get("record_ref") or "",
            r.get("record_id") or "",
            r.get("stored_path") or "",
            r.get("file_size") or 0,
            r.get("content_type") or "",
            r.get("step_name") or "",
            uploaded_at,
            r.get("uploaded_by_name") or "",
            r.get("uploaded_by_role") or "",
        ])
    _add_sheet(wb, "Files",
               ["File", "Type", "Source", "Linked Record", "Linked Record ID",
                "Path", "Size (bytes)", "Content Type", "Step",
                "Uploaded At", "Uploaded By", "Uploaded By Role"],
               data)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/export/excel")
def export_full_project_database(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    """ADMIN / PROJECT_OWNER only — bundle the entire project's data into a
    single multi-sheet xlsx workbook."""
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Project owner or admin access required")

    from openpyxl import Workbook

    pid = user.project_id
    proj = db.query(models.Project).filter_by(id=pid).first()
    if not proj:
        raise HTTPException(404, "Project not found")

    wb = Workbook()
    # Strip the default empty 'Sheet' so the bundle starts on the first
    # real domain sheet.
    default = wb.active
    wb.remove(default)

    # Sheets — order intentional (organisation → execution → safety/files)
    builders = [
        ("contacts",              lambda: _sheet_contacts(wb, db, pid)),
        ("packages",              lambda: _sheet_packages(wb, db, pid)),
        ("areas",                 lambda: _sheet_areas(wb, db, pid)),
        ("meeting points",        lambda: _sheet_meeting_points(wb, db, pid)),
        ("tasks",                 lambda: _sheet_tasks(wb, db, pid)),
        ("budget overview",       lambda: _sheet_budget_overview(wb, user, db)),
        ("procurement plan",      lambda: _sheet_procurement_plan(wb, db, pid)),
        ("procurement register",  lambda: _sheet_procurement_register(wb, db, pid)),
        ("scope changes",         lambda: _sheet_scope_changes(wb, db, pid)),
        ("documents",             lambda: _sheet_documents(wb, db, pid)),
        ("itp",                   lambda: _sheet_itp(wb, db, pid)),
        ("punch",                 lambda: _sheet_punch(wb, db, pid)),
        ("daily reports",         lambda: _sheet_daily_reports(wb, db, pid)),
        ("workers",               lambda: _sheet_workers(wb, db, pid)),
        ("subcontractors",        lambda: _sheet_subcontractors(wb, db, pid)),
        ("work permits",          lambda: _sheet_work_permits(wb, db, pid)),
        ("loto",                  lambda: _sheet_lotos(wb, db, pid)),
        ("safety observations",   lambda: _sheet_safety_observations(wb, db, pid)),
        ("safety incidents",      lambda: _sheet_safety_incidents(wb, db, pid)),
        ("safety toolboxes",      lambda: _sheet_safety_toolboxes(wb, db, pid)),
        ("files",                 lambda: _sheet_files(wb, user, db)),
    ]
    for name, fn in builders:
        try:
            fn()
        except Exception as exc:
            # Leave a placeholder sheet so the user knows which section
            # failed without aborting the whole export.
            from openpyxl.styles import Font
            ws = wb.create_sheet(title=f"{name[:26]} ERR")
            ws.append([f"Failed to export '{name}': {exc!r}"])
            ws.cell(row=1, column=1).font = Font(bold=True, color="C00000")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_proj = "".join(
        ch if ch.isalnum() or ch in "-_" else "_" for ch in (proj.project_number or f"project_{pid}")
    )
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    filename = f"{safe_proj}_full_database_{date_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
