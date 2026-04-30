import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
import database, models, auth
from routers.audit import set_created, set_updated, check_lock, audit_dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(prefix="/api/risks", tags=["risks"])


def _get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _fmt_risk(r: models.Risk):
    return {
        "id": r.id,
        "seq_id": r.project_seq_id,
        "title": r.title,
        "description": r.description,
        "status": r.status,
        "category_id": r.category_id,
        "category_name": r.category.name if r.category else None,
        "phase_id": r.phase_id,
        "phase_name": r.phase.name if r.phase else None,
        "date_opened": r.date_opened,
        "date_closed": r.date_closed,
        "owner_id": r.owner_id,
        "owner_name": r.owner.name if r.owner else None,
        "prob_score_before": r.prob_score_before,
        "capex_score_before": r.capex_score_before,
        "schedule_score_before": r.schedule_score_before,
        "capex_value": r.capex_value,
        "schedule_value": r.schedule_value,
        "mitigation_type": r.mitigation_type,
        "mitigation_action": r.mitigation_action,
        "action_due_date": r.action_due_date,
        "action_status": r.action_status,
        "prob_score_after": r.prob_score_after,
        "capex_score_after": r.capex_score_after,
        "schedule_score_after": r.schedule_score_after,
        "secondary_effects": r.secondary_effects,
        **audit_dict(r),
        "notes": [
            {
                "id": n.id,
                "content": n.content,
                "created_at": n.created_at.isoformat() + 'Z' if n.created_at else None,
                "author_name": n.author.name if n.author else None,
            }
            for n in (r.notes or [])
        ],
    }


# ── Score Setup (global — not project-scoped) ─────────────────────────────────

class ScoreSetupBody(BaseModel):
    probability_pct: float = 0.0
    capex_impact_pct: float = 0.0
    schedule_impact_pct: float = 0.0


@router.get("/score-setup")
def get_score_setup(db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    rows = db.query(models.RiskScoreSetup).order_by(models.RiskScoreSetup.score).all()
    return [
        {
            "score": r.score,
            "probability_pct": r.probability_pct,
            "capex_impact_pct": r.capex_impact_pct,
            "schedule_impact_pct": r.schedule_impact_pct,
        }
        for r in rows
    ]


@router.put("/score-setup/{score}")
def update_score_setup(
    score: int,
    body: ScoreSetupBody,
    db: Session = Depends(_get_db),
    user=Depends(auth.get_project_user),
):
    if not auth.has_owner_or_lead_access(user, "Risk Register", db):
        raise HTTPException(403, "Not authorized")
    if score < 1 or score > 5:
        raise HTTPException(400, "Score must be 1-5")
    row = db.query(models.RiskScoreSetup).filter_by(score=score).first()
    if not row:
        raise HTTPException(404, "Score not found")
    row.probability_pct = body.probability_pct
    row.capex_impact_pct = body.capex_impact_pct
    row.schedule_impact_pct = body.schedule_impact_pct
    db.commit()
    return {"score": row.score, "probability_pct": row.probability_pct,
            "capex_impact_pct": row.capex_impact_pct, "schedule_impact_pct": row.schedule_impact_pct}


# ── Risk Matrix (global) ───────────────────────────────────────────────────────

class MatrixCellBody(BaseModel):
    level: str


@router.get("/matrix")
def get_matrix(db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    cells = db.query(models.RiskMatrixCell).all()
    return [{"prob_score": c.prob_score, "impact_score": c.impact_score, "level": c.level} for c in cells]


@router.put("/matrix/{prob_score}/{impact_score}")
def update_matrix_cell(
    prob_score: int,
    impact_score: int,
    body: MatrixCellBody,
    db: Session = Depends(_get_db),
    user=Depends(auth.get_project_user),
):
    if not auth.has_owner_or_lead_access(user, "Risk Register", db):
        raise HTTPException(403, "Not authorized")
    if body.level not in ("LOW", "MEDIUM", "HIGH"):
        raise HTTPException(400, "Level must be LOW, MEDIUM, or HIGH")
    cell = db.query(models.RiskMatrixCell).filter_by(
        prob_score=prob_score, impact_score=impact_score
    ).first()
    if not cell:
        raise HTTPException(404, "Cell not found")
    cell.level = body.level
    db.commit()
    return {"prob_score": cell.prob_score, "impact_score": cell.impact_score, "level": cell.level}


# ── Categories (per project) ───────────────────────────────────────────────────

class CategoryBody(BaseModel):
    name: str
    description: Optional[str] = None
    sort_order: Optional[int] = 0


@router.get("/categories")
def get_categories(db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    cats = db.query(models.RiskCategory).filter(
        models.RiskCategory.project_id == user.project_id
    ).order_by(models.RiskCategory.sort_order, models.RiskCategory.name).all()
    return [{"id": c.id, "name": c.name, "description": c.description, "sort_order": c.sort_order} for c in cats]


@router.post("/categories")
def create_category(body: CategoryBody, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if not auth.has_owner_or_lead_access(user, "Risk Register", db):
        raise HTTPException(403, "Not authorized")
    c = models.RiskCategory(
        project_id=user.project_id,
        name=body.name, description=body.description, sort_order=body.sort_order or 0
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return {"id": c.id, "name": c.name, "description": c.description, "sort_order": c.sort_order}


@router.put("/categories/{cat_id}")
def update_category(cat_id: int, body: CategoryBody, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if not auth.has_owner_or_lead_access(user, "Risk Register", db):
        raise HTTPException(403, "Not authorized")
    c = db.query(models.RiskCategory).filter_by(id=cat_id, project_id=user.project_id).first()
    if not c:
        raise HTTPException(404, "Category not found")
    c.name = body.name
    c.description = body.description
    c.sort_order = body.sort_order or 0
    db.commit()
    return {"id": c.id, "name": c.name, "description": c.description, "sort_order": c.sort_order}


@router.delete("/categories/{cat_id}")
def delete_category(cat_id: int, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if not auth.has_owner_or_lead_access(user, "Risk Register", db):
        raise HTTPException(403, "Not authorized")
    c = db.query(models.RiskCategory).filter_by(id=cat_id, project_id=user.project_id).first()
    if not c:
        raise HTTPException(404, "Not found")
    db.delete(c)
    db.commit()
    return {"ok": True}


# ── Phases (per project) ───────────────────────────────────────────────────────

class PhaseBody(BaseModel):
    name: str
    sort_order: Optional[int] = 0


@router.get("/phases")
def get_phases(db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    phases = db.query(models.RiskPhase).filter(
        models.RiskPhase.project_id == user.project_id
    ).order_by(models.RiskPhase.sort_order, models.RiskPhase.name).all()
    return [{"id": p.id, "name": p.name, "sort_order": p.sort_order} for p in phases]


@router.post("/phases")
def create_phase(body: PhaseBody, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if not auth.has_owner_or_lead_access(user, "Risk Register", db):
        raise HTTPException(403, "Not authorized")
    p = models.RiskPhase(project_id=user.project_id, name=body.name, sort_order=body.sort_order or 0)
    db.add(p)
    db.commit()
    db.refresh(p)
    return {"id": p.id, "name": p.name, "sort_order": p.sort_order}


@router.put("/phases/{phase_id}")
def update_phase(phase_id: int, body: PhaseBody, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if not auth.has_owner_or_lead_access(user, "Risk Register", db):
        raise HTTPException(403, "Not authorized")
    p = db.query(models.RiskPhase).filter_by(id=phase_id, project_id=user.project_id).first()
    if not p:
        raise HTTPException(404, "Phase not found")
    p.name = body.name
    p.sort_order = body.sort_order or 0
    db.commit()
    return {"id": p.id, "name": p.name, "sort_order": p.sort_order}


@router.delete("/phases/{phase_id}")
def delete_phase(phase_id: int, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if not auth.has_owner_or_lead_access(user, "Risk Register", db):
        raise HTTPException(403, "Not authorized")
    p = db.query(models.RiskPhase).filter_by(id=phase_id, project_id=user.project_id).first()
    if not p:
        raise HTTPException(404, "Not found")
    db.delete(p)
    db.commit()
    return {"ok": True}


# ── Risks (per project) ────────────────────────────────────────────────────────

class RiskBody(BaseModel):
    title: str
    description: Optional[str] = None
    status: Optional[str] = "OPEN"
    category_id: Optional[int] = None
    phase_id: Optional[int] = None
    date_opened: Optional[str] = None
    date_closed: Optional[str] = None
    owner_id: Optional[int] = None
    prob_score_before: Optional[int] = None
    capex_score_before: Optional[int] = None
    schedule_score_before: Optional[int] = None
    capex_value: Optional[float] = None
    schedule_value: Optional[float] = None
    mitigation_type: Optional[str] = None
    mitigation_action: Optional[str] = None
    action_due_date: Optional[str] = None
    action_status: Optional[str] = "NOT_STARTED"
    prob_score_after: Optional[int] = None
    capex_score_after: Optional[int] = None
    schedule_score_after: Optional[int] = None
    secondary_effects: Optional[str] = None
    updated_at: Optional[str] = None


@router.get("/my-open")
def get_my_open_risks(db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if not user.contact_id:
        return []
    risks = (
        db.query(models.Risk)
        .filter(
            models.Risk.project_id == user.project_id,
            models.Risk.owner_id == user.contact_id,
            models.Risk.status == "OPEN",
        )
        .order_by(models.Risk.created_at.desc())
        .all()
    )
    return [_fmt_risk(r) for r in risks]


@router.get("")
def get_risks(
    status: Optional[str] = None,
    category_id: Optional[int] = None,
    phase_id: Optional[int] = None,
    db: Session = Depends(_get_db),
    user=Depends(auth.get_project_user),
):
    q = db.query(models.Risk).filter(models.Risk.project_id == user.project_id)
    if status:
        q = q.filter(models.Risk.status == status)
    if category_id:
        q = q.filter(models.Risk.category_id == category_id)
    if phase_id:
        q = q.filter(models.Risk.phase_id == phase_id)
    risks = q.order_by(models.Risk.created_at.desc()).all()
    return [_fmt_risk(r) for r in risks]


@router.post("")
def create_risk(body: RiskBody, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if user.role == "VENDOR":
        raise HTTPException(403, "Vendors cannot create risks")
    r = models.Risk(
        project_id=user.project_id,
        project_seq_id=models.next_project_seq(db, models.Risk, user.project_id),
        **body.model_dump(exclude={"updated_at"}),
    )
    set_created(r, user.id)
    db.add(r)
    db.commit()
    db.refresh(r)
    return _fmt_risk(r)


# ── Export Excel ─────────────────────────────────────────────────────────────
# NOTE: must be registered before /{risk_id} to avoid route conflict

def _calc_impact(risk, score_setup, use_after=False):
    capex_val = risk.capex_value or 0
    sched_val = risk.schedule_value or 0
    ps = risk.prob_score_after if use_after else risk.prob_score_before
    cs = risk.capex_score_after if use_after else risk.capex_score_before
    ss = risk.schedule_score_after if use_after else risk.schedule_score_before
    capex_impact = 0.0
    sched_impact = 0.0
    if ps and capex_val:
        p = score_setup.get(ps)
        c = score_setup.get(cs or 1)
        if p and c:
            capex_impact = capex_val * (p.probability_pct / 100) * (c.capex_impact_pct / 100)
    if ps and sched_val:
        p = score_setup.get(ps)
        s = score_setup.get(ss or 1)
        if p and s:
            sched_impact = sched_val * (p.probability_pct / 100) * (s.schedule_impact_pct / 100)
    return capex_impact, sched_impact


@router.get("/export/excel")
def export_risks_excel(
    status: Optional[str] = None,
    category_id: Optional[int] = None,
    phase_id: Optional[int] = None,
    db: Session = Depends(_get_db),
    user=Depends(auth.get_project_user),
):
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(500, "openpyxl is not installed")

    q = db.query(models.Risk).filter(models.Risk.project_id == user.project_id)
    if status:
        q = q.filter(models.Risk.status == status)
    if category_id:
        q = q.filter(models.Risk.category_id == category_id)
    if phase_id:
        q = q.filter(models.Risk.phase_id == phase_id)
    risks = q.order_by(models.Risk.created_at.desc()).all()

    setup_rows = db.query(models.RiskScoreSetup).all()
    score_setup = {s.score: s for s in setup_rows}

    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    headers = [
        "ID", "Title", "Description", "Status",
        "Category", "Phase",
        "Date Opened", "Date Closed", "Owner",
        "Prob Score (Before)", "CAPEX Score (Before)", "Schedule Score (Before)",
        "Risk Score (Before)",
        "CAPEX at Risk", "Schedule at Risk (months)",
        "Budget Impact (Before)", "Budget Impact (After)",
        "Schedule Impact (Before)", "Schedule Impact (After)",
        "Mitigation Type", "Mitigation Action", "Action Due Date", "Action Status",
        "Prob Score (After)", "CAPEX Score (After)", "Schedule Score (After)",
        "Risk Score (After)",
        "Secondary Effects",
    ]
    ws.append(headers)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill

    for r in risks:
        before_capex, before_sched = _calc_impact(r, score_setup, use_after=False)
        after_capex, after_sched = _calc_impact(r, score_setup, use_after=True)

        risk_score_before = None
        if r.prob_score_before:
            impact_before = max(r.capex_score_before or 0, r.schedule_score_before or 0)
            risk_score_before = r.prob_score_before * impact_before if impact_before else None

        risk_score_after = None
        if r.prob_score_after:
            impact_after = max(r.capex_score_after or 0, r.schedule_score_after or 0)
            risk_score_after = r.prob_score_after * impact_after if impact_after else None

        ws.append([
            f"RI-{(r.project_seq_id or r.id):06d}",
            r.title,
            r.description,
            r.status,
            r.category.name if r.category else "",
            r.phase.name if r.phase else "",
            r.date_opened or "",
            r.date_closed or "",
            r.owner.name if r.owner else "",
            r.prob_score_before,
            r.capex_score_before,
            r.schedule_score_before,
            risk_score_before,
            r.capex_value or 0,
            r.schedule_value or 0,
            round(before_capex, 2) if before_capex else 0,
            round(after_capex, 2) if after_capex else 0,
            round(before_sched, 2) if before_sched else 0,
            round(after_sched, 2) if after_sched else 0,
            r.mitigation_type or "",
            r.mitigation_action or "",
            r.action_due_date or "",
            r.action_status or "",
            r.prob_score_after,
            r.capex_score_after,
            r.schedule_score_after,
            risk_score_after,
            r.secondary_effects or "",
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="risk_register.xlsx"'},
    )


@router.get("/{risk_id}")
def get_risk(risk_id: int, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    r = db.query(models.Risk).filter_by(id=risk_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Risk not found")
    return _fmt_risk(r)


@router.put("/{risk_id}")
def update_risk(risk_id: int, body: RiskBody, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if user.role == "VENDOR":
        raise HTTPException(403, "Vendors cannot edit risks")
    r = db.query(models.Risk).filter_by(id=risk_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Risk not found")
    check_lock(r.updated_at, body.updated_at, "risk")
    for k, v in body.model_dump(exclude_unset=True, exclude={"updated_at"}).items():
        setattr(r, k, v)
    set_updated(r, user.id)
    db.commit()
    db.refresh(r)
    return _fmt_risk(r)


@router.delete("/{risk_id}")
def delete_risk(risk_id: int, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    if not auth.has_owner_or_lead_access(user, "Risk Register", db):
        raise HTTPException(403, "Not authorized")
    r = db.query(models.Risk).filter_by(id=risk_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Not found")
    db.delete(r)
    db.commit()
    return {"ok": True}


# ── Risk Notes ────────────────────────────────────────────────────────────────

class NoteBody(BaseModel):
    content: str


@router.post("/{risk_id}/notes")
def add_risk_note(risk_id: int, body: NoteBody, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    r = db.query(models.Risk).filter_by(id=risk_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Risk not found")
    note = models.RiskNote(risk_id=risk_id, content=body.content, created_by_id=user.id)
    db.add(note)
    db.commit()
    db.refresh(note)
    return {
        "id": note.id,
        "content": note.content,
        "created_at": note.created_at.isoformat() + 'Z' if note.created_at else None,
        "author_name": note.author.name if note.author else None,
    }


@router.delete("/{risk_id}/notes/{note_id}")
def delete_risk_note(risk_id: int, note_id: int, db: Session = Depends(_get_db), user=Depends(auth.get_project_user)):
    note = db.query(models.RiskNote).filter_by(id=note_id, risk_id=risk_id).first()
    if not note:
        raise HTTPException(404, "Note not found")
    if not auth.has_owner_or_lead_access(user, "Risk Register", db) and note.created_by_id != user.id:
        raise HTTPException(403, "Not authorized")
    db.delete(note)
    db.commit()
    return {"ok": True}
