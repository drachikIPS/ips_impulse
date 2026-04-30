import io
import os
import re
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

import auth
import models
from database import get_db

router = APIRouter(prefix="/api/attachments", tags=["attachments"])

UPLOAD_ROOT = Path("uploads")
DEFAULT_MAX_UPLOAD_MB = 100
MAX_UPLOAD_MB_HARD_CAP = 500


def _max_upload_bytes(db: Session, project_id: int) -> int:
    """Per-project upload limit, settable from Project Settings."""
    row = db.query(models.Setting).filter_by(project_id=project_id, key="max_upload_mb").first()
    if row and row.value:
        try:
            mb = int(str(row.value).strip())
            if 1 <= mb <= MAX_UPLOAD_MB_HARD_CAP:
                return mb * 1024 * 1024
        except ValueError:
            pass
    return DEFAULT_MAX_UPLOAD_MB * 1024 * 1024

VALID_RECORD_TYPES = {
    "meeting_point", "order", "invoice", "scope_change",
    "progress_report", "document", "document_version", "procurement_entry", "task", "itp", "punch",
    "worker", "safety_observation", "incident", "safety_toolbox",
}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _sanitize(name: str) -> str:
    """Replace filesystem-unsafe characters."""
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', str(name)).strip("_. ") or "unknown"


def _folder_for_record(record_type: str, record_id: int, project_number: str, db: Session) -> Path:
    """
    Return the absolute upload folder for a given record.
    Raises 404/400 when the record is missing or mandatory relations absent.
    """
    pn = _sanitize(project_number)

    if record_type == "meeting_point":
        point = db.query(models.MeetingPoint).filter_by(id=record_id).first()
        if not point:
            raise HTTPException(404, "Meeting point not found")
        link = (
            db.query(models.MeetingPointLink)
            .filter_by(meeting_point_id=record_id)
            .order_by(models.MeetingPointLink.created_at)
            .first()
        )
        if not link:
            raise HTTPException(
                400,
                "Meeting point must be linked to a meeting before adding attachments",
            )
        meeting = db.query(models.Meeting).filter_by(id=link.meeting_id).first()
        type_name = _sanitize(
            meeting.meeting_type.name
            if meeting and meeting.meeting_type
            else "Unassigned"
        )
        return UPLOAD_ROOT / pn / "Meetings" / type_name

    elif record_type == "order":
        order = db.query(models.Order).filter_by(id=record_id).first()
        if not order:
            raise HTTPException(404, "Order not found")
        if not order.package:
            raise HTTPException(400, "Order must be linked to a package")
        return UPLOAD_ROOT / pn / _sanitize(order.package.tag_number) / "Orders"

    elif record_type == "invoice":
        inv = db.query(models.Invoice).filter_by(id=record_id).first()
        if not inv:
            raise HTTPException(404, "Invoice not found")
        pkg = inv.package or (inv.order.package if inv.order else None)
        if not pkg:
            raise HTTPException(400, "Invoice must be linked to a package")
        return UPLOAD_ROOT / pn / _sanitize(pkg.tag_number) / "Invoices"

    elif record_type == "scope_change":
        sc = db.query(models.ScopeChange).filter_by(id=record_id).first()
        if not sc:
            raise HTTPException(404, "Scope change not found")
        if not sc.package:
            raise HTTPException(
                400, "Scope change must be linked to a package before adding attachments"
            )
        return UPLOAD_ROOT / pn / _sanitize(sc.package.tag_number) / "Scope Changes"

    elif record_type == "progress_report":
        pr = db.query(models.ProgressReport).filter_by(id=record_id).first()
        if not pr:
            raise HTTPException(404, "Progress report not found")
        if not pr.package:
            raise HTTPException(
                400,
                "Progress report must be linked to a package before adding attachments",
            )
        return UPLOAD_ROOT / pn / _sanitize(pr.package.tag_number) / "Progress Reports"

    elif record_type == "document":
        doc = db.query(models.Document).filter_by(id=record_id).first()
        if not doc:
            raise HTTPException(404, "Document not found")
        if not doc.package:
            raise HTTPException(400, "Document must be linked to a package")
        subfolder = (
            "Technical Documents" if doc.document_type == "TECHNICAL" else "Commercial Documents"
        )
        return UPLOAD_ROOT / pn / _sanitize(doc.package.tag_number) / subfolder

    elif record_type == "document_version":
        dv = db.query(models.DocumentVersion).filter_by(id=record_id).first()
        if not dv:
            raise HTTPException(404, "Document version not found")
        doc = db.query(models.Document).filter_by(id=dv.document_id).first()
        if not doc or not doc.package:
            raise HTTPException(400, "Document version must be linked to a package")
        subfolder = (
            "Technical Documents" if doc.document_type == "TECHNICAL" else "Commercial Documents"
        )
        version_folder = f"V{dv.version:02d}"
        return UPLOAD_ROOT / pn / _sanitize(doc.package.tag_number) / subfolder / version_folder

    elif record_type == "task":
        task = db.query(models.Task).filter_by(id=record_id).first()
        if not task:
            raise HTTPException(404, "Task not found")
        if not task.package:
            raise HTTPException(400, "Task must be linked to a package")
        return UPLOAD_ROOT / pn / _sanitize(task.package.tag_number) / "Tasks"

    elif record_type == "procurement_entry":
        entry = db.query(models.ProcurementEntry).filter_by(id=record_id).first()
        if not entry:
            raise HTTPException(404, "Procurement entry not found")
        if not entry.package or not entry.company:
            raise HTTPException(400, "Procurement entry must have a package and company")
        folder_name = f"{_sanitize(entry.package.tag_number)}-{_sanitize(entry.company.name)}"
        return UPLOAD_ROOT / pn / "Procurement" / folder_name

    elif record_type == "itp":
        itp = db.query(models.ITPRecord).filter_by(id=record_id).first()
        if not itp:
            raise HTTPException(404, "ITP record not found")
        if not itp.package:
            raise HTTPException(400, "ITP record must be linked to a package")
        return UPLOAD_ROOT / pn / _sanitize(itp.package.tag_number) / "ITP"

    elif record_type == "punch":
        punch = db.query(models.PunchItem).filter_by(id=record_id).first()
        if not punch:
            raise HTTPException(404, "Punch item not found")
        if not punch.package:
            raise HTTPException(400, "Punch item must be linked to a package")
        return UPLOAD_ROOT / pn / _sanitize(punch.package.tag_number) / "Punchlist"

    elif record_type == "worker":
        worker = db.query(models.Worker).filter_by(id=record_id).first()
        if not worker:
            raise HTTPException(404, "Worker not found")
        if not worker.package:
            raise HTTPException(400, "Worker must be linked to a package")
        return UPLOAD_ROOT / pn / _sanitize(worker.package.tag_number) / "Worker certificates"

    elif record_type == "safety_observation":
        obs = db.query(models.SafetyObservation).filter_by(id=record_id).first()
        if not obs:
            raise HTTPException(404, "Safety observation not found")
        if not obs.package:
            raise HTTPException(400, "Safety observation must be linked to a package")
        return UPLOAD_ROOT / pn / _sanitize(obs.package.tag_number) / "Safety Observations"

    elif record_type == "incident":
        inc = db.query(models.SafetyIncident).filter_by(id=record_id).first()
        if not inc:
            raise HTTPException(404, "Safety incident not found")
        if not inc.package:
            raise HTTPException(400, "Safety incident must be linked to a package")
        return UPLOAD_ROOT / pn / _sanitize(inc.package.tag_number) / "Incidents"

    elif record_type == "safety_toolbox":
        tbx = db.query(models.SafetyToolbox).filter_by(id=record_id).first()
        if not tbx:
            raise HTTPException(404, "Toolbox talk not found")
        # Toolbox talks can span multiple packages, so attachments are stored
        # in a project-level "Safety Toolboxes" folder rather than per-package.
        return UPLOAD_ROOT / pn / "Safety Toolboxes"

    else:
        raise HTTPException(400, f"Unknown record type: {record_type}")


def _fmt_attachment(a: models.FileAttachment, db: Session = None, project_id: int = None) -> dict:
    # Resolve step name (cheap when db provided — also used by the bidder
    # portal to group documents per procurement step)
    step_name = None
    if a.step_id and a.step is not None:
        step_name = a.step.step_id

    # Project-scoped role of the uploader, used by the bidder portal to split
    # attachments into "project documents" vs "my uploads".
    uploaded_by_role = None
    if a.uploaded_by:
        uploaded_by_role = a.uploaded_by.role
        if db is not None and project_id is not None:
            up = db.query(models.UserProject).filter_by(
                user_id=a.uploaded_by_id, project_id=project_id,
            ).first()
            if up and up.role:
                uploaded_by_role = up.role

    return {
        "id": a.id,
        "key": f"a-{a.id}",
        "source": "attachment",
        "project_id": a.project_id,
        "record_type": a.record_type,
        "record_id": a.record_id,
        "original_filename": a.original_filename,
        "stored_path": a.stored_path,
        "file_size": a.file_size,
        "content_type": a.content_type,
        "uploaded_at": a.uploaded_at.isoformat() + 'Z' if a.uploaded_at else None,
        "uploaded_by_name": a.uploaded_by.name if a.uploaded_by else None,
        "uploaded_by_role": uploaded_by_role,
        "step_id": a.step_id,
        "step_name": step_name,
        "view_url": f"/api/attachments/{a.id}/view",
        "download_url": f"/api/attachments/{a.id}/download",
    }


def _fmt_floorplan_as_file(fp: "models.Floorplan") -> dict:
    """Floorplans live in their own table (linked to Areas, used for pin
    overlays) but should still appear in the project Files list."""
    return {
        "id": fp.id,
        "key": f"fp-{fp.id}",
        "source": "floorplan",
        "project_id": fp.project_id,
        "record_type": "floorplan",
        "record_id": fp.id,
        "record_type_label": "Floorplan",
        "record_ref": fp.name,
        "original_filename": fp.original_filename or fp.name,
        "stored_path": fp.stored_path,
        "file_size": fp.file_size,
        "content_type": fp.content_type or "image/jpeg",
        "uploaded_at": fp.uploaded_at.isoformat() + 'Z' if fp.uploaded_at else None,
        "uploaded_by_name": fp.uploaded_by.name if fp.uploaded_by else None,
        "view_url": f"/api/floorplans/{fp.id}/image",
        "download_url": f"/api/floorplans/{fp.id}/image",
    }


def _fmt_report_as_file(r: "models.Report") -> dict:
    """Background-generated PDF reports (Safety / Punch List) — surfaced in
    the Files list once their status is READY."""
    kind_label = {"safety": "Safety Report", "punch": "Punch List Report"}.get(r.kind, (r.kind or "report").title() + " Report")
    fname = r.title or f"{kind_label} {r.id}"
    if not fname.lower().endswith(".pdf"):
        fname = f"{fname}.pdf"
    return {
        "id": r.id,
        "key": f"r-{r.id}",
        "source": "report",
        "project_id": r.project_id,
        "record_type": "report",
        "record_id": r.id,
        "record_type_label": kind_label,
        "record_ref": r.title or f"Report {r.id}",
        "original_filename": fname,
        "stored_path": r.stored_path,
        "file_size": r.file_size,
        "content_type": "application/pdf",
        "uploaded_at": (r.completed_at or r.requested_at).isoformat() + 'Z' if (r.completed_at or r.requested_at) else None,
        "uploaded_by_name": r.requested_by.name if r.requested_by else None,
        "view_url": None,  # Reports use their own download endpoint; no inline preview
        "download_url": f"/api/reports/{r.id}/download",
    }


_PREFIX_CODES = {
    "meeting_point":    ("MP", "meeting_points"),
    "scope_change":     ("SC", "scope_changes"),
    "document":         ("DO", "documents"),
    "task":             ("TA", "tasks"),
    "progress_report":  ("PR", None),
    "itp":              ("IT", "itp_records"),
    "punch":            ("PI", "punch_items"),
    "order":            ("OR", None),
    "invoice":          ("IN", None),
    "procurement_entry":("PE", None),
    "worker":           ("WK", "workers"),
    "subcontractor":    ("SU", "subcontractors"),
    "safety_observation": ("SO", "safety_observations"),
    "incident":           ("IR", "safety_incidents"),
    "safety_toolbox":     ("TB", "safety_toolboxes"),
}


def _id_prefix(record_type: str, record_id: int, db: Session) -> str:
    """Return the formatted ID prefix using per-project seq_id where available."""
    if record_type == "document_version":
        dv = db.query(models.DocumentVersion).filter_by(id=record_id).first()
        if dv:
            doc = db.query(models.Document).filter_by(id=dv.document_id).first()
            if doc:
                seq = doc.project_seq_id or doc.id
                return f"DO-{seq:06d}-V{dv.version:02d}"
        return f"DV-{record_id:06d}"
    entry = _PREFIX_CODES.get(record_type)
    if not entry:
        return f"{record_type[:2].upper()}-{record_id:06d}"
    code, table = entry
    # Try to fetch project_seq_id from the relevant model
    seq_id = None
    if table:
        try:
            row = db.execute(
                text(f"SELECT project_seq_id FROM {table} WHERE id = :rid"),
                {"rid": record_id}
            ).fetchone()
            if row and row[0] is not None:
                seq_id = row[0]
        except Exception:
            pass
    num = seq_id if seq_id is not None else record_id
    return f"{code}-{num:06d}"


def _record_type_label(rt: str) -> str:
    return {
        "meeting_point": "Meeting Point",
        "order": "Order",
        "invoice": "Invoice",
        "scope_change": "Scope Change",
        "progress_report": "Progress Report",
        "document": "Document",
        "procurement_entry": "Procurement",
        "task": "Task",
        "itp": "ITP Record",
        "punch": "Punch Item",
        "worker": "Worker",
        "subcontractor": "Subcontractor",
        "safety_observation": "Safety Observation",
        "incident": "Safety Incident",
        "safety_toolbox": "Toolbox Talk",
    }.get(rt, rt)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_attachment(
    record_type: str = Form(...),
    record_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    if record_type not in VALID_RECORD_TYPES:
        raise HTTPException(400, f"Invalid record type: {record_type}")

    # Read file content (enforce per-project size limit)
    max_bytes = _max_upload_bytes(db, user.project_id)
    content = await file.read()
    if len(content) > max_bytes:
        max_mb = max_bytes // (1024 * 1024)
        raise HTTPException(400, f"File exceeds the project's maximum upload size of {max_mb} MB")

    # Get project number for folder path
    project = db.query(models.Project).filter_by(id=user.project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    folder = _folder_for_record(record_type, record_id, project.project_number, db)
    folder.mkdir(parents=True, exist_ok=True)

    # Build a safe unique filename: {id_prefix}_{stem_truncated}_{uuid8}.{ext}
    raw_name = file.filename or "file"
    suffix   = Path(raw_name).suffix[:10]        # keep extension, max 10 chars
    stem     = _sanitize(Path(raw_name).stem) or "file"
    stem     = stem[:60]                         # truncate stem to leave room for prefix
    uid8     = uuid.uuid4().hex[:8]              # 8-char unique suffix
    prefix   = _id_prefix(record_type, record_id, db)
    stored_name = f"{prefix}_{stem}_{uid8}{suffix}"
    dest = folder / stored_name

    dest.write_bytes(content)

    # For procurement entries, auto-stamp the step the entry is currently on.
    # This gives the bidder portal a free "uploaded at step X" overview for
    # both project- and bidder-side uploads with no extra UI input.
    step_id_at_upload = None
    if record_type == "procurement_entry":
        entry = db.query(models.ProcurementEntry).filter_by(
            id=record_id, project_id=user.project_id,
        ).first()
        if entry and entry.current_step_id:
            step_id_at_upload = entry.current_step_id

        # BIDDERs are locked out of uploading once they've submitted at the
        # entry's current step — they have to wait for the step to advance.
        if user.role == "BIDDER" and entry and entry.current_step_id:
            existing = db.query(models.BidderSubmittal).filter_by(
                entry_id=entry.id, step_id=entry.current_step_id,
            ).first()
            if existing:
                raise HTTPException(
                    409,
                    "You have already submitted for the current step. Uploads "
                    "are locked until the project team advances the step."
                )

    attachment = models.FileAttachment(
        project_id=user.project_id,
        record_type=record_type,
        record_id=record_id,
        original_filename=file.filename or "file",
        stored_path=str(dest.relative_to(UPLOAD_ROOT)),
        file_size=len(content),
        content_type=file.content_type or "application/octet-stream",
        uploaded_at=datetime.utcnow(),
        uploaded_by_id=user.id,
        step_id=step_id_at_upload,
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    return _fmt_attachment(attachment, db=db, project_id=user.project_id)


@router.get("")
def list_attachments(
    record_type: str,
    record_id: int,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    attachments = (
        db.query(models.FileAttachment)
        .filter_by(project_id=user.project_id, record_type=record_type, record_id=record_id)
        .order_by(models.FileAttachment.uploaded_at)
        .all()
    )
    return [_fmt_attachment(a, db=db, project_id=user.project_id) for a in attachments]


@router.get("/all")
def list_all_attachments(
    record_type: Optional[str] = None,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    """Master list — every project file in one view: FileAttachment rows plus
    Floorplan images (separate table — used for area pin overlays) and ready
    background-generated PDF Reports (separate Report table)."""
    result = []

    # ── Standard attachments ────────────────────────────────────────────────
    if not record_type or record_type not in ("floorplan", "report"):
        q = db.query(models.FileAttachment).filter_by(project_id=user.project_id)
        if record_type:
            q = q.filter_by(record_type=record_type)
        for a in q.order_by(models.FileAttachment.uploaded_at.desc()).all():
            row = _fmt_attachment(a, db=db, project_id=user.project_id)
            row["record_type_label"] = _record_type_label(a.record_type)
            row["record_ref"] = _record_ref(a, db)
            result.append(row)

    # ── Floorplans (only when no filter, or filter targets floorplans) ──────
    if not record_type or record_type == "floorplan":
        for fp in db.query(models.Floorplan).filter_by(project_id=user.project_id).all():
            result.append(_fmt_floorplan_as_file(fp))

    # ── Generated reports — only those READY for download ───────────────────
    if not record_type or record_type == "report":
        ready = db.query(models.Report).filter_by(
            project_id=user.project_id, status="READY",
        ).all()
        for r in ready:
            result.append(_fmt_report_as_file(r))

    # Sort the unioned set by uploaded_at desc (None last)
    result.sort(key=lambda x: x.get("uploaded_at") or "", reverse=True)
    return result


@router.get("/all/export/excel")
def export_all_attachments_excel(
    record_type: Optional[str] = None,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    """Excel dump of every file in the master Files list.
    Same column set as the on-screen table, plus a few extra technical columns
    (source kind, content type, file size in bytes) for downstream workflows."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = list_all_attachments(record_type=record_type, db=db, user=user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Files"
    ws.append([
        "File", "Type", "Source",
        "Linked Record", "Linked Record ID",
        "Path", "Size (bytes)", "Content Type",
        "Step",
        "Uploaded At", "Uploaded By", "Uploaded By Role",
    ])
    fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    for r in rows:
        uploaded_at = r.get("uploaded_at") or ""
        # Strip the ISO 'T' marker for nicer display in Excel
        if isinstance(uploaded_at, str) and "T" in uploaded_at:
            uploaded_at = uploaded_at.replace("T", " ").rstrip("Z")
        ws.append([
            r.get("original_filename") or "",
            r.get("record_type_label") or r.get("record_type") or "",
            (r.get("source") or "").title(),
            r.get("record_ref") or "",
            r.get("record_id") or "",
            r.get("stored_path") or "",
            r.get("file_size") or 0,
            r.get("content_type") or "",
            r.get("step_name") or "",
            uploaded_at,
            r.get("uploaded_by_name") or "",
            r.get("uploaded_by_role") or "",
        ])

    # Auto-width (capped)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="files_{date_str}.xlsx"'},
    )


def _record_ref(a: models.FileAttachment, db: Session) -> str:
    """Short human-readable identifier for the linked record."""
    try:
        if a.record_type == "meeting_point":
            p = db.query(models.MeetingPoint).filter_by(id=a.record_id).first()
            return p.topic[:60] if p else f"#{a.record_id}"
        elif a.record_type == "order":
            o = db.query(models.Order).filter_by(id=a.record_id).first()
            return o.po_number if o else f"#{a.record_id}"
        elif a.record_type == "invoice":
            i = db.query(models.Invoice).filter_by(id=a.record_id).first()
            return i.invoice_number if i else f"#{a.record_id}"
        elif a.record_type == "scope_change":
            sc = db.query(models.ScopeChange).filter_by(id=a.record_id).first()
            return f"SC-{a.record_id:06d}" + (f" {sc.description[:40]}" if sc else "")
        elif a.record_type == "progress_report":
            pr = db.query(models.ProgressReport).filter_by(id=a.record_id).first()
            if pr and pr.package:
                return f"PR-{a.record_id:06d} {pr.package.tag_number}"
            return f"PR-{a.record_id:06d}"
        elif a.record_type == "document":
            doc = db.query(models.Document).filter_by(id=a.record_id).first()
            return f"DO-{a.record_id:06d}" + (f" {doc.description[:40]}" if doc else "")
        elif a.record_type == "procurement_entry":
            entry = db.query(models.ProcurementEntry).filter_by(id=a.record_id).first()
            if entry:
                pkg = entry.package.tag_number if entry.package else "?"
                co = entry.company.name if entry.company else "?"
                return f"{pkg} — {co}"
            return f"#{a.record_id}"
        elif a.record_type == "task":
            task = db.query(models.Task).filter_by(id=a.record_id).first()
            if task:
                return task.description[:60]
            return f"TASK-{a.record_id}"
        elif a.record_type == "itp":
            itp = db.query(models.ITPRecord).filter_by(id=a.record_id).first()
            if itp:
                pkg = itp.package.tag_number if itp.package else "?"
                return f"IT-{a.record_id:06d} {pkg} — {itp.test[:40] if itp.test else ''}"
            return f"IT-{a.record_id:06d}"
        elif a.record_type == "punch":
            punch = db.query(models.PunchItem).filter_by(id=a.record_id).first()
            if punch:
                pkg = punch.package.tag_number if punch.package else "?"
                return f"PI-{a.record_id:06d} {pkg} — {punch.topic[:40] if punch.topic else ''}"
            return f"PI-{a.record_id:06d}"
    except Exception:
        pass
    return f"#{a.record_id}"


@router.delete("/{attachment_id}")
def delete_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only Admins and Project Owners can delete attachments")

    a = db.query(models.FileAttachment).filter_by(
        id=attachment_id, project_id=user.project_id
    ).first()
    if not a:
        raise HTTPException(404, "Attachment not found")

    # Remove from disk
    disk_path = UPLOAD_ROOT / a.stored_path
    try:
        if disk_path.exists():
            disk_path.unlink()
    except Exception:
        pass  # Log but don't block DB deletion

    db.delete(a)
    db.commit()
    return {"deleted": True}


@router.get("/{attachment_id}/view")
def view_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    """Serve the file inline (for PDF/image preview)."""
    a = db.query(models.FileAttachment).filter_by(
        id=attachment_id, project_id=user.project_id
    ).first()
    if not a:
        raise HTTPException(404, "Attachment not found")

    disk_path = UPLOAD_ROOT / a.stored_path
    if not disk_path.exists():
        raise HTTPException(404, "File not found on disk")

    return FileResponse(
        path=str(disk_path),
        media_type=a.content_type or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{a.original_filename}"'},
    )


@router.get("/{attachment_id}/download")
def download_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    """Serve the file as a download."""
    a = db.query(models.FileAttachment).filter_by(
        id=attachment_id, project_id=user.project_id
    ).first()
    if not a:
        raise HTTPException(404, "Attachment not found")

    disk_path = UPLOAD_ROOT / a.stored_path
    if not disk_path.exists():
        raise HTTPException(404, "File not found on disk")

    return FileResponse(
        path=str(disk_path),
        media_type=a.content_type or "application/octet-stream",
        filename=a.original_filename,
        headers={"Content-Disposition": f'attachment; filename="{a.original_filename}"'},
    )


# ── ZIP download — used by "Download all" buttons in the procurement panes ──

def _safe_zip_filename(name: str) -> str:
    """Sanitise a filename for a zip member entry."""
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', str(name)).strip("_. ")
    return cleaned or "file"


@router.get("/zip")
def download_attachments_zip(
    ids: str = Query(..., description="Comma-separated FileAttachment ids"),
    filename: Optional[str] = Query(None, description="Suggested name for the resulting zip"),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    """Build a ZIP of the requested attachments. Every id must belong to the
    current project — others are silently dropped. Empty result yields 404."""
    raw = (ids or "").split(",")
    requested = []
    for s in raw:
        s = s.strip()
        if not s: continue
        try:
            requested.append(int(s))
        except ValueError:
            continue
    if not requested:
        raise HTTPException(400, "No valid attachment ids provided")

    rows = db.query(models.FileAttachment).filter(
        models.FileAttachment.id.in_(requested),
        models.FileAttachment.project_id == user.project_id,
    ).all()
    if not rows:
        raise HTTPException(404, "No matching attachments")

    buf = io.BytesIO()
    seen_names: dict = {}
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for r in rows:
            disk_path = UPLOAD_ROOT / r.stored_path
            if not disk_path.exists():
                continue
            base = _safe_zip_filename(r.original_filename or f"file_{r.id}")
            # Disambiguate duplicates: foo.pdf, foo (2).pdf, foo (3).pdf
            n = seen_names.get(base, 0) + 1
            seen_names[base] = n
            if n == 1:
                arc_name = base
            else:
                stem, dot, ext = base.rpartition(".")
                arc_name = f"{stem} ({n}).{ext}" if dot else f"{base} ({n})"
            try:
                zf.write(disk_path, arcname=arc_name)
            except OSError:
                continue
    buf.seek(0)

    out_name = _safe_zip_filename(filename) if filename else "attachments.zip"
    if not out_name.lower().endswith(".zip"):
        out_name = out_name + ".zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )
