"""
Export / Import router
Supports Excel-based export and import for:
  - Documents         (unique key: id)
  - Tasks             (unique key: id)
  - Procurement Plans (unique key: package_tag)

The export file doubles as the import template: it includes the data rows
on the first sheet and a Lookups sheet with valid reference values.
"""
import io
from datetime import datetime, date as date_type
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_

import models
import auth
from database import get_db

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(prefix="/api/export-import", tags=["export-import"])


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def _check_openpyxl():
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(500, "openpyxl is not installed. Run: pip install openpyxl")


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header(ws, row_num: int = 1):
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill


def _style_section_header(ws, row_num: int = 1):
    sec_fill = PatternFill("solid", fgColor="2E75B6")
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = sec_fill


def _autowidth(ws, max_width: int = 45):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


def _parse_date(val) -> Optional[str]:
    """
    Convert any date value to a YYYY-MM-DD string.
    Handles:
      - Python datetime / date objects (returned by openpyxl for date-formatted cells)
      - Strings in ISO format (YYYY-MM-DD) or common alternatives
    Returns None if the value cannot be interpreted as a date.
    """
    if val is None:
        return None
    # datetime.datetime and datetime.date both have .strftime()
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # ISO format (preferred)
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return s
    except ValueError:
        pass
    # Common alternative text formats
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_bool(val) -> bool:
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("TRUE", "YES", "1")


# ─────────────────────────────────────────────────────────────────────────────
# DOCUMENTS
# ─────────────────────────────────────────────────────────────────────────────

DOC_COLS = [
    ("id",                  "ID (unique key – leave blank for new records)"),
    ("package_tag",         "Package Tag *"),
    ("subservice_code",     "Subservice Code *"),
    ("document_type",       "Document Type * (TECHNICAL / COMMERCIAL)"),
    ("description",         "Description *"),
    ("area_tag",            "Area Tag"),
    ("unit_tag",            "Unit Tag"),
    ("require_area_review", "Require Area Review (TRUE / FALSE)"),
    ("require_unit_review", "Require Unit Review (TRUE / FALSE)"),
    ("start_date",          "Start Date (YYYY-MM-DD)"),
    ("first_issue_date",    "First Issue Date (YYYY-MM-DD)"),
    ("approval_due_date",   "Approval Due Date (YYYY-MM-DD)"),
    ("weight",              "Weight (default 8)"),
]


def _build_doc_workbook(db: Session, project_id: int) -> "Workbook":
    """Build the documents workbook (data + lookups). Used for both export and as import template."""
    docs = (
        db.query(models.Document)
        .filter_by(project_id=project_id)
        .order_by(models.Document.id)
        .all()
    )
    packages = (
        db.query(models.Package)
        .filter_by(project_id=project_id)
        .order_by(models.Package.tag_number)
        .all()
    )
    subservices = (
        db.query(models.Subservice)
        .filter(or_(
            models.Subservice.project_id == project_id,
            models.Subservice.project_id.is_(None),
        ))
        .order_by(models.Subservice.subservice_code)
        .all()
    )
    areas = (
        db.query(models.Area)
        .filter_by(project_id=project_id)
        .order_by(models.Area.tag)
        .all()
    )
    units = (
        db.query(models.Unit)
        .filter_by(project_id=project_id)
        .order_by(models.Unit.tag)
        .all()
    )

    wb = Workbook()

    # ── Data sheet ──
    ws = wb.active
    ws.title = "Documents"
    ws.append([c[1] for c in DOC_COLS])
    _style_header(ws, 1)
    for doc in docs:
        ws.append([
            doc.id,
            doc.package.tag_number if doc.package else "",
            doc.subservice.subservice_code if doc.subservice else "",
            doc.document_type,
            doc.description,
            doc.area.tag if doc.area else "",
            doc.unit.tag if doc.unit else "",
            "TRUE" if doc.require_area_review else "FALSE",
            "TRUE" if doc.require_unit_review else "FALSE",
            doc.start_date or "",
            doc.first_issue_date or "",
            doc.approval_due_date or "",
            doc.weight if doc.weight is not None else 8,
        ])
    _autowidth(ws)

    # ── Lookups sheet ──
    lws = wb.create_sheet("Lookups")
    lws.append(["PACKAGES", "", "SUBSERVICES", "", "AREAS", "UNITS", "", "DOCUMENT TYPES"])
    _style_section_header(lws, 1)
    lws.append(["tag_number", "name", "subservice_code", "subservice_name", "area_tag", "unit_tag", "", "document_type"])
    for cell in lws[2]:
        if cell.value:
            cell.font = Font(bold=True)

    doc_types = ["TECHNICAL", "COMMERCIAL"]
    n = max(len(packages), len(subservices), len(areas), len(units), 2)
    for i in range(n):
        lws.append([
            packages[i].tag_number    if i < len(packages)    else "",
            packages[i].name          if i < len(packages)    else "",
            subservices[i].subservice_code if i < len(subservices) else "",
            subservices[i].subservice_name if i < len(subservices) else "",
            areas[i].tag              if i < len(areas)       else "",
            units[i].tag              if i < len(units)       else "",
            "",
            doc_types[i]              if i < len(doc_types)   else "",
        ])
    _autowidth(lws)
    return wb


@router.get("/documents/export")
def export_documents(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    wb = _build_doc_workbook(db, user.project_id)
    return _xlsx_response(wb, "documents_export.xlsx")


def _parse_doc_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Documents" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Documents' not found in the uploaded file")
    ws = wb["Documents"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header

    packages = {p.tag_number: p for p in db.query(models.Package).filter_by(project_id=project_id).all()}
    subservices = {
        s.subservice_code: s
        for s in db.query(models.Subservice).filter(
            or_(models.Subservice.project_id == project_id, models.Subservice.project_id.is_(None))
        ).all()
    }
    areas    = {a.tag: a for a in db.query(models.Area).filter_by(project_id=project_id).all()}
    units    = {u.tag: u for u in db.query(models.Unit).filter_by(project_id=project_id).all()}
    existing = {d.id: d for d in db.query(models.Document).filter_by(project_id=project_id).all()}

    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        vals += [None] * max(0, len(DOC_COLS) - len(vals))

        raw_id = vals[0]
        try:
            doc_id = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            doc_id = None

        pkg_tag      = str(vals[1]  or "").strip()
        ss_code      = str(vals[2]  or "").strip()
        doc_type     = str(vals[3]  or "").strip().upper()
        description  = str(vals[4]  or "").strip()
        area_tag     = str(vals[5]  or "").strip() or None
        unit_tag     = str(vals[6]  or "").strip() or None
        req_area     = _parse_bool(vals[7])
        req_unit     = _parse_bool(vals[8])
        start_date   = _parse_date(vals[9])
        first_issue  = _parse_date(vals[10])
        approval_due = _parse_date(vals[11])
        try:
            weight = int(vals[12]) if vals[12] is not None and str(vals[12]).strip() != "" else 8
        except (ValueError, TypeError):
            weight = 8

        errors, warnings = [], []
        action = "UPDATE" if doc_id is not None else "CREATE"

        if doc_id is not None and doc_id not in existing:
            errors.append(f"ID {doc_id} not found in this project")
        if not pkg_tag:
            errors.append("Package Tag is required")
        elif pkg_tag not in packages:
            errors.append(f"Package '{pkg_tag}' not found")
        if not ss_code:
            errors.append("Subservice Code is required")
        elif ss_code not in subservices:
            errors.append(f"Subservice '{ss_code}' not found")
        if doc_type not in ("TECHNICAL", "COMMERCIAL"):
            errors.append(f"Document Type must be TECHNICAL or COMMERCIAL (got '{doc_type}')")
        if not description:
            errors.append("Description is required")
        if area_tag and area_tag not in areas:
            errors.append(f"Area '{area_tag}' not found")
        if unit_tag and unit_tag not in units:
            errors.append(f"Unit '{unit_tag}' not found")

        result.append({
            "row_num": row_idx,
            "id": doc_id,
            "package_tag": pkg_tag,
            "subservice_code": ss_code,
            "document_type": doc_type,
            "description": description,
            "area_tag": area_tag,
            "unit_tag": unit_tag,
            "require_area_review": req_area,
            "require_unit_review": req_unit,
            "start_date": start_date,
            "first_issue_date": first_issue,
            "approval_due_date": approval_due,
            "weight": weight,
            "action": action,
            "errors": errors,
            "warnings": warnings,
        })
    return result


@router.post("/documents/preview")
async def preview_documents_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import documents")
    file_bytes = await file.read()
    rows      = _parse_doc_rows(file_bytes, db, user.project_id)
    creates   = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates   = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {"rows": rows, "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count}}


@router.post("/documents/apply")
def apply_documents_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import documents")

    rows = payload.get("rows", [])
    packages = {p.tag_number: p for p in db.query(models.Package).filter_by(project_id=user.project_id).all()}
    subservices = {
        s.subservice_code: s
        for s in db.query(models.Subservice).filter(
            or_(models.Subservice.project_id == user.project_id, models.Subservice.project_id.is_(None))
        ).all()
    }
    areas = {a.tag: a for a in db.query(models.Area).filter_by(project_id=user.project_id).all()}
    units = {u.tag: u for u in db.query(models.Unit).filter_by(project_id=user.project_id).all()}
    now   = datetime.utcnow()
    created = updated = skipped = 0

    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue
        pkg = packages.get(row.get("package_tag", ""))
        ss  = subservices.get(row.get("subservice_code", ""))
        if not pkg or not ss:
            skipped += 1
            continue
        area = areas.get(row["area_tag"]) if row.get("area_tag") else None
        unit = units.get(row["unit_tag"]) if row.get("unit_tag") else None

        if row["action"] == "UPDATE" and row.get("id"):
            doc = db.query(models.Document).filter_by(id=row["id"], project_id=user.project_id).first()
            if not doc:
                skipped += 1
                continue
            doc.package_id          = pkg.id
            doc.subservice_id       = ss.id
            doc.document_type       = row["document_type"]
            doc.description         = row["description"]
            doc.area_id             = area.id if area else None
            doc.unit_id             = unit.id if unit else None
            doc.require_area_review = row.get("require_area_review", False)
            doc.require_unit_review = row.get("require_unit_review", False)
            doc.start_date          = row.get("start_date")
            doc.first_issue_date    = row.get("first_issue_date")
            doc.approval_due_date   = row.get("approval_due_date")
            doc.weight              = row.get("weight", 8)
            doc.updated_at          = now
            doc.updated_by_id       = user.id
            updated += 1
        else:
            db.add(models.Document(
                project_id          = user.project_id,
                package_id          = pkg.id,
                subservice_id       = ss.id,
                document_type       = row["document_type"],
                description         = row["description"],
                area_id             = area.id if area else None,
                unit_id             = unit.id if unit else None,
                require_area_review = row.get("require_area_review", False),
                require_unit_review = row.get("require_unit_review", False),
                start_date          = row.get("start_date"),
                first_issue_date    = row.get("first_issue_date"),
                approval_due_date   = row.get("approval_due_date"),
                weight              = row.get("weight", 8),
                status              = "NOT_STARTED",
                current_version     = 0,
                created_at          = now,
                created_by_id       = user.id,
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ─────────────────────────────────────────────────────────────────────────────
# TASKS
# ─────────────────────────────────────────────────────────────────────────────

TASK_COLS = [
    ("id",               "ID (unique key – leave blank for new records)"),
    ("package_tag",      "Package Tag"),
    ("description",      "Description *"),
    ("details",          "Details"),
    ("start_date",       "Start Date (YYYY-MM-DD)"),
    ("finish_date",      "Finish Date (YYYY-MM-DD)"),
    ("financial_weight", "Financial Weight"),
    ("area_tag",         "Area Tag"),
    ("unit_tag",         "Unit Tag"),
]


def _build_task_workbook(db: Session, project_id: int) -> "Workbook":
    """Build the tasks workbook (data + lookups). Used for both export and as import template."""
    tasks = (
        db.query(models.Task)
        .filter_by(project_id=project_id)
        .order_by(models.Task.id)
        .all()
    )
    packages = (
        db.query(models.Package)
        .filter_by(project_id=project_id)
        .order_by(models.Package.tag_number)
        .all()
    )
    areas = (
        db.query(models.Area)
        .filter_by(project_id=project_id)
        .order_by(models.Area.tag)
        .all()
    )
    units = (
        db.query(models.Unit)
        .filter_by(project_id=project_id)
        .order_by(models.Unit.tag)
        .all()
    )

    wb = Workbook()

    # ── Data sheet ──
    ws = wb.active
    ws.title = "Tasks"
    ws.append([c[1] for c in TASK_COLS])
    _style_header(ws, 1)
    for t in tasks:
        ws.append([
            t.id,
            t.package.tag_number if t.package else "",
            t.description,
            t.details or "",
            t.start_date or "",
            t.finish_date or "",
            t.financial_weight if t.financial_weight is not None else "",
            t.area.tag if t.area else "",
            t.unit.tag if t.unit else "",
        ])
    _autowidth(ws)

    # ── Lookups sheet ──
    lws = wb.create_sheet("Lookups")
    lws.append(["PACKAGES"])
    _style_section_header(lws, 1)
    lws.append(["tag_number", "name"])
    for cell in lws[2]:
        if cell.value:
            cell.font = Font(bold=True)
    for pkg in packages:
        lws.append([pkg.tag_number, pkg.name])
    lws.append([])
    row_offset = len(packages) + 4
    lws.append(["AREAS"])
    _style_section_header(lws, row_offset)
    lws.append(["tag", "description"])
    for cell in lws[row_offset + 1]:
        if cell.value:
            cell.font = Font(bold=True)
    for a in areas:
        lws.append([a.tag, a.description])
    lws.append([])
    row_offset2 = row_offset + len(areas) + 3
    lws.append(["UNITS"])
    _style_section_header(lws, row_offset2)
    lws.append(["tag", "description"])
    for cell in lws[row_offset2 + 1]:
        if cell.value:
            cell.font = Font(bold=True)
    for u in units:
        lws.append([u.tag, u.description])
    _autowidth(lws)
    return wb


@router.get("/tasks/export")
def export_tasks(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    wb = _build_task_workbook(db, user.project_id)
    return _xlsx_response(wb, "tasks_export.xlsx")


def _parse_task_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Tasks" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Tasks' not found in the uploaded file")
    ws = wb["Tasks"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header

    packages = {p.tag_number: p for p in db.query(models.Package).filter_by(project_id=project_id).all()}
    areas    = {a.tag: a for a in db.query(models.Area).filter_by(project_id=project_id).all()}
    units    = {u.tag: u for u in db.query(models.Unit).filter_by(project_id=project_id).all()}
    existing = {t.id: t for t in db.query(models.Task).filter_by(project_id=project_id).all()}

    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        vals += [None] * max(0, len(TASK_COLS) - len(vals))

        raw_id = vals[0]
        try:
            task_id = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            task_id = None

        pkg_tag          = str(vals[1] or "").strip() or None
        description      = str(vals[2] or "").strip()
        details          = str(vals[3] or "").strip() or None
        start_date       = _parse_date(vals[4])
        finish_date      = _parse_date(vals[5])
        try:
            financial_weight = float(vals[6]) if vals[6] is not None and str(vals[6]).strip() != "" else None
        except (ValueError, TypeError):
            financial_weight = None
        area_tag = str(vals[7] or "").strip() or None if len(vals) > 7 else None
        unit_tag = str(vals[8] or "").strip() or None if len(vals) > 8 else None

        errors, warnings = [], []
        action = "UPDATE" if task_id is not None else "CREATE"

        if task_id is not None and task_id not in existing:
            errors.append(f"Task ID {task_id} not found in this project")
        if not description:
            errors.append("Description is required")
        if pkg_tag and pkg_tag not in packages:
            errors.append(f"Package '{pkg_tag}' not found")
        if area_tag and area_tag not in areas:
            errors.append(f"Area '{area_tag}' not found")
        if unit_tag and unit_tag not in units:
            errors.append(f"Unit '{unit_tag}' not found")

        result.append({
            "row_num": row_idx,
            "id": task_id,
            "package_tag": pkg_tag,
            "description": description,
            "details": details,
            "start_date": start_date,
            "finish_date": finish_date,
            "financial_weight": financial_weight,
            "area_tag": area_tag,
            "unit_tag": unit_tag,
            "action": action,
            "errors": errors,
            "warnings": warnings,
        })
    return result


@router.post("/tasks/preview")
async def preview_tasks_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import tasks")
    file_bytes = await file.read()
    rows      = _parse_task_rows(file_bytes, db, user.project_id)
    creates   = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates   = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {"rows": rows, "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count}}


@router.post("/tasks/apply")
def apply_tasks_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import tasks")
    rows     = payload.get("rows", [])
    packages = {p.tag_number: p for p in db.query(models.Package).filter_by(project_id=user.project_id).all()}
    areas    = {a.tag: a for a in db.query(models.Area).filter_by(project_id=user.project_id).all()}
    units    = {u.tag: u for u in db.query(models.Unit).filter_by(project_id=user.project_id).all()}
    now      = datetime.utcnow()
    created  = updated = skipped = 0

    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue
        pkg  = packages.get(row["package_tag"]) if row.get("package_tag") else None
        area = areas.get(row["area_tag"]) if row.get("area_tag") else None
        unit = units.get(row["unit_tag"]) if row.get("unit_tag") else None

        if row["action"] == "UPDATE" and row.get("id"):
            task = db.query(models.Task).filter_by(id=row["id"], project_id=user.project_id).first()
            if not task:
                skipped += 1
                continue
            task.package_id       = pkg.id if pkg else None
            task.description      = row["description"]
            task.details          = row.get("details")
            task.start_date       = row.get("start_date")
            task.finish_date      = row.get("finish_date")
            task.financial_weight = row.get("financial_weight")
            task.area_id          = area.id if area else None
            task.unit_id          = unit.id if unit else None
            task.updated_at       = now
            task.updated_by_id    = user.id
            updated += 1
        else:
            db.add(models.Task(
                project_id       = user.project_id,
                package_id       = pkg.id if pkg else None,
                description      = row["description"],
                details          = row.get("details"),
                start_date       = row.get("start_date"),
                finish_date      = row.get("finish_date"),
                financial_weight = row.get("financial_weight"),
                area_id          = area.id if area else None,
                unit_id          = unit.id if unit else None,
                created_at       = now,
                created_by_id    = user.id,
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ─────────────────────────────────────────────────────────────────────────────
# PROCUREMENT PLANS
# ─────────────────────────────────────────────────────────────────────────────

def _build_procurement_workbook(db: Session, project_id: int) -> "Workbook":
    """Build the procurement plans workbook (data + lookups). Used for both export and as import template."""
    packages  = (
        db.query(models.Package).filter_by(project_id=project_id).order_by(models.Package.tag_number).all()
    )
    steps     = (
        db.query(models.ProcurementStep).filter_by(project_id=project_id).order_by(models.ProcurementStep.sort_order).all()
    )
    ctypes    = (
        db.query(models.ContractType).filter_by(project_id=project_id).order_by(models.ContractType.sort_order).all()
    )
    companies = (
        db.query(models.BiddingCompany).filter_by(project_id=project_id).order_by(models.BiddingCompany.name).all()
    )

    wb = Workbook()

    # ── Data sheet ──
    ws = wb.active
    ws.title = "Procurement Plans"
    headers = [
        "Package Tag (unique key)",
        "Package Name",
        "Contract Type",
        "Notes",
        "Bidding Companies (comma-separated)",
    ] + [f"Step: {s.step_id} ({s.description or ''})" for s in steps]
    ws.append(headers)
    _style_header(ws, 1)

    for pkg in packages:
        plan     = db.query(models.PackagePlan).filter_by(package_id=pkg.id, project_id=project_id).first()
        ct_name  = plan.contract_type.name if plan and plan.contract_type else ""
        notes    = plan.notes or "" if plan else ""
        bidders  = ""
        step_map = {}
        if plan:
            bidders  = ", ".join(b.company.name for b in plan.bidders if b.company)
            step_map = {sd.step_id: sd.due_date or "" for sd in plan.step_dates}
        row = [pkg.tag_number, pkg.name, ct_name, notes, bidders]
        for s in steps:
            row.append(step_map.get(s.id, ""))
        ws.append(row)
    _autowidth(ws)

    # ── Lookups sheet ──
    lws = wb.create_sheet("Lookups")
    lws.append(["PACKAGES", "", "CONTRACT TYPES", "", "BIDDING COMPANIES", "", "STEPS", ""])
    _style_section_header(lws, 1)
    lws.append(["tag_number", "name", "contract_type", "", "company_name", "", "step_id", "description"])
    for cell in lws[2]:
        if cell.value:
            cell.font = Font(bold=True)
    n = max(len(packages), len(ctypes), len(companies), len(steps), 1)
    for i in range(n):
        lws.append([
            packages[i].tag_number  if i < len(packages)  else "",
            packages[i].name        if i < len(packages)  else "",
            ctypes[i].name          if i < len(ctypes)    else "",
            "",
            companies[i].name       if i < len(companies) else "",
            "",
            steps[i].step_id        if i < len(steps)     else "",
            steps[i].description    if i < len(steps)     else "",
        ])
    _autowidth(lws)
    return wb


@router.get("/procurement/export")
def export_procurement_plans(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    wb = _build_procurement_workbook(db, user.project_id)
    return _xlsx_response(wb, "procurement_plans_export.xlsx")


def _parse_procurement_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Procurement Plans" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Procurement Plans' not found in the uploaded file")
    ws = wb["Procurement Plans"]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise HTTPException(400, "File is empty")

    # Detect step columns from header (columns 6+)
    step_ids_in_header = []
    for h in header_row[5:]:
        h_str = str(h or "").strip()
        if h_str.startswith("Step: "):
            rest    = h_str[6:]
            step_id = rest.split(" (")[0].strip() if " (" in rest else rest.strip()
            step_ids_in_header.append(step_id)
        else:
            step_ids_in_header.append(h_str)

    packages  = {p.tag_number: p  for p in db.query(models.Package).filter_by(project_id=project_id).all()}
    ctypes    = {ct.name: ct       for ct in db.query(models.ContractType).filter_by(project_id=project_id).all()}
    companies = {c.name: c         for c in db.query(models.BiddingCompany).filter_by(project_id=project_id).all()}
    steps     = {s.step_id: s      for s in db.query(models.ProcurementStep).filter_by(project_id=project_id).all()}
    existing_plans = {
        plan.package_id: plan
        for plan in db.query(models.PackagePlan).filter_by(project_id=project_id).all()
    }

    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue

        pkg_tag = str(vals[0] if len(vals) > 0 else "").strip()
        ct_name = str(vals[2] if len(vals) > 2 else "").strip() or None
        notes   = str(vals[3] if len(vals) > 3 else "").strip() or None
        b_str   = str(vals[4] if len(vals) > 4 else "").strip()
        b_names = [b.strip() for b in b_str.split(",") if b.strip()] if b_str else []

        step_dates = {}
        for i, sid in enumerate(step_ids_in_header):
            raw = vals[5 + i] if 5 + i < len(vals) else None
            step_dates[sid] = _parse_date(raw)

        errors, warnings = [], []

        if not pkg_tag:
            errors.append("Package Tag is required")
        elif pkg_tag not in packages:
            errors.append(f"Package '{pkg_tag}' not found")

        if ct_name and ct_name not in ctypes:
            errors.append(f"Contract type '{ct_name}' not found")

        unknown_co = [n for n in b_names if n not in companies]
        if unknown_co:
            warnings.append(f"Unknown bidding companies (skipped): {', '.join(unknown_co)}")

        unknown_st = [sid for sid in step_dates if sid not in steps]
        if unknown_st:
            warnings.append(f"Unknown steps (skipped): {', '.join(unknown_st)}")

        pkg    = packages.get(pkg_tag)
        action = "UPDATE" if (pkg and pkg.id in existing_plans) else "CREATE"

        result.append({
            "row_num": row_idx,
            "package_tag": pkg_tag,
            "contract_type_name": ct_name,
            "notes": notes,
            "bidder_names": b_names,
            "step_dates": step_dates,
            "action": action,
            "errors": errors,
            "warnings": warnings,
        })
    return result


@router.post("/procurement/preview")
async def preview_procurement_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import procurement plans")
    file_bytes = await file.read()
    rows      = _parse_procurement_rows(file_bytes, db, user.project_id)
    creates   = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates   = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {"rows": rows, "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count}}


@router.post("/procurement/apply")
def apply_procurement_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import procurement plans")
    rows      = payload.get("rows", [])
    packages  = {p.tag_number: p  for p in db.query(models.Package).filter_by(project_id=user.project_id).all()}
    ctypes    = {ct.name: ct       for ct in db.query(models.ContractType).filter_by(project_id=user.project_id).all()}
    companies = {c.name: c         for c in db.query(models.BiddingCompany).filter_by(project_id=user.project_id).all()}
    steps     = {s.step_id: s      for s in db.query(models.ProcurementStep).filter_by(project_id=user.project_id).all()}
    now       = datetime.utcnow()
    created   = updated = skipped = 0

    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue
        pkg = packages.get(row.get("package_tag", ""))
        if not pkg:
            skipped += 1
            continue
        ct = ctypes.get(row["contract_type_name"]) if row.get("contract_type_name") else None

        plan = db.query(models.PackagePlan).filter_by(package_id=pkg.id, project_id=user.project_id).first()
        if plan:
            plan.contract_type_id = ct.id if ct else plan.contract_type_id
            plan.notes            = row.get("notes")
            plan.updated_at       = now
            plan.updated_by_id    = user.id
            updated += 1
        else:
            plan = models.PackagePlan(
                project_id       = user.project_id,
                package_id       = pkg.id,
                contract_type_id = ct.id if ct else None,
                notes            = row.get("notes"),
                created_at       = now,
                created_by_id    = user.id,
            )
            db.add(plan)
            db.flush()
            created += 1

        # Bidders — add only new ones (never remove existing)
        existing_bidder_ids = {b.company_id for b in plan.bidders}
        for bname in (row.get("bidder_names") or []):
            co = companies.get(bname)
            if co and co.id not in existing_bidder_ids:
                db.add(models.PackagePlanBidder(plan_id=plan.id, company_id=co.id))
                existing_bidder_ids.add(co.id)

        # Step dates — upsert
        existing_step_dates = {sd.step_id: sd for sd in plan.step_dates}
        for sid_str, date_val in (row.get("step_dates") or {}).items():
            step = steps.get(sid_str)
            if not step:
                continue
            if step.id in existing_step_dates:
                existing_step_dates[step.id].due_date = date_val
            else:
                db.add(models.PackagePlanStepDate(plan_id=plan.id, step_id=step.id, due_date=date_val))

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ─────────────────────────────────────────────────────────────────────────────
# ITP RECORDS
# ─────────────────────────────────────────────────────────────────────────────

ITP_COLS = [
    ("id",                   "ID (unique key – leave blank for new records)"),
    ("package_tag",          "Package Tag *"),
    ("test_type",            "Test Type *"),
    ("test",                 "Test *"),
    ("details",              "Details"),
    ("witness_level",        "Witness Level Code * (H / W / R / I)"),
    ("area_tag",             "Area Tag"),
    ("unit_tag",             "Unit Tag"),
    ("acceptance_criteria",  "Acceptance Criteria"),
    ("planned_date",         "Planned Date (YYYY-MM-DD)"),
    ("status",               "Status (DRAFT / PLANNED / PASSED / FAILED)"),
]


def _build_itp_workbook(db: Session, project_id: int) -> "Workbook":
    """Build the ITP workbook (data + lookups). Used for both export and as import template."""
    records = (
        db.query(models.ITPRecord)
        .filter_by(project_id=project_id)
        .order_by(models.ITPRecord.id)
        .all()
    )
    packages = (
        db.query(models.Package)
        .filter_by(project_id=project_id)
        .order_by(models.Package.tag_number)
        .all()
    )
    test_types = (
        db.query(models.ITPTestType)
        .filter_by(project_id=project_id)
        .order_by(models.ITPTestType.name)
        .all()
    )
    witness_levels = (
        db.query(models.ITPWitnessLevel)
        .filter_by(project_id=project_id)
        .order_by(models.ITPWitnessLevel.sort_order)
        .all()
    )
    areas = (
        db.query(models.Area)
        .filter_by(project_id=project_id)
        .order_by(models.Area.tag)
        .all()
    )
    units = (
        db.query(models.Unit)
        .filter_by(project_id=project_id)
        .order_by(models.Unit.tag)
        .all()
    )

    wb = Workbook()

    # ── Data sheet ──
    ws = wb.active
    ws.title = "ITP Records"
    ws.append([c[1] for c in ITP_COLS])
    _style_header(ws, 1)
    for r in records:
        ws.append([
            r.id,
            r.package.tag_number if r.package else "",
            r.test_type.name if r.test_type else "",
            r.test or "",
            r.details or "",
            r.witness_level.code if r.witness_level else "",
            r.area.tag if r.area else "",
            r.unit.tag if r.unit else "",
            r.acceptance_criteria or "",
            r.planned_date or "",
            r.status or "DRAFT",
        ])
    _autowidth(ws)

    # ── Lookups sheet ──
    lws = wb.create_sheet("Lookups")
    lws.append(["PACKAGES", "", "TEST TYPES", "WITNESS LEVELS", "", "AREAS", "UNITS", "STATUSES"])
    _style_section_header(lws, 1)
    lws.append(["tag_number", "name", "test_type_name", "code", "wl_name", "area_tag", "unit_tag", "status"])
    for cell in lws[2]:
        if cell.value:
            cell.font = Font(bold=True)

    statuses = ["DRAFT", "PLANNED", "PASSED", "FAILED"]
    n = max(len(packages), len(test_types), len(witness_levels), len(areas), len(units), len(statuses))
    for i in range(n):
        lws.append([
            packages[i].tag_number       if i < len(packages)       else "",
            packages[i].name             if i < len(packages)       else "",
            test_types[i].name           if i < len(test_types)     else "",
            witness_levels[i].code       if i < len(witness_levels) else "",
            witness_levels[i].name       if i < len(witness_levels) else "",
            areas[i].tag                 if i < len(areas)          else "",
            units[i].tag                 if i < len(units)          else "",
            statuses[i]                  if i < len(statuses)       else "",
        ])
    _autowidth(lws)
    return wb


@router.get("/itp/export")
def export_itp(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    wb = _build_itp_workbook(db, user.project_id)
    return _xlsx_response(wb, "itp_export.xlsx")


def _parse_itp_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "ITP Records" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'ITP Records' not found in the uploaded file")
    ws = wb["ITP Records"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header

    packages       = {p.tag_number: p for p in db.query(models.Package).filter_by(project_id=project_id).all()}
    test_types     = {tt.name: tt for tt in db.query(models.ITPTestType).filter_by(project_id=project_id).all()}
    witness_levels = {wl.code: wl for wl in db.query(models.ITPWitnessLevel).filter_by(project_id=project_id).all()}
    areas          = {a.tag: a for a in db.query(models.Area).filter_by(project_id=project_id).all()}
    units          = {u.tag: u for u in db.query(models.Unit).filter_by(project_id=project_id).all()}
    existing       = {r.id: r for r in db.query(models.ITPRecord).filter_by(project_id=project_id).all()}

    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        vals += [None] * max(0, len(ITP_COLS) - len(vals))

        raw_id = vals[0]
        try:
            rec_id = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            rec_id = None

        pkg_tag      = str(vals[1] or "").strip()
        test_type_nm = str(vals[2] or "").strip()
        test         = str(vals[3] or "").strip()
        details      = str(vals[4] or "").strip() or None
        wl_code      = str(vals[5] or "").strip().upper()
        area_tag     = str(vals[6] or "").strip() or None
        unit_tag     = str(vals[7] or "").strip() or None
        acc_criteria = str(vals[8] or "").strip() or None
        planned_date = _parse_date(vals[9])
        status_raw   = str(vals[10] or "").strip().upper() or "DRAFT"

        errors, warnings = [], []
        action = "UPDATE" if rec_id is not None else "CREATE"

        if rec_id is not None and rec_id not in existing:
            errors.append(f"ID {rec_id} not found in this project")
        if not pkg_tag:
            errors.append("Package Tag is required")
        elif pkg_tag not in packages:
            errors.append(f"Package '{pkg_tag}' not found")
        if not test_type_nm:
            errors.append("Test Type is required")
        elif test_type_nm not in test_types:
            errors.append(f"Test Type '{test_type_nm}' not found (check Lookups sheet)")
        if not test:
            errors.append("Test is required")
        if not wl_code:
            errors.append("Witness Level Code is required")
        elif wl_code not in witness_levels:
            errors.append(f"Witness Level '{wl_code}' not found (check Lookups sheet)")
        if area_tag and area_tag not in areas:
            errors.append(f"Area '{area_tag}' not found")
        if unit_tag and unit_tag not in units:
            errors.append(f"Unit '{unit_tag}' not found")
        if status_raw not in ("DRAFT", "PLANNED", "PASSED", "FAILED"):
            warnings.append(f"Unknown status '{status_raw}', will default to DRAFT")
            status_raw = "DRAFT"

        result.append({
            "row_num":             row_idx,
            "id":                  rec_id,
            "package_tag":         pkg_tag,
            "test_type":           test_type_nm,
            "test":                test,
            "details":             details,
            "witness_level":       wl_code,
            "area_tag":            area_tag,
            "unit_tag":            unit_tag,
            "acceptance_criteria": acc_criteria,
            "planned_date":        planned_date,
            "status":              status_raw,
            "action":              action,
            "errors":              errors,
            "warnings":            warnings,
        })
    return result


@router.post("/itp/preview")
async def preview_itp_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    # VENDORs are also allowed to import ITP records for packages they are
    # linked to; per-row authz is enforced when applying.
    if user.role not in ("ADMIN", "PROJECT_OWNER", "VENDOR"):
        raise HTTPException(403, "Only ADMIN, PROJECT_OWNER or VENDOR can import ITP records")
    file_bytes = await file.read()
    rows      = _parse_itp_rows(file_bytes, db, user.project_id)
    creates   = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates   = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {"rows": rows, "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count}}


@router.post("/itp/apply")
def apply_itp_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    # VENDORs can import ITP records, but only for packages they are linked
    # to; other packages in the payload are silently skipped.
    if user.role not in ("ADMIN", "PROJECT_OWNER", "VENDOR"):
        raise HTTPException(403, "Only ADMIN, PROJECT_OWNER or VENDOR can import ITP records")

    rows           = payload.get("rows", [])
    packages       = {p.tag_number: p for p in db.query(models.Package).filter_by(project_id=user.project_id).all()}
    test_types     = {tt.name: tt for tt in db.query(models.ITPTestType).filter_by(project_id=user.project_id).all()}
    witness_levels = {wl.code: wl for wl in db.query(models.ITPWitnessLevel).filter_by(project_id=user.project_id).all()}
    areas          = {a.tag: a for a in db.query(models.Area).filter_by(project_id=user.project_id).all()}
    units          = {u.tag: u for u in db.query(models.Unit).filter_by(project_id=user.project_id).all()}
    # For vendors, precompute the package IDs they are linked to.
    allowed_pkg_ids = None
    if user.role == "VENDOR":
        if not user.contact_id:
            allowed_pkg_ids = set()
        else:
            allowed_pkg_ids = {
                pc.package_id for pc in db.query(models.PackageContact)
                    .filter_by(contact_id=user.contact_id).all()
            }
    now            = datetime.utcnow()
    created = updated = skipped = 0

    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue
        pkg = packages.get(row.get("package_tag", ""))
        tt  = test_types.get(row.get("test_type", ""))
        wl  = witness_levels.get(row.get("witness_level", ""))
        if not pkg or not tt or not wl:
            skipped += 1
            continue
        if allowed_pkg_ids is not None and pkg.id not in allowed_pkg_ids:
            skipped += 1
            continue
        area = areas.get(row["area_tag"]) if row.get("area_tag") else None
        unit = units.get(row["unit_tag"]) if row.get("unit_tag") else None

        if row["action"] == "UPDATE" and row.get("id"):
            rec = db.query(models.ITPRecord).filter_by(id=row["id"], project_id=user.project_id).first()
            if not rec:
                skipped += 1
                continue
            rec.package_id          = pkg.id
            rec.test_type_id        = tt.id
            rec.test                = row["test"]
            rec.details             = row.get("details")
            rec.witness_level_id    = wl.id
            rec.area_id             = area.id if area else None
            rec.unit_id             = unit.id if unit else None
            rec.acceptance_criteria = row.get("acceptance_criteria")
            rec.planned_date        = row.get("planned_date")
            rec.status              = row.get("status", "DRAFT")
            rec.updated_at          = now
            rec.updated_by_id       = user.id
            updated += 1
        else:
            db.add(models.ITPRecord(
                project_id          = user.project_id,
                package_id          = pkg.id,
                test_type_id        = tt.id,
                test                = row["test"],
                details             = row.get("details"),
                witness_level_id    = wl.id,
                area_id             = area.id if area else None,
                unit_id             = unit.id if unit else None,
                acceptance_criteria = row.get("acceptance_criteria"),
                planned_date        = row.get("planned_date"),
                status              = row.get("status", "DRAFT"),
                approval_status     = "TO_SUBMIT",
                created_at          = now,
                created_by_id       = user.id,
                project_seq_id      = models.next_project_seq(db, models.ITPRecord, user.project_id),
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ─────────────────────────────────────────────────────────────────────────────
# RISK REGISTER
# ─────────────────────────────────────────────────────────────────────────────

RISK_COLS = [
    ("id",                    "ID"),
    ("title",                 "Title"),
    ("description",           "Description"),
    ("status",                "Status"),
    ("category",              "Category"),
    ("phase",                 "Phase"),
    ("date_opened",           "Date Opened"),
    ("date_closed",           "Date Closed"),
    ("owner",                 "Owner"),
    ("prob_score_before",     "Prob Score (Before)"),
    ("capex_score_before",    "CAPEX Score (Before)"),
    ("schedule_score_before", "Schedule Score (Before)"),
    ("capex_value",           "CAPEX at Risk"),
    ("schedule_value",        "Schedule at Risk (months)"),
    ("mitigation_type",       "Mitigation Type"),
    ("mitigation_action",     "Mitigation Action"),
    ("action_due_date",       "Action Due Date"),
    ("action_status",         "Action Status"),
    ("prob_score_after",      "Prob Score (After)"),
    ("capex_score_after",     "CAPEX Score (After)"),
    ("schedule_score_after",  "Schedule Score (After)"),
    ("secondary_effects",     "Secondary Effects"),
]


def _build_risk_workbook(db: Session, project_id: int) -> "Workbook":
    risks = (
        db.query(models.Risk)
        .filter_by(project_id=project_id)
        .order_by(models.Risk.id)
        .all()
    )
    categories = (
        db.query(models.RiskCategory)
        .filter_by(project_id=project_id)
        .order_by(models.RiskCategory.name)
        .all()
    )
    phases = (
        db.query(models.RiskPhase)
        .filter_by(project_id=project_id)
        .order_by(models.RiskPhase.name)
        .all()
    )
    contacts = (
        db.query(models.Contact)
        .filter_by(project_id=project_id)
        .order_by(models.Contact.name)
        .all()
    )

    wb = Workbook()

    # ── Data sheet ──
    ws = wb.active
    ws.title = "Risk Register"
    ws.append([c[1] for c in RISK_COLS])
    _style_header(ws, 1)

    for r in risks:
        ws.append([
            r.id,
            r.title,
            r.description or "",
            r.status,
            r.category.name if r.category else "",
            r.phase.name if r.phase else "",
            r.date_opened or "",
            r.date_closed or "",
            r.owner.name if r.owner else "",
            r.prob_score_before,
            r.capex_score_before,
            r.schedule_score_before,
            r.capex_value if r.capex_value is not None else "",
            r.schedule_value if r.schedule_value is not None else "",
            r.mitigation_type or "",
            r.mitigation_action or "",
            r.action_due_date or "",
            r.action_status or "",
            r.prob_score_after,
            r.capex_score_after,
            r.schedule_score_after,
            r.secondary_effects or "",
        ])
    _autowidth(ws)

    # ── Lookups sheet ──
    lws = wb.create_sheet("Lookups")

    lws.append(["CATEGORIES"])
    _style_section_header(lws, 1)
    lws.append(["name"])
    for cell in lws[2]:
        if cell.value:
            cell.font = Font(bold=True)
    for c in categories:
        lws.append([c.name])

    lws.append([])
    row_offset = len(categories) + 4
    lws.append(["PHASES"])
    _style_section_header(lws, row_offset)
    lws.append(["name"])
    for cell in lws[row_offset + 1]:
        if cell.value:
            cell.font = Font(bold=True)
    for p in phases:
        lws.append([p.name])

    lws.append([])
    row_offset2 = row_offset + len(phases) + 3
    lws.append(["OWNERS (contacts)"])
    _style_section_header(lws, row_offset2)
    lws.append(["name"])
    for cell in lws[row_offset2 + 1]:
        if cell.value:
            cell.font = Font(bold=True)
    for ct in contacts:
        lws.append([ct.name])

    lws.append([])
    row_offset3 = row_offset2 + len(contacts) + 3
    lws.append(["VALID VALUES"])
    _style_section_header(lws, row_offset3)
    lws.append(["Field", "Allowed Values"])
    for cell in lws[row_offset3 + 1]:
        if cell.value:
            cell.font = Font(bold=True)
    lws.append(["Status", "OPEN, MONITORING, CLOSED"])
    lws.append(["Mitigation Type", "AVOID, REDUCE, TRANSFER, ACCEPT"])
    lws.append(["Action Status", "NOT_STARTED, IN_PROGRESS, CLOSED, ON_HOLD"])
    lws.append(["Prob/CAPEX/Schedule Scores", "1, 2, 3, 4, 5"])

    _autowidth(lws)
    return wb


@router.get("/risks/export")
def export_risks(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    wb = _build_risk_workbook(db, user.project_id)
    return _xlsx_response(wb, "risk_register_export.xlsx")


def _parse_risk_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Risk Register" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Risk Register' not found in the uploaded file")
    ws = wb["Risk Register"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header

    categories = {c.name.lower(): c for c in db.query(models.RiskCategory).filter_by(project_id=project_id).all()}
    phases     = {p.name.lower(): p for p in db.query(models.RiskPhase).filter_by(project_id=project_id).all()}
    contacts   = {c.name.lower(): c for c in db.query(models.Contact).filter_by(project_id=project_id).all()}
    existing   = {r.id: r for r in db.query(models.Risk).filter_by(project_id=project_id).all()}

    valid_statuses = {"OPEN", "MONITORING", "CLOSED"}
    valid_mitigation = {"AVOID", "REDUCE", "TRANSFER", "ACCEPT", ""}
    valid_action = {"NOT_STARTED", "IN_PROGRESS", "CLOSED", "ON_HOLD", ""}

    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        vals += [None] * max(0, len(RISK_COLS) - len(vals))

        errors, warnings = [], []

        raw_id = vals[0]
        try:
            risk_id = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            risk_id = None

        action = "UPDATE" if risk_id is not None else "CREATE"

        title              = str(vals[1] or "").strip()
        description        = str(vals[2] or "").strip()
        status             = str(vals[3] or "OPEN").strip().upper()
        category_name      = str(vals[4] or "").strip()
        phase_name         = str(vals[5] or "").strip()
        date_opened        = _parse_date(vals[6])
        date_closed        = _parse_date(vals[7])
        owner_name         = str(vals[8] or "").strip()

        def _parse_score(v):
            if v is None or str(v).strip() == "":
                return None
            try:
                s = int(v)
                return s if 1 <= s <= 5 else None
            except (ValueError, TypeError):
                return None

        prob_score_before     = _parse_score(vals[9])
        capex_score_before    = _parse_score(vals[10])
        schedule_score_before = _parse_score(vals[11])

        try:
            capex_value = float(vals[12]) if vals[12] is not None and str(vals[12]).strip() != "" else None
        except (ValueError, TypeError):
            capex_value = None
        try:
            schedule_value = float(vals[13]) if vals[13] is not None and str(vals[13]).strip() != "" else None
        except (ValueError, TypeError):
            schedule_value = None

        mitigation_type    = str(vals[14] or "").strip().upper()
        mitigation_action  = str(vals[15] or "").strip()
        action_due_date    = _parse_date(vals[16])
        action_status      = str(vals[17] or "NOT_STARTED").strip().upper()

        prob_score_after     = _parse_score(vals[18])
        capex_score_after    = _parse_score(vals[19])
        schedule_score_after = _parse_score(vals[20])

        secondary_effects  = str(vals[21] or "").strip() if len(vals) > 21 else ""

        # Validate
        if risk_id is not None and risk_id not in existing:
            errors.append(f"Risk ID {risk_id} not found in this project")
        if not title:
            errors.append("Title is required")
        if not description:
            errors.append("Description is required")
        if status not in valid_statuses:
            errors.append(f"Invalid status '{status}'")
        if not category_name:
            errors.append("Category is required")
        elif category_name.lower() not in categories:
            errors.append(f"Category '{category_name}' not found")
        if not phase_name:
            errors.append("Phase is required")
        elif phase_name.lower() not in phases:
            errors.append(f"Phase '{phase_name}' not found")
        if owner_name and owner_name.lower() not in contacts:
            errors.append(f"Owner '{owner_name}' not found in contacts")
        if mitigation_type and mitigation_type not in valid_mitigation:
            errors.append(f"Invalid mitigation type '{mitigation_type}'")
        if action_status and action_status not in valid_action:
            errors.append(f"Invalid action status '{action_status}'")

        result.append({
            "row_num": row_idx,
            "id": risk_id,
            "title": title,
            "description": description,
            "status": status,
            "category_name": category_name,
            "phase_name": phase_name,
            "date_opened": date_opened,
            "date_closed": date_closed,
            "owner_name": owner_name,
            "prob_score_before": prob_score_before,
            "capex_score_before": capex_score_before,
            "schedule_score_before": schedule_score_before,
            "capex_value": capex_value,
            "schedule_value": schedule_value,
            "mitigation_type": mitigation_type or None,
            "mitigation_action": mitigation_action or None,
            "action_due_date": action_due_date,
            "action_status": action_status or "NOT_STARTED",
            "prob_score_after": prob_score_after,
            "capex_score_after": capex_score_after,
            "schedule_score_after": schedule_score_after,
            "secondary_effects": secondary_effects or None,
            "action": action,
            "errors": errors,
            "warnings": warnings,
        })
    return result


@router.post("/risks/preview")
async def preview_risks_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import risks")
    file_bytes = await file.read()
    rows      = _parse_risk_rows(file_bytes, db, user.project_id)
    creates   = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates   = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {"rows": rows, "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count}}


@router.post("/risks/apply")
def apply_risks_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import risks")
    rows       = payload.get("rows", [])
    categories = {c.name.lower(): c for c in db.query(models.RiskCategory).filter_by(project_id=user.project_id).all()}
    phases     = {p.name.lower(): p for p in db.query(models.RiskPhase).filter_by(project_id=user.project_id).all()}
    contacts   = {c.name.lower(): c for c in db.query(models.Contact).filter_by(project_id=user.project_id).all()}
    now        = datetime.utcnow()
    created = updated = skipped = 0

    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue
        cat     = categories.get(row["category_name"].lower()) if row.get("category_name") else None
        phase   = phases.get(row["phase_name"].lower()) if row.get("phase_name") else None
        owner   = contacts.get(row["owner_name"].lower()) if row.get("owner_name") else None

        if row["action"] == "UPDATE" and row.get("id"):
            risk = db.query(models.Risk).filter_by(id=row["id"], project_id=user.project_id).first()
            if not risk:
                skipped += 1
                continue
            risk.title                 = row["title"]
            risk.description           = row.get("description")
            risk.status                = row.get("status", "OPEN")
            risk.category_id           = cat.id if cat else None
            risk.phase_id              = phase.id if phase else None
            risk.date_opened           = row.get("date_opened")
            risk.date_closed           = row.get("date_closed")
            risk.owner_id              = owner.id if owner else None
            risk.prob_score_before     = row.get("prob_score_before")
            risk.capex_score_before    = row.get("capex_score_before")
            risk.schedule_score_before = row.get("schedule_score_before")
            risk.capex_value           = row.get("capex_value")
            risk.schedule_value        = row.get("schedule_value")
            risk.mitigation_type       = row.get("mitigation_type")
            risk.mitigation_action     = row.get("mitigation_action")
            risk.action_due_date       = row.get("action_due_date")
            risk.action_status         = row.get("action_status", "NOT_STARTED")
            risk.prob_score_after      = row.get("prob_score_after")
            risk.capex_score_after     = row.get("capex_score_after")
            risk.schedule_score_after  = row.get("schedule_score_after")
            risk.secondary_effects     = row.get("secondary_effects")
            risk.updated_at            = now
            risk.updated_by_id         = user.id
            updated += 1
        else:
            db.add(models.Risk(
                project_id             = user.project_id,
                project_seq_id         = models.next_project_seq(db, models.Risk, user.project_id),
                title                  = row["title"],
                description            = row.get("description"),
                status                 = row.get("status", "OPEN"),
                category_id            = cat.id if cat else None,
                phase_id               = phase.id if phase else None,
                date_opened            = row.get("date_opened"),
                date_closed            = row.get("date_closed"),
                owner_id               = owner.id if owner else None,
                prob_score_before      = row.get("prob_score_before"),
                capex_score_before     = row.get("capex_score_before"),
                schedule_score_before  = row.get("schedule_score_before"),
                capex_value            = row.get("capex_value"),
                schedule_value         = row.get("schedule_value"),
                mitigation_type        = row.get("mitigation_type"),
                mitigation_action      = row.get("mitigation_action"),
                action_due_date        = row.get("action_due_date"),
                action_status          = row.get("action_status", "NOT_STARTED"),
                prob_score_after       = row.get("prob_score_after"),
                capex_score_after      = row.get("capex_score_after"),
                schedule_score_after   = row.get("schedule_score_after"),
                secondary_effects      = row.get("secondary_effects"),
                created_at             = now,
                created_by_id          = user.id,
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ─────────────────────────────────────────────────────────────────────────────
# WORKERS / SUBCONTRACTORS IMPORT
# Two data sheets in one workbook (Subcontractors + Workers) + Lookups sheet.
# Mirrors the Risk Register import pattern: Export template, Preview, Apply.
# ─────────────────────────────────────────────────────────────────────────────

SUB_COLS = [
    ("id",              "ID"),
    ("package_tag",     "Package (tag)"),
    ("company",         "Company"),
    ("contact_person",  "Contact Person"),
    ("phone",           "Phone"),
    ("email",           "Email"),
    ("description",     "Scope / Description"),
]

WORKER_COLS = [
    ("id",                    "ID"),
    ("package_tag",           "Package (tag)"),
    ("name",                  "Name"),
    ("phone",                 "Phone"),
    ("is_subcontractor",      "Is Subcontractor (Y/N)"),
    ("subcontractor_company", "Subcontractor Company"),
]


def _build_workers_subs_workbook(db: Session, project_id: int) -> "Workbook":
    """Excel workbook with two data sheets (Subcontractors, Workers) + Lookups."""
    packages = (
        db.query(models.Package)
          .filter(models.Package.project_id == project_id)
          .order_by(models.Package.tag_number.asc())
          .all()
    )
    subs = (
        db.query(models.Subcontractor)
          .filter(models.Subcontractor.project_id == project_id)
          .order_by(models.Subcontractor.id.asc())
          .all()
    )
    workers = (
        db.query(models.Worker)
          .filter(models.Worker.project_id == project_id)
          .order_by(models.Worker.id.asc())
          .all()
    )

    wb = Workbook()

    # ── Subcontractors sheet ──
    ws = wb.active
    ws.title = "Subcontractors"
    ws.append([c[1] for c in SUB_COLS])
    _style_header(ws, 1)
    for s in subs:
        ws.append([
            s.id,
            s.package.tag_number if s.package else "",
            s.company,
            s.contact_person or "",
            s.phone or "",
            s.email or "",
            s.description or "",
        ])
    _autowidth(ws)

    # ── Workers sheet ──
    wws = wb.create_sheet("Workers")
    wws.append([c[1] for c in WORKER_COLS])
    _style_header(wws, 1)
    for w in workers:
        wws.append([
            w.id,
            w.package.tag_number if w.package else "",
            w.name,
            w.phone or "",
            "Y" if w.is_subcontractor else "N",
            w.subcontractor.company if w.subcontractor else "",
        ])
    _autowidth(wws)

    # ── Lookups sheet ──
    lws = wb.create_sheet("Lookups")
    lws.append(["PACKAGES"])
    _style_section_header(lws, 1)
    lws.append(["tag", "name"])
    for cell in lws[2]:
        if cell.value:
            cell.font = Font(bold=True)
    for p in packages:
        lws.append([p.tag_number, p.name or ""])

    lws.append([])
    row_offset = len(packages) + 4
    lws.append(["SUBCONTRACTORS (existing companies)"])
    _style_section_header(lws, row_offset)
    lws.append(["company", "package tag"])
    for cell in lws[row_offset + 1]:
        if cell.value:
            cell.font = Font(bold=True)
    for s in subs:
        lws.append([s.company, s.package.tag_number if s.package else ""])

    lws.append([])
    row_offset2 = row_offset + len(subs) + 3
    lws.append(["NOTES"])
    _style_section_header(lws, row_offset2)
    lws.append(["Workers reference a subcontractor by its Company name. The subcontractor can either"])
    lws.append(["already exist in this project, or be created in the Subcontractors sheet of this same file."])
    lws.append(["Leave the ID column blank to create a new row; fill in an existing ID to update."])

    _autowidth(lws)
    return wb


@router.get("/workers-subs/export")
def export_workers_subs(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    wb = _build_workers_subs_workbook(db, user.project_id)
    return _xlsx_response(wb, "workers_subcontractors_export.xlsx")


def _parse_workers_subs_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    """Return a unified list of preview rows for both Subcontractor and Worker sheets.

    Each row has a `kind` of 'subcontractor' or 'worker' so the frontend can render
    them in one table. Workers that reference a subcontractor being CREATED in the
    same file are flagged as valid (forward reference within the import)."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Subcontractors" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Subcontractors' not found in the uploaded file")
    if "Workers" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Workers' not found in the uploaded file")

    packages = {p.tag_number.lower(): p for p in db.query(models.Package).filter_by(project_id=project_id).all() if p.tag_number}
    existing_subs = {s.id: s for s in db.query(models.Subcontractor).filter_by(project_id=project_id).all()}
    existing_workers = {w.id: w for w in db.query(models.Worker).filter_by(project_id=project_id).all()}
    subs_by_company_pkg = {}  # (company_lower, package_id) → Subcontractor
    for s in existing_subs.values():
        subs_by_company_pkg[(s.company.lower(), s.package_id)] = s

    result = []

    # ── Subcontractors sheet ──
    ws = wb["Subcontractors"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header

    # Track new subs being created here, keyed by (company_lower, package_tag_lower)
    new_subs_index = {}

    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        vals += [None] * max(0, len(SUB_COLS) - len(vals))

        errors, warnings = [], []
        raw_id = vals[0]
        try:
            sub_id = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            sub_id = None

        action = "UPDATE" if sub_id is not None else "CREATE"

        package_tag    = str(vals[1] or "").strip()
        company        = str(vals[2] or "").strip()
        contact_person = str(vals[3] or "").strip()
        phone          = str(vals[4] or "").strip()
        email          = str(vals[5] or "").strip()
        description    = str(vals[6] or "").strip()

        pkg = packages.get(package_tag.lower()) if package_tag else None

        if sub_id is not None and sub_id not in existing_subs:
            errors.append(f"Subcontractor ID {sub_id} not found in this project")
        if not company:
            errors.append("Company is required")
        if not package_tag:
            errors.append("Package (tag) is required")
        elif not pkg:
            errors.append(f"Package '{package_tag}' not found")

        if pkg and company:
            new_subs_index[(company.lower(), pkg.tag_number.lower())] = True

        result.append({
            "kind": "subcontractor",
            "row_num": row_idx,
            "id": sub_id,
            "package_tag": package_tag,
            "package_id": pkg.id if pkg else None,
            "company": company,
            "contact_person": contact_person,
            "phone": phone,
            "email": email,
            "description": description,
            "action": action,
            "errors": errors,
            "warnings": warnings,
        })

    # ── Workers sheet ──
    wws = wb["Workers"]
    rows_iter = wws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header

    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        vals += [None] * max(0, len(WORKER_COLS) - len(vals))

        errors, warnings = [], []
        raw_id = vals[0]
        try:
            worker_id = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            worker_id = None

        action = "UPDATE" if worker_id is not None else "CREATE"

        package_tag           = str(vals[1] or "").strip()
        name                  = str(vals[2] or "").strip()
        phone                 = str(vals[3] or "").strip()
        is_sub_raw            = str(vals[4] or "").strip().upper()
        subcontractor_company = str(vals[5] or "").strip()

        is_subcontractor = is_sub_raw in ("Y", "YES", "TRUE", "1")

        pkg = packages.get(package_tag.lower()) if package_tag else None

        if worker_id is not None and worker_id not in existing_workers:
            errors.append(f"Worker ID {worker_id} not found in this project")
        if not name:
            errors.append("Name is required")
        if not package_tag:
            errors.append("Package (tag) is required")
        elif not pkg:
            errors.append(f"Package '{package_tag}' not found")
        if is_subcontractor:
            if not subcontractor_company:
                errors.append("Subcontractor Company is required when 'Is Subcontractor' is Y")
            elif pkg:
                # Valid if existing in DB matching this package, or in new_subs_index for this package
                key_existing = (subcontractor_company.lower(), pkg.id)
                key_new      = (subcontractor_company.lower(), pkg.tag_number.lower())
                if key_existing not in subs_by_company_pkg and key_new not in new_subs_index:
                    errors.append(
                        f"Subcontractor '{subcontractor_company}' not found on package "
                        f"'{pkg.tag_number}' (neither in DB nor in this import's Subcontractors sheet)"
                    )

        result.append({
            "kind": "worker",
            "row_num": row_idx,
            "id": worker_id,
            "package_tag": package_tag,
            "package_id": pkg.id if pkg else None,
            "name": name,
            "phone": phone,
            "is_subcontractor": is_subcontractor,
            "subcontractor_company": subcontractor_company,
            "action": action,
            "errors": errors,
            "warnings": warnings,
        })

    return result


@router.post("/workers-subs/preview")
async def preview_workers_subs_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import workers/subcontractors")
    file_bytes = await file.read()
    rows = _parse_workers_subs_rows(file_bytes, db, user.project_id)
    creates = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {
        "rows": rows,
        "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count},
    }


@router.post("/workers-subs/apply")
def apply_workers_subs_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import workers/subcontractors")
    rows = payload.get("rows", [])
    now = datetime.utcnow()
    created = updated = skipped = 0

    # Process subcontractors FIRST so workers that reference newly-created subs
    # can resolve them by (company, package_id).
    sub_rows     = [r for r in rows if r.get("kind") == "subcontractor"]
    worker_rows  = [r for r in rows if r.get("kind") == "worker"]

    # Build lookup of ALL subcontractors in-project (rebuilt after sub inserts).
    def _subs_by_company_pkg():
        return {
            (s.company.lower(), s.package_id): s
            for s in db.query(models.Subcontractor).filter_by(project_id=user.project_id).all()
        }

    for row in sub_rows:
        if row.get("errors"):
            skipped += 1
            continue
        if row["action"] == "UPDATE" and row.get("id"):
            sub = db.query(models.Subcontractor).filter_by(id=row["id"], project_id=user.project_id).first()
            if not sub:
                skipped += 1
                continue
            sub.package_id     = row.get("package_id") or sub.package_id
            sub.company        = row["company"]
            sub.contact_person = row.get("contact_person") or None
            sub.phone          = row.get("phone") or None
            sub.email          = row.get("email") or None
            sub.description    = row.get("description") or None
            sub.updated_at     = now
            sub.updated_by_id  = user.id
            updated += 1
        else:
            if not row.get("package_id"):
                skipped += 1
                continue
            db.add(models.Subcontractor(
                project_id     = user.project_id,
                project_seq_id = models.next_project_seq(db, models.Subcontractor, user.project_id),
                package_id     = row["package_id"],
                company        = row["company"],
                contact_person = row.get("contact_person") or None,
                phone          = row.get("phone") or None,
                email          = row.get("email") or None,
                description    = row.get("description") or None,
                created_at     = now,
                created_by_id  = user.id,
            ))
            created += 1

    # Flush so newly-created subs are visible in the subsequent lookup.
    db.flush()
    subs_idx = _subs_by_company_pkg()

    for row in worker_rows:
        if row.get("errors"):
            skipped += 1
            continue

        sub_id = None
        if row.get("is_subcontractor") and row.get("subcontractor_company") and row.get("package_id"):
            sub = subs_idx.get((row["subcontractor_company"].lower(), row["package_id"]))
            sub_id = sub.id if sub else None
            if not sub_id:
                # Should have been caught in preview, but guard anyway.
                skipped += 1
                continue

        if row["action"] == "UPDATE" and row.get("id"):
            w = db.query(models.Worker).filter_by(id=row["id"], project_id=user.project_id).first()
            if not w:
                skipped += 1
                continue
            w.package_id        = row.get("package_id") or w.package_id
            w.name              = row["name"]
            w.phone             = row.get("phone") or None
            w.is_subcontractor  = bool(row.get("is_subcontractor"))
            w.subcontractor_id  = sub_id if row.get("is_subcontractor") else None
            w.updated_at        = now
            w.updated_by_id     = user.id
            updated += 1
        else:
            if not row.get("package_id"):
                skipped += 1
                continue
            db.add(models.Worker(
                project_id        = user.project_id,
                project_seq_id    = models.next_project_seq(db, models.Worker, user.project_id),
                package_id        = row["package_id"],
                name              = row["name"],
                phone             = row.get("phone") or None,
                is_subcontractor  = bool(row.get("is_subcontractor")),
                subcontractor_id  = sub_id if row.get("is_subcontractor") else None,
                status            = "PENDING",
                submitted_at      = now,
                created_at        = now,
                created_by_id     = user.id,
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ─────────────────────────────────────────────────────────────────────────────
# INVOICE IMPORT
# ─────────────────────────────────────────────────────────────────────────────

INVOICE_COLS = [
    ("id",              "ID"),
    ("invoice_number",  "Invoice Number"),
    ("po_number",       "Linked Order (PO Number)"),
    ("invoice_date",    "Invoice Date"),
    ("amount",          "Amount"),
    ("description",     "Description"),
]


def _build_invoice_workbook(db: Session, project_id: int) -> "Workbook":
    """Build an Excel workbook with existing invoices + Lookups sheet with orders."""
    invoices = (
        db.query(models.Invoice)
        .join(models.Order, models.Invoice.order_id == models.Order.id)
        .filter(models.Order.package_id.in_(
            db.query(models.Package.id).filter_by(project_id=project_id)
        ))
        .order_by(models.Invoice.id)
        .all()
    )
    orders = (
        db.query(models.Order)
        .join(models.Package, models.Order.package_id == models.Package.id)
        .filter(
            models.Package.project_id == project_id,
            models.Order.status.notin_(["CANCELLED", "DRAFT"]),
        )
        .order_by(models.Order.po_number)
        .all()
    )

    wb = Workbook()

    # ── Sheet 1: Invoices ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Invoices"

    headers = [h for _, h in INVOICE_COLS]
    ws.append(headers)
    _style_header(ws)

    for inv in invoices:
        ws.append([
            inv.id,
            inv.invoice_number,
            inv.order.po_number if inv.order else "",
            inv.invoice_date or "",
            inv.amount,
            inv.description or "",
        ])

    _autowidth(ws)

    # ── Sheet 2: Lookups ──────────────────────────────────────────────────────
    lk = wb.create_sheet("Lookups")
    row = 1

    lk.cell(row=row, column=1, value="ORDERS (PO Number)")
    lk.cell(row=row, column=2, value="Package")
    lk.cell(row=row, column=3, value="Vendor")
    lk.cell(row=row, column=4, value="PO Amount")
    lk.cell(row=row, column=5, value="Currency")
    _style_section_header(lk, row)
    row += 1
    for o in orders:
        pkg_tag = o.package.tag_number if o.package else ""
        pkg_name = o.package.name if o.package else ""
        lk.cell(row=row, column=1, value=o.po_number)
        lk.cell(row=row, column=2, value=f"{pkg_tag} — {pkg_name}")
        lk.cell(row=row, column=3, value=o.vendor_name or "")
        lk.cell(row=row, column=4, value=o.amount)
        lk.cell(row=row, column=5, value=o.currency or "EUR")
        row += 1

    row += 1

    lk.cell(row=row, column=1, value="VALID VALUES")
    _style_section_header(lk, row)
    row += 1
    lk.cell(row=row, column=1, value="Field")
    lk.cell(row=row, column=2, value="Allowed values")
    _style_header(lk, row)
    row += 1
    lk.cell(row=row, column=1, value="Note")
    lk.cell(row=row, column=2, value="Status is NOT imported — all imported invoices start as DRAFT")
    row += 1
    lk.cell(row=row, column=1, value="ID")
    lk.cell(row=row, column=2, value="Leave blank to create new; fill in existing ID to update")
    row += 1
    lk.cell(row=row, column=1, value="Linked Order (PO Number)")
    lk.cell(row=row, column=2, value="Must match a PO Number from the ORDERS list above")
    row += 1

    _autowidth(lk)

    return wb


def _parse_invoice_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    """Parse uploaded Excel rows, validate against project orders, return preview rows."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Invoices" not in wb.sheetnames:
        raise HTTPException(400, "Workbook must contain an 'Invoices' sheet")
    ws = wb["Invoices"]

    orders = (
        db.query(models.Order)
        .join(models.Package, models.Order.package_id == models.Package.id)
        .filter(models.Package.project_id == project_id)
        .all()
    )
    order_by_po = {o.po_number.strip().lower(): o for o in orders}

    existing_invoices = {}
    for o in orders:
        for inv in o.invoices:
            existing_invoices[inv.id] = inv

    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row) + [None] * 10
        if all(v is None or str(v).strip() == "" for v in vals[:6]):
            continue

        raw_id = vals[0]
        inv_id = None
        if raw_id is not None:
            try:
                inv_id = int(raw_id)
            except (ValueError, TypeError):
                pass

        action = "UPDATE" if inv_id else "CREATE"

        invoice_number = str(vals[1] or "").strip()
        po_number      = str(vals[2] or "").strip()
        invoice_date   = _parse_date(vals[3])
        raw_amount     = vals[4]
        description    = str(vals[5] or "").strip()

        amount = 0.0
        if raw_amount is not None:
            try:
                amount = float(raw_amount)
            except (ValueError, TypeError):
                pass

        errors = []
        warnings = []

        if inv_id and inv_id not in existing_invoices:
            errors.append(f"Invoice ID {inv_id} not found in this project")

        if not invoice_number:
            errors.append("Invoice Number is required")
        if not po_number:
            errors.append("Linked Order (PO Number) is required")
        if not invoice_date:
            errors.append("Invoice Date is required")

        order_match = order_by_po.get(po_number.lower()) if po_number else None
        if po_number and not order_match:
            errors.append(f"PO Number '{po_number}' not found in this project. Check the Lookups sheet for valid values.")
        elif order_match and order_match.status in ("CANCELLED", "DRAFT"):
            warnings.append(f"PO '{po_number}' has status {order_match.status}")

        package_tag = ""
        if order_match and order_match.package:
            package_tag = order_match.package.tag_number

        rows.append({
            "row_num":        idx,
            "id":             inv_id,
            "invoice_number": invoice_number,
            "po_number":      po_number,
            "package_tag":    package_tag,
            "invoice_date":   invoice_date,
            "amount":         amount,
            "description":    description,
            "action":         action,
            "errors":         errors,
            "warnings":       warnings,
        })

    return rows


@router.get("/invoices/export")
def export_invoices_template(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    wb = _build_invoice_workbook(db, user.project_id)
    return _xlsx_response(wb, "invoices_import_template.xlsx")


@router.post("/invoices/preview")
async def preview_invoices_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import invoices")
    file_bytes = await file.read()
    rows      = _parse_invoice_rows(file_bytes, db, user.project_id)
    creates   = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates   = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {
        "rows": rows,
        "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count},
    }


@router.post("/invoices/apply")
def apply_invoices_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    if user.role not in ("ADMIN", "PROJECT_OWNER"):
        raise HTTPException(403, "Only ADMIN or PROJECT_OWNER can import invoices")

    rows = payload.get("rows", [])

    orders = (
        db.query(models.Order)
        .join(models.Package, models.Order.package_id == models.Package.id)
        .filter(models.Package.project_id == user.project_id)
        .all()
    )
    order_by_po = {o.po_number.strip().lower(): o for o in orders}

    now = datetime.utcnow()
    created = updated = skipped = 0

    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue

        order = order_by_po.get(row["po_number"].strip().lower()) if row.get("po_number") else None
        if not order:
            skipped += 1
            continue

        if row["action"] == "UPDATE" and row.get("id"):
            inv = db.query(models.Invoice).filter_by(id=row["id"]).first()
            if not inv:
                skipped += 1
                continue
            inv.invoice_number = row["invoice_number"]
            inv.order_id       = order.id
            inv.package_id     = order.package_id
            inv.invoice_date   = row.get("invoice_date")
            inv.amount         = row.get("amount", 0.0)
            inv.description    = row.get("description")
            inv.updated_at     = now
            inv.updated_by_id  = user.id
            updated += 1
        else:
            db.add(models.Invoice(
                order_id        = order.id,
                package_id      = order.package_id,
                invoice_number  = row["invoice_number"],
                description     = row.get("description"),
                amount          = row.get("amount", 0.0),
                currency        = order.currency or "EUR",
                invoice_date    = row.get("invoice_date"),
                status          = "DRAFT",
                pmc_approved    = False,
                client_approved = False,
                created_by_id   = user.id,
                created_at      = now,
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ─────────────────────────────────────────────────────────────────────────────
# PROJECT ORGANIZATION IMPORTS — Contacts, Subservices, Areas, Units
# Follow the same Export-template / Preview / Apply flow as Risks.
# All gated to ADMIN or PROJECT_OWNER.
# ─────────────────────────────────────────────────────────────────────────────

_ORG_ROLES = ("ADMIN", "PROJECT_OWNER")


def _check_org_role(user: "auth.ProjectContext"):
    if user.role not in _ORG_ROLES:
        raise HTTPException(403, "Only Project Owners can import organization data")


# ── Contacts ──────────────────────────────────────────────────────────────────

CONTACT_COLS = [
    ("id",       "ID (leave blank to create)"),
    ("name",     "Name *"),
    ("company",  "Company"),
    ("function", "Function"),
    ("email",    "Email"),
    ("phone",    "Phone"),
]


def _build_contacts_workbook(db: Session, project_id: int) -> "Workbook":
    contacts = (
        db.query(models.Contact)
        .filter_by(project_id=project_id)
        .order_by(models.Contact.name)
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append([c[1] for c in CONTACT_COLS])
    _style_header(ws, 1)
    for c in contacts:
        ws.append([c.id, c.name or "", c.company or "", c.function or "", c.email or "", c.phone or ""])
    _autowidth(ws)

    notes = wb.create_sheet("Notes")
    notes.append(["Field", "Notes"])
    _style_header(notes, 1)
    notes.append(["ID", "Leave blank to create a new contact. Fill in to update an existing one."])
    notes.append(["Name", "Required."])
    notes.append(["Email", "If set on a new contact, must be unique within the project."])
    _autowidth(notes)
    return wb


def _parse_contact_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Contacts" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Contacts' not found in the uploaded file")
    ws = wb["Contacts"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)

    existing = {c.id: c for c in db.query(models.Contact).filter_by(project_id=project_id).all()}
    emails = {(c.email or "").strip().lower(): c for c in existing.values() if c.email}

    result = []
    seen_emails = set()
    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row) + [None] * max(0, len(CONTACT_COLS) - len(row or []))
        if all(v is None or str(v).strip() == "" for v in vals):
            continue

        errors, warnings = [], []
        raw_id = vals[0]
        try:
            cid = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            cid = None
        action = "UPDATE" if cid is not None else "CREATE"

        name     = str(vals[1] or "").strip()
        company  = str(vals[2] or "").strip()
        function = str(vals[3] or "").strip()
        email    = str(vals[4] or "").strip()
        phone    = str(vals[5] or "").strip()

        if cid is not None and cid not in existing:
            errors.append(f"Contact ID {cid} not found in this project")
        if not name:
            errors.append("Name is required")
        if email:
            key = email.lower()
            if action == "CREATE" and key in emails:
                errors.append(f"A contact with email '{email}' already exists")
            elif action == "UPDATE" and key in emails and emails[key].id != cid:
                errors.append(f"Email '{email}' is already used by another contact")
            if key in seen_emails:
                errors.append(f"Email '{email}' is duplicated within the file")
            seen_emails.add(key)

        result.append({
            "row_num": row_idx, "id": cid, "action": action,
            "name": name, "company": company or None, "function": function or None,
            "email": email or None, "phone": phone or None,
            "errors": errors, "warnings": warnings,
        })
    return result


@router.get("/contacts/export")
def export_contacts_template(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    return _xlsx_response(_build_contacts_workbook(db, user.project_id), "contacts_import_template.xlsx")


@router.post("/contacts/preview")
async def preview_contacts_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    _check_org_role(user)
    rows = _parse_contact_rows(await file.read(), db, user.project_id)
    creates = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {"rows": rows, "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count}}


@router.post("/contacts/apply")
def apply_contacts_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_org_role(user)
    rows = payload.get("rows", [])
    now = datetime.utcnow()
    created = updated = skipped = 0
    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue
        if row["action"] == "UPDATE" and row.get("id"):
            c = db.query(models.Contact).filter_by(id=row["id"], project_id=user.project_id).first()
            if not c:
                skipped += 1
                continue
            c.name     = row["name"]
            c.company  = row.get("company")
            c.function = row.get("function")
            c.email    = row.get("email")
            c.phone    = row.get("phone")
            c.updated_at = now
            c.updated_by_id = user.id
            updated += 1
        else:
            db.add(models.Contact(
                project_id=user.project_id,
                name=row["name"], company=row.get("company"), function=row.get("function"),
                email=row.get("email"), phone=row.get("phone"),
                created_at=now, created_by_id=user.id,
            ))
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ── Subservices ───────────────────────────────────────────────────────────────

SUBSERVICE_COLS = [
    ("id",              "ID (leave blank to create)"),
    ("service_code",    "Service Code *"),
    ("service_name",    "Service Name *"),
    ("subservice_code", "Subservice Code *"),
    ("subservice_name", "Subservice Name *"),
    ("pmc_reviewer",    "PMC Reviewer (contact name)"),
    ("client_reviewer", "Client Reviewer (contact name)"),
    ("sort_order",      "Sort Order"),
]


def _build_subservices_workbook(db: Session, project_id: int) -> "Workbook":
    subs = (
        db.query(models.Subservice)
        .filter_by(project_id=project_id)
        .order_by(models.Subservice.service_code, models.Subservice.sort_order)
        .all()
    )
    contacts = (
        db.query(models.Contact)
        .filter_by(project_id=project_id)
        .order_by(models.Contact.name)
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Subservices"
    ws.append([c[1] for c in SUBSERVICE_COLS])
    _style_header(ws, 1)
    for s in subs:
        ws.append([
            s.id, s.service_code, s.service_name, s.subservice_code, s.subservice_name,
            s.pmc_reviewer.name if s.pmc_reviewer else "",
            s.client_reviewer.name if s.client_reviewer else "",
            s.sort_order or 0,
        ])
    _autowidth(ws)

    lws = wb.create_sheet("Lookups")
    lws.append(["CONTACTS (use exact name for reviewer columns)"])
    _style_section_header(lws, 1)
    lws.append(["name"])
    for cell in lws[2]:
        if cell.value:
            cell.font = Font(bold=True)
    for ct in contacts:
        lws.append([ct.name])
    _autowidth(lws)
    return wb


def _parse_subservice_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Subservices" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Subservices' not found in the uploaded file")
    ws = wb["Subservices"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)

    existing = {s.id: s for s in db.query(models.Subservice).filter_by(project_id=project_id).all()}
    contacts = {c.name.lower(): c for c in db.query(models.Contact).filter_by(project_id=project_id).all()}

    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row) + [None] * max(0, len(SUBSERVICE_COLS) - len(row or []))
        if all(v is None or str(v).strip() == "" for v in vals):
            continue

        errors, warnings = [], []
        raw_id = vals[0]
        try:
            sid = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            sid = None
        action = "UPDATE" if sid is not None else "CREATE"

        service_code    = str(vals[1] or "").strip()
        service_name    = str(vals[2] or "").strip()
        subservice_code = str(vals[3] or "").strip()
        subservice_name = str(vals[4] or "").strip()
        pmc_name        = str(vals[5] or "").strip()
        client_name     = str(vals[6] or "").strip()
        try:
            sort_order = int(vals[7]) if vals[7] is not None and str(vals[7]).strip() != "" else 0
        except (ValueError, TypeError):
            sort_order = 0

        if sid is not None and sid not in existing:
            errors.append(f"Subservice ID {sid} not found in this project")
        if not service_code:    errors.append("Service Code is required")
        if not service_name:    errors.append("Service Name is required")
        if not subservice_code: errors.append("Subservice Code is required")
        if not subservice_name: errors.append("Subservice Name is required")
        if pmc_name and pmc_name.lower() not in contacts:
            errors.append(f"PMC Reviewer '{pmc_name}' not found in contacts")
        if client_name and client_name.lower() not in contacts:
            errors.append(f"Client Reviewer '{client_name}' not found in contacts")

        result.append({
            "row_num": row_idx, "id": sid, "action": action,
            "service_code": service_code, "service_name": service_name,
            "subservice_code": subservice_code, "subservice_name": subservice_name,
            "pmc_reviewer_name": pmc_name or None,
            "client_reviewer_name": client_name or None,
            "sort_order": sort_order,
            "errors": errors, "warnings": warnings,
        })
    return result


@router.get("/subservices/export")
def export_subservices_template(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    return _xlsx_response(_build_subservices_workbook(db, user.project_id), "subservices_import_template.xlsx")


@router.post("/subservices/preview")
async def preview_subservices_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    _check_org_role(user)
    rows = _parse_subservice_rows(await file.read(), db, user.project_id)
    creates = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {"rows": rows, "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count}}


@router.post("/subservices/apply")
def apply_subservices_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_org_role(user)
    rows = payload.get("rows", [])
    contacts = {c.name.lower(): c for c in db.query(models.Contact).filter_by(project_id=user.project_id).all()}
    created = updated = skipped = 0
    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue
        pmc    = contacts.get((row.get("pmc_reviewer_name") or "").lower()) if row.get("pmc_reviewer_name") else None
        client = contacts.get((row.get("client_reviewer_name") or "").lower()) if row.get("client_reviewer_name") else None
        if row["action"] == "UPDATE" and row.get("id"):
            s = db.query(models.Subservice).filter_by(id=row["id"], project_id=user.project_id).first()
            if not s:
                skipped += 1
                continue
            s.service_code       = row["service_code"]
            s.service_name       = row["service_name"]
            s.subservice_code    = row["subservice_code"]
            s.subservice_name    = row["subservice_name"]
            s.pmc_reviewer_id    = pmc.id if pmc else None
            s.client_reviewer_id = client.id if client else None
            s.sort_order         = row.get("sort_order") or 0
            updated += 1
        else:
            db.add(models.Subservice(
                project_id=user.project_id,
                service_code=row["service_code"], service_name=row["service_name"],
                subservice_code=row["subservice_code"], subservice_name=row["subservice_name"],
                pmc_reviewer_id=pmc.id if pmc else None,
                client_reviewer_id=client.id if client else None,
                sort_order=row.get("sort_order") or 0,
            ))
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ── Areas ─────────────────────────────────────────────────────────────────────

AREA_COLS = [
    ("id",                "ID (leave blank to create)"),
    ("tag",               "Tag *"),
    ("description",       "Description *"),
    ("details",           "Details"),
    ("owner",             "Owner (contact name)"),
    ("site_supervisors",  "Site Supervisors (comma-separated contact names)"),
]


def _eligible_supervisor_ids(db: Session, project_id: int) -> set:
    """Contacts whose linked user is ADMIN, PROJECT_OWNER, PROJECT_TEAM or CLIENT
    on this project — same rule as routers/areas_units._is_eligible_supervisor."""
    rows = (
        db.query(models.Contact, models.User, models.UserProject)
          .join(models.User, models.User.contact_id == models.Contact.id)
          .outerjoin(
              models.UserProject,
              (models.UserProject.user_id == models.User.id)
              & (models.UserProject.project_id == project_id),
          )
          .filter(models.Contact.project_id == project_id)
          .all()
    )
    out = set()
    for (c, u, up) in rows:
        project_role = (up.role if up else u.role) if u else None
        if u and (u.role == "ADMIN" or project_role in {"PROJECT_OWNER", "PROJECT_TEAM", "CLIENT"}):
            out.add(c.id)
    return out


def _build_areas_workbook(db: Session, project_id: int) -> "Workbook":
    areas = (
        db.query(models.Area).filter_by(project_id=project_id).order_by(models.Area.tag).all()
    )
    contacts = (
        db.query(models.Contact).filter_by(project_id=project_id).order_by(models.Contact.name).all()
    )
    eligible = _eligible_supervisor_ids(db, project_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Areas"
    ws.append([c[1] for c in AREA_COLS])
    _style_header(ws, 1)
    for a in areas:
        sups = ", ".join((s.contact.name for s in (a.site_supervisors or []) if s.contact))
        ws.append([a.id, a.tag, a.description, a.details or "", a.owner.name if a.owner else "", sups])
    _autowidth(ws)

    lws = wb.create_sheet("Lookups")
    lws.append(["OWNERS (any contact)"])
    _style_section_header(lws, 1)
    lws.append(["name"])
    for cell in lws[2]:
        if cell.value: cell.font = Font(bold=True)
    for ct in contacts:
        lws.append([ct.name])

    lws.append([])
    off = len(contacts) + 4
    lws.append(["ELIGIBLE SITE SUPERVISORS (Project Owners / Project Team / Client)"])
    _style_section_header(lws, off)
    lws.append(["name"])
    for cell in lws[off + 1]:
        if cell.value: cell.font = Font(bold=True)
    for ct in contacts:
        if ct.id in eligible:
            lws.append([ct.name])
    _autowidth(lws)
    return wb


def _parse_area_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Areas" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Areas' not found in the uploaded file")
    ws = wb["Areas"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)

    existing = {a.id: a for a in db.query(models.Area).filter_by(project_id=project_id).all()}
    contacts = {c.name.lower(): c for c in db.query(models.Contact).filter_by(project_id=project_id).all()}
    eligible = _eligible_supervisor_ids(db, project_id)

    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row) + [None] * max(0, len(AREA_COLS) - len(row or []))
        if all(v is None or str(v).strip() == "" for v in vals):
            continue

        errors, warnings = [], []
        raw_id = vals[0]
        try:
            aid = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            aid = None
        action = "UPDATE" if aid is not None else "CREATE"

        tag         = str(vals[1] or "").strip()
        description = str(vals[2] or "").strip()
        details     = str(vals[3] or "").strip() or None
        owner_name  = str(vals[4] or "").strip()
        sups_str    = str(vals[5] or "").strip()

        if aid is not None and aid not in existing:
            errors.append(f"Area ID {aid} not found in this project")
        if not tag:         errors.append("Tag is required")
        if not description: errors.append("Description is required")
        if owner_name and owner_name.lower() not in contacts:
            errors.append(f"Owner '{owner_name}' not found in contacts")

        sup_names = [s.strip() for s in sups_str.split(",") if s.strip()] if sups_str else []
        sup_ids = []
        for nm in sup_names:
            ct = contacts.get(nm.lower())
            if not ct:
                errors.append(f"Site Supervisor '{nm}' not found in contacts")
                continue
            if ct.id not in eligible:
                errors.append(f"Site Supervisor '{nm}' is not an eligible supervisor (must be Project Owner / Project Team / Client)")
                continue
            sup_ids.append(ct.id)

        result.append({
            "row_num": row_idx, "id": aid, "action": action,
            "tag": tag, "description": description, "details": details,
            "owner_name": owner_name or None,
            "site_supervisor_names": sup_names,
            "site_supervisor_ids": sup_ids,
            "errors": errors, "warnings": warnings,
        })
    return result


@router.get("/areas/export")
def export_areas_template(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    return _xlsx_response(_build_areas_workbook(db, user.project_id), "areas_import_template.xlsx")


@router.post("/areas/preview")
async def preview_areas_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    _check_org_role(user)
    rows = _parse_area_rows(await file.read(), db, user.project_id)
    creates = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {"rows": rows, "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count}}


@router.post("/areas/apply")
def apply_areas_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_org_role(user)
    rows = payload.get("rows", [])
    contacts = {c.name.lower(): c for c in db.query(models.Contact).filter_by(project_id=user.project_id).all()}
    now = datetime.utcnow()
    created = updated = skipped = 0
    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue
        owner = contacts.get((row.get("owner_name") or "").lower()) if row.get("owner_name") else None
        sup_ids = row.get("site_supervisor_ids") or []

        if row["action"] == "UPDATE" and row.get("id"):
            a = db.query(models.Area).filter_by(id=row["id"], project_id=user.project_id).first()
            if not a:
                skipped += 1
                continue
            a.tag = row["tag"]
            a.description = row["description"]
            a.details = row.get("details")
            a.owner_id = owner.id if owner else None
            a.updated_at = now
            a.updated_by_id = user.id
            db.query(models.AreaSiteSupervisor).filter_by(area_id=a.id).delete()
            db.flush()
            for sid in set(sup_ids):
                db.add(models.AreaSiteSupervisor(area_id=a.id, contact_id=sid))
            updated += 1
        else:
            a = models.Area(
                project_id=user.project_id,
                tag=row["tag"], description=row["description"], details=row.get("details"),
                owner_id=owner.id if owner else None,
                created_at=now, created_by_id=user.id,
            )
            db.add(a)
            db.flush()
            for sid in set(sup_ids):
                db.add(models.AreaSiteSupervisor(area_id=a.id, contact_id=sid))
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}


# ── Units ─────────────────────────────────────────────────────────────────────

UNIT_COLS = [
    ("id",          "ID (leave blank to create)"),
    ("tag",         "Tag *"),
    ("description", "Description *"),
    ("details",     "Details"),
    ("owner",       "Owner (contact name)"),
]


def _build_units_workbook(db: Session, project_id: int) -> "Workbook":
    units = db.query(models.Unit).filter_by(project_id=project_id).order_by(models.Unit.tag).all()
    contacts = db.query(models.Contact).filter_by(project_id=project_id).order_by(models.Contact.name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Units"
    ws.append([c[1] for c in UNIT_COLS])
    _style_header(ws, 1)
    for u in units:
        ws.append([u.id, u.tag, u.description, u.details or "", u.owner.name if u.owner else ""])
    _autowidth(ws)

    lws = wb.create_sheet("Lookups")
    lws.append(["OWNERS (any contact)"])
    _style_section_header(lws, 1)
    lws.append(["name"])
    for cell in lws[2]:
        if cell.value: cell.font = Font(bold=True)
    for ct in contacts:
        lws.append([ct.name])
    _autowidth(lws)
    return wb


def _parse_unit_rows(file_bytes: bytes, db: Session, project_id: int) -> list:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Units" not in wb.sheetnames:
        raise HTTPException(400, "Sheet 'Units' not found in the uploaded file")
    ws = wb["Units"]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)

    existing = {u.id: u for u in db.query(models.Unit).filter_by(project_id=project_id).all()}
    contacts = {c.name.lower(): c for c in db.query(models.Contact).filter_by(project_id=project_id).all()}

    result = []
    for row_idx, row in enumerate(rows_iter, start=2):
        vals = list(row) + [None] * max(0, len(UNIT_COLS) - len(row or []))
        if all(v is None or str(v).strip() == "" for v in vals):
            continue

        errors, warnings = [], []
        raw_id = vals[0]
        try:
            uid = int(raw_id) if raw_id is not None and str(raw_id).strip() != "" else None
        except (ValueError, TypeError):
            uid = None
        action = "UPDATE" if uid is not None else "CREATE"

        tag         = str(vals[1] or "").strip()
        description = str(vals[2] or "").strip()
        details     = str(vals[3] or "").strip() or None
        owner_name  = str(vals[4] or "").strip()

        if uid is not None and uid not in existing:
            errors.append(f"Unit ID {uid} not found in this project")
        if not tag:         errors.append("Tag is required")
        if not description: errors.append("Description is required")
        if owner_name and owner_name.lower() not in contacts:
            errors.append(f"Owner '{owner_name}' not found in contacts")

        result.append({
            "row_num": row_idx, "id": uid, "action": action,
            "tag": tag, "description": description, "details": details,
            "owner_name": owner_name or None,
            "errors": errors, "warnings": warnings,
        })
    return result


@router.get("/units/export")
def export_units_template(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    return _xlsx_response(_build_units_workbook(db, user.project_id), "units_import_template.xlsx")


@router.post("/units/preview")
async def preview_units_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_openpyxl()
    _check_org_role(user)
    rows = _parse_unit_rows(await file.read(), db, user.project_id)
    creates = sum(1 for r in rows if r["action"] == "CREATE" and not r["errors"])
    updates = sum(1 for r in rows if r["action"] == "UPDATE" and not r["errors"])
    err_count = sum(1 for r in rows if r["errors"])
    return {"rows": rows, "summary": {"total": len(rows), "creates": creates, "updates": updates, "errors": err_count}}


@router.post("/units/apply")
def apply_units_import(
    payload: dict,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _check_org_role(user)
    rows = payload.get("rows", [])
    contacts = {c.name.lower(): c for c in db.query(models.Contact).filter_by(project_id=user.project_id).all()}
    now = datetime.utcnow()
    created = updated = skipped = 0
    for row in rows:
        if row.get("errors"):
            skipped += 1
            continue
        owner = contacts.get((row.get("owner_name") or "").lower()) if row.get("owner_name") else None
        if row["action"] == "UPDATE" and row.get("id"):
            u = db.query(models.Unit).filter_by(id=row["id"], project_id=user.project_id).first()
            if not u:
                skipped += 1
                continue
            u.tag = row["tag"]
            u.description = row["description"]
            u.details = row.get("details")
            u.owner_id = owner.id if owner else None
            u.updated_at = now
            u.updated_by_id = user.id
            updated += 1
        else:
            db.add(models.Unit(
                project_id=user.project_id,
                tag=row["tag"], description=row["description"], details=row.get("details"),
                owner_id=owner.id if owner else None,
                created_at=now, created_by_id=user.id,
            ))
            created += 1
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped}
