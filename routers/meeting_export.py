import io
import os
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
import models
import auth

from fastapi import Query as FQuery

router = APIRouter(tags=["meeting-export"])

LOGO_PATH = os.path.join("static", "assets", "impulse-logo-light@2x.png")

_DARK_BLUE  = "1E3A5F"
_LIGHT_BLUE = "D6E4F0"
_WHITE      = "FFFFFF"

_STATUS_LABELS = {
    "NOT_STARTED": "Not Started",
    "IN_PROGRESS": "In Progress",
    "ON_HOLD":     "On Hold",
    "URGENT":      "Urgent",
    "CLOSED":      "Closed",
    "PLANNED":     "Planned",
    "COMPLETED":   "Completed",
    "CANCELLED":   "Cancelled",
}


def _safe_filename(title: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in title)


def _get_meeting_and_points(meeting_id: int, user, db: Session, selected_only: bool = False):
    meeting = db.query(models.Meeting).filter(
        models.Meeting.id == meeting_id,
        models.Meeting.project_id == user.project_id,
    ).first()
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    q = (
        db.query(models.MeetingPoint)
        .join(models.MeetingPointLink)
        .filter(models.MeetingPointLink.meeting_id == meeting_id)
    )
    if selected_only:
        q = q.filter(models.MeetingPointLink.for_preparation == True)
    points = q.order_by(models.MeetingPoint.project_seq_id).all()
    return meeting, points


def _format_notes(notes) -> str:
    """Concatenate notes as 'Author - Date\\nnote\\n\\nAuthor - Date\\nnote'."""
    parts = []
    for n in sorted(notes, key=lambda x: x.created_at or datetime.min):
        date_str = n.created_at.strftime("%Y-%m-%d") if n.created_at else ""
        author   = n.author.name if n.author else "?"
        parts.append(f"{author} - {date_str}\n{n.content or ''}")
    return "\n\n".join(parts)


# ── Excel ─────────────────────────────────────────────────────────────────────

@router.get("/api/meetings/{meeting_id}/export/excel")
def export_excel(
    meeting_id: int,
    selected_only: bool = FQuery(False),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    meeting, points = _get_meeting_and_points(meeting_id, user, db, selected_only=selected_only)

    wb = Workbook()

    # ── Sheet 1: Meeting Overview ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Meeting"

    ws1.merge_cells("A1:E1")
    c = ws1["A1"]
    c.value = meeting.title
    c.font = Font(bold=True, size=14, color=_WHITE)
    c.fill = PatternFill("solid", fgColor=_DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 30

    details = [
        ("Date",     str(meeting.date) if meeting.date else ""),
        ("Time",     meeting.time or ""),
        ("Location", meeting.location or ""),
        ("Type",     meeting.meeting_type.name if meeting.meeting_type else ""),
        ("Status",   _STATUS_LABELS.get(meeting.status, meeting.status or "")),
        ("Notes",    meeting.notes or ""),
    ]

    for i, (label, value) in enumerate(details, start=2):
        lc = ws1.cell(row=i, column=1, value=label)
        lc.font = Font(bold=True)
        lc.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
        vc = ws1.cell(row=i, column=2, value=value)
        vc.alignment = Alignment(wrap_text=True)
        ws1.merge_cells(f"B{i}:E{i}")

    next_row = len(details) + 3

    ph = ws1.cell(row=next_row, column=1, value="Participants")
    ph.font = Font(bold=True, size=11, color=_WHITE)
    ph.fill = PatternFill("solid", fgColor=_DARK_BLUE)
    ws1.merge_cells(f"A{next_row}:E{next_row}")
    next_row += 1

    for hdr, col in [("Name", 1), ("Company", 2), ("Attendance", 3)]:
        hc = ws1.cell(row=next_row, column=col, value=hdr)
        hc.font = Font(bold=True)
        hc.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
    next_row += 1

    for p in meeting.participants:
        contact = db.query(models.Contact).filter(models.Contact.id == p.contact_id).first()
        if contact:
            ws1.cell(row=next_row, column=1, value=contact.name)
            ws1.cell(row=next_row, column=2, value=contact.company or "")
            ws1.cell(row=next_row, column=3, value="Present" if p.present else "Absent")
            next_row += 1

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 35
    ws1.column_dimensions["C"].width = 18

    # ── Sheet 2: Meeting Points ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Points")

    pt_headers = ["ID", "Type", "Topic", "Details", "Notes", "Responsible", "Due Date", "Status", "Closed At"]
    pt_widths   = [12,   12,     32,      50,         50,      25,            14,         15,        20]

    for col, (h, w) in enumerate(zip(pt_headers, pt_widths), start=1):
        hc = ws2.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 20

    for ri, p in enumerate(points, start=2):
        responsible = (
            db.query(models.Contact).filter(models.Contact.id == p.responsible_id).first()
            if p.responsible_id else None
        )
        notes_text = _format_notes(p.notes)
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None

        row_vals = [
            f"MP-{str(p.project_seq_id or p.id).zfill(6)}",
            p.type,
            p.topic,
            p.details or "",
            notes_text,
            responsible.name if responsible else "",
            str(p.due_date) if p.due_date else "",
            _STATUS_LABELS.get(p.status, p.status),
            p.closed_at.strftime("%Y-%m-%d %H:%M") if p.closed_at else "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"meeting_{_safe_filename(meeting.title)}_{meeting.date or 'nodate'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── PDF ───────────────────────────────────────────────────────────────────────

@router.get("/api/meetings/{meeting_id}/export/pdf")
def export_pdf(
    meeting_id: int,
    selected_only: bool = FQuery(False),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from fpdf import FPDF, XPos, YPos

    meeting, points = _get_meeting_and_points(meeting_id, user, db, selected_only=selected_only)

    # Colours as (R, G, B)
    BLUE  = (30,  58,  95)
    LBLUE = (214, 228, 240)
    WHITE = (255, 255, 255)
    GRAY  = (100, 100, 100)
    BLACK = (30,  30,  30)
    RED   = (180, 30,  30)
    GREEN = (30,  130, 60)

    TYPE_COLORS = {
        "ACTION":   (124, 58,  237),
        "DECISION": (37,  99,  235),
        "INFO":     (100, 100, 100),
    }
    STATUS_COLORS = {
        "NOT_STARTED": (100, 100, 100),
        "IN_PROGRESS": (180, 120, 0),
        "ON_HOLD":     (37,  99,  235),
        "URGENT":      (180, 30,  30),
        "CLOSED":      (30,  130, 60),
    }

    # ── helper: line after cell ───────────────────────────────────────────────
    NL = dict(new_x=XPos.LMARGIN, new_y=YPos.NEXT)   # move to next line
    ST = dict(new_x=XPos.RIGHT,   new_y=YPos.TOP)     # stay on same line (default)

    class MeetingPDF(FPDF):
        def header(self):
            if os.path.exists(LOGO_PATH):
                self.image(LOGO_PATH, x=10, y=6, h=16)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(*GRAY)
            self.set_y(8)
            self.cell(0, 5, f"Page {self.page_no()}", align="R", **NL)
            self.ln(10)

        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(*GRAY)
            generated = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
            self.cell(0, 5, f"Generated: {generated}", align="C")

    pdf = MeetingPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    W = pdf.w - pdf.l_margin - pdf.r_margin  # usable width

    # ── convenience wrappers ──────────────────────────────────────────────────
    def tc(*rgb):
        pdf.set_text_color(*rgb)

    def fc(*rgb):
        pdf.set_fill_color(*rgb)

    def divider():
        pdf.set_draw_color(*LBLUE)
        pdf.set_line_width(0.4)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.l_margin + W, pdf.get_y())
        pdf.ln(3)

    # ── Title banner ──────────────────────────────────────────────────────────
    fc(*BLUE)
    tc(*WHITE)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(W, 10, meeting.title, fill=True, **NL)
    pdf.ln(3)

    # ── Meta grid (two columns) ───────────────────────────────────────────────
    half = W / 2

    def meta_pair(lbl1, v1, lbl2, v2):
        pdf.set_font("Helvetica", "B", 9)
        tc(*BLUE)
        pdf.cell(28, 5.5, lbl1 + ":", **ST)
        pdf.set_font("Helvetica", "", 9)
        tc(*BLACK)
        pdf.cell(half - 28, 5.5, v1 or "-", **ST)
        if lbl2:
            pdf.set_font("Helvetica", "B", 9)
            tc(*BLUE)
            pdf.cell(28, 5.5, lbl2 + ":", **ST)
            pdf.set_font("Helvetica", "", 9)
            tc(*BLACK)
            pdf.cell(half - 28, 5.5, v2 or "-", **NL)
        else:
            pdf.ln(5.5)

    meta_pair("Date",     str(meeting.date) if meeting.date else "",
              "Time",     meeting.time or "")
    meta_pair("Location", meeting.location or "",
              "Type",     meeting.meeting_type.name if meeting.meeting_type else "")
    meta_pair("Status",   _STATUS_LABELS.get(meeting.status, meeting.status or ""),
              "",         "")
    pdf.ln(2)

    if meeting.notes:
        pdf.set_font("Helvetica", "B", 9)
        tc(*BLUE)
        pdf.cell(28, 5, "Notes:", **ST)
        pdf.set_font("Helvetica", "", 9)
        tc(*BLACK)
        pdf.multi_cell(W - 28, 5, meeting.notes, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(1)

    divider()

    # ── Participants ──────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 11)
    tc(*BLUE)
    pdf.cell(W, 7, "Participants", **NL)
    pdf.ln(1)

    if meeting.participants:
        col_name = W * 0.45
        col_co   = W * 0.38
        col_att  = W * 0.17

        fc(*BLUE)
        tc(*WHITE)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(col_name, 6, "Name",       fill=True, **ST)
        pdf.cell(col_co,   6, "Company",    fill=True, **ST)
        pdf.cell(col_att,  6, "Attendance", fill=True, **NL)

        pdf.set_font("Helvetica", "", 8)
        for i, p in enumerate(meeting.participants):
            contact = db.query(models.Contact).filter(models.Contact.id == p.contact_id).first()
            if not contact:
                continue
            fc(*(LBLUE if i % 2 == 0 else WHITE))
            tc(*BLACK)
            pdf.cell(col_name, 5, contact.name,          fill=True, **ST)
            pdf.cell(col_co,   5, contact.company or "",  fill=True, **ST)
            tc(*(GREEN if p.present else GRAY))
            pdf.cell(col_att,  5, "Present" if p.present else "Absent", fill=True, **NL)
    else:
        pdf.set_font("Helvetica", "I", 9)
        tc(*GRAY)
        pdf.cell(W, 5, "No participants recorded.", **NL)

    pdf.ln(4)
    divider()

    # ── Meeting Points ────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 11)
    tc(*BLUE)
    pdf.cell(W, 7, f"Meeting Points  ({len(points)})", **NL)
    pdf.ln(2)

    if not points:
        pdf.set_font("Helvetica", "I", 9)
        tc(*GRAY)
        pdf.cell(W, 5, "No meeting points.", **NL)

    for p in points:
        if pdf.get_y() > pdf.h - 70:
            pdf.add_page()

        # Header bar
        fc(*LBLUE)
        bar_y = pdf.get_y()
        pdf.cell(W, 7, "", fill=True, **NL)
        pdf.set_y(bar_y + 1)
        pdf.set_x(pdf.l_margin + 1)

        seq = f"MP-{str(p.project_seq_id or p.id).zfill(6)}"
        pdf.set_font("Helvetica", "B", 8)
        tc(*BLUE)
        pdf.cell(24, 5, seq, **ST)

        # Type badge
        fc(*TYPE_COLORS.get(p.type, GRAY))
        tc(*WHITE)
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(20, 5, p.type, fill=True, **ST)
        pdf.cell(3,  5, "",               **ST)

        # Status badge
        fc(*STATUS_COLORS.get(p.status, GRAY))
        pdf.cell(26, 5, _STATUS_LABELS.get(p.status, p.status), fill=True, **ST)
        pdf.cell(3,  5, "",               **ST)

        # Responsible
        tc(*BLACK)
        pdf.set_font("Helvetica", "", 8)
        resp = p.responsible.name if p.responsible else "-"
        remaining = W - 1 - 24 - 20 - 3 - 26 - 3
        pdf.cell(remaining, 5, f"Responsible: {resp}", **NL)

        # Due date
        pdf.ln(2)
        if p.due_date:
            today   = datetime.utcnow().date()
            overdue = str(p.due_date) < str(today) and p.status != "CLOSED"
            pdf.set_x(pdf.l_margin + 2)
            pdf.set_font("Helvetica", "B" if overdue else "", 8)
            tc(*(RED if overdue else GRAY))
            label = f"Due: {p.due_date}" + (" (OVERDUE)" if overdue else "")
            pdf.cell(W, 4, label, **NL)

        # Topic
        pdf.set_x(pdf.l_margin + 2)
        pdf.set_font("Helvetica", "B", 10)
        tc(*BLACK)
        pdf.multi_cell(W - 2, 5.5, p.topic, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        # Details
        if p.details:
            pdf.set_x(pdf.l_margin + 2)
            pdf.set_font("Helvetica", "", 9)
            tc(*GRAY)
            pdf.multi_cell(W - 2, 4.5, p.details, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        # Closed info
        if p.closed_at:
            pdf.set_x(pdf.l_margin + 2)
            pdf.set_font("Helvetica", "I", 8)
            tc(*GREEN)
            pdf.cell(W, 4, f"Closed: {p.closed_at.strftime('%Y-%m-%d %H:%M')}", **NL)

        # Notes
        sorted_notes = sorted(p.notes, key=lambda x: x.created_at or datetime.min)
        if sorted_notes:
            pdf.ln(1)
            pdf.set_x(pdf.l_margin + 2)
            pdf.set_font("Helvetica", "B", 8)
            tc(*BLUE)
            pdf.cell(W, 4.5, "Notes:", **NL)
            for n in sorted_notes:
                date_str = n.created_at.strftime("%Y-%m-%d") if n.created_at else ""
                author   = n.author.name if n.author else "?"
                pdf.set_x(pdf.l_margin + 5)
                pdf.set_font("Helvetica", "B", 8)
                tc(*BLACK)
                pdf.cell(W - 5, 4, f"{author} - {date_str}", **NL)
                if n.content:
                    pdf.set_x(pdf.l_margin + 5)
                    pdf.set_font("Helvetica", "", 8)
                    tc(*GRAY)
                    pdf.multi_cell(W - 5, 4, n.content, new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        pdf.ln(3)
        divider()

    buf = io.BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    fname = f"meeting_{_safe_filename(meeting.title)}_{meeting.date or 'nodate'}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── All Points Excel export ───────────────────────────────────────────────────

@router.get("/api/meeting-points/export/excel")
def export_all_points_excel(
    status: str = FQuery(None),
    point_type: str = FQuery(None),
    responsible_id: int = FQuery(None),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(models.MeetingPoint).filter(
        models.MeetingPoint.project_id == user.project_id
    )
    if status:
        q = q.filter(models.MeetingPoint.status == status)
    if point_type:
        q = q.filter(models.MeetingPoint.type == point_type)
    if responsible_id:
        q = q.filter(models.MeetingPoint.responsible_id == responsible_id)

    points = q.order_by(models.MeetingPoint.project_seq_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Meeting Points"

    headers   = ["ID", "Type", "Topic", "Details", "Notes", "Responsible", "Due Date", "Status", "Closed At"]
    col_widths = [12,   12,     32,      50,         50,      25,            14,         15,        20]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    for ri, p in enumerate(points, start=2):
        responsible = (
            db.query(models.Contact).filter(models.Contact.id == p.responsible_id).first()
            if p.responsible_id else None
        )
        notes_text = _format_notes(p.notes)
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None

        row_vals = [
            f"MP-{str(p.project_seq_id or p.id).zfill(6)}",
            p.type,
            p.topic,
            p.details or "",
            notes_text,
            responsible.name if responsible else "",
            str(p.due_date) if p.due_date else "",
            _STATUS_LABELS.get(p.status, p.status),
            p.closed_at.strftime("%Y-%m-%d %H:%M") if p.closed_at else "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="meeting_points_{date_str}.xlsx"'},
    )


# ── Budget: Orders Excel export ───────────────────────────────────────────────

@router.get("/api/budget/orders/export/excel")
def export_orders_excel(
    package_id: int = FQuery(None),
    status: str = FQuery(None),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(models.Order).join(models.Package).filter(
        models.Package.project_id == user.project_id
    )
    if package_id:
        q = q.filter(models.Order.package_id == package_id)
    if status:
        q = q.filter(models.Order.status == status)
    orders = q.order_by(models.Order.order_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers    = ["PO Number", "Package", "Vendor", "Description", "Order Date", "Amount", "Currency", "Status"]
    col_widths = [18,           20,         25,        45,            14,           15,        10,          14]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    STATUS_LABELS = {"DRAFT": "Draft", "COMMITTED": "Committed", "CANCELLED": "Cancelled"}

    for ri, o in enumerate(orders, start=2):
        pkg = db.query(models.Package).filter(models.Package.id == o.package_id).first() if o.package_id else None
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        row_vals = [
            o.po_number or "",
            f"{pkg.tag_number} - {pkg.name}" if pkg else "",
            o.vendor_name or "",
            o.description or "",
            str(o.order_date) if o.order_date else "",
            o.amount,
            o.currency or "",
            STATUS_LABELS.get(o.status, o.status or ""),
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill
        # Right-align amount column
        ws.cell(row=ri, column=6).alignment = Alignment(horizontal="right", vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="orders_{date_str}.xlsx"'},
    )


# ── Budget: Invoices Excel export ─────────────────────────────────────────────

@router.get("/api/budget/invoices/export/excel")
def export_invoices_excel(
    package_id: int = FQuery(None),
    status: str = FQuery(None),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(models.Invoice).join(models.Order).join(
        models.Package, models.Order.package_id == models.Package.id
    ).filter(models.Package.project_id == user.project_id)
    if package_id:
        q = q.filter(models.Invoice.package_id == package_id)
    if status:
        q = q.filter(models.Invoice.status == status)
    invoices = q.order_by(models.Invoice.invoice_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers    = ["Invoice #", "PO Number", "Package", "Description", "Invoice Date", "Amount", "Currency", "PMC Approved", "Client Approved", "Status", "Review Comment"]
    col_widths = [16,           16,           20,         45,            14,             15,        10,          14,              14,               14,        35]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    STATUS_LABELS = {"DRAFT": "Draft", "PENDING": "Pending", "APPROVED": "Approved", "REJECTED": "Rejected", "CANCELLED": "Cancelled"}

    for ri, inv in enumerate(invoices, start=2):
        pkg   = db.query(models.Package).filter(models.Package.id == inv.package_id).first() if inv.package_id else None
        order = db.query(models.Order).filter(models.Order.id == inv.order_id).first() if inv.order_id else None
        fill  = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        row_vals = [
            inv.invoice_number or "",
            order.po_number if order else "",
            f"{pkg.tag_number} - {pkg.name}" if pkg else "",
            inv.description or "",
            str(inv.invoice_date) if inv.invoice_date else "",
            inv.amount,
            inv.currency or "",
            "Yes" if inv.pmc_approved else "No",
            "Yes" if inv.client_approved else "No",
            STATUS_LABELS.get(inv.status, inv.status or ""),
            inv.review_comment or "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill
        ws.cell(row=ri, column=6).alignment = Alignment(horizontal="right", vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="invoices_{date_str}.xlsx"'},
    )


# ── Risk Register Excel export ────────────────────────────────────────────────

@router.get("/api/risks/export/excel")
def export_risks_excel(
    status: str = FQuery(None),
    category_id: int = FQuery(None),
    phase_id: int = FQuery(None),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(models.Risk).filter(models.Risk.project_id == user.project_id)
    if status:
        q = q.filter(models.Risk.status == status)
    if category_id:
        q = q.filter(models.Risk.category_id == category_id)
    if phase_id:
        q = q.filter(models.Risk.phase_id == phase_id)
    risks = q.order_by(models.Risk.project_seq_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    headers = [
        "ID", "Title", "Category", "Phase", "Status",
        "Prob Before", "CAPEX Score Before", "Schedule Score Before",
        "CAPEX Value", "Schedule Value",
        "Mitigation Type", "Mitigation Action", "Action Due Date", "Action Status",
        "Prob After", "CAPEX Score After", "Schedule Score After",
        "Owner", "Date Opened", "Date Closed",
        "Secondary Effects", "Notes",
    ]
    col_widths = [
        10, 35, 20, 20, 14,
        12, 18, 20,
        15, 15,
        18, 45, 16, 16,
        12, 16, 18,
        25, 14, 14,
        35, 45,
    ]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    STATUS_LABELS = {
        "OPEN": "Open", "MITIGATED": "Mitigated", "CLOSED": "Closed",
        "ACCEPTED": "Accepted", "TRANSFERRED": "Transferred",
    }
    ACTION_STATUS_LABELS = {
        "PENDING": "Pending", "IN_PROGRESS": "In Progress",
        "COMPLETED": "Completed", "OVERDUE": "Overdue",
    }

    for ri, r in enumerate(risks, start=2):
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        row_vals = [
            f"R-{str(r.project_seq_id or r.id).zfill(4)}",
            r.title or "",
            r.category.name if r.category else "",
            r.phase.name if r.phase else "",
            STATUS_LABELS.get(r.status, r.status or ""),
            r.prob_score_before,
            r.capex_score_before,
            r.schedule_score_before,
            r.capex_value,
            r.schedule_value,
            r.mitigation_type or "",
            r.mitigation_action or "",
            str(r.action_due_date) if r.action_due_date else "",
            ACTION_STATUS_LABELS.get(r.action_status, r.action_status or ""),
            r.prob_score_after,
            r.capex_score_after,
            r.schedule_score_after,
            r.owner.name if r.owner else "",
            str(r.date_opened) if r.date_opened else "",
            str(r.date_closed) if r.date_closed else "",
            r.secondary_effects or "",
            "\n\n".join(n.content for n in sorted(r.notes, key=lambda x: x.created_at or datetime.min) if n.content),
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="risk_register_{date_str}.xlsx"'},
    )


# ── Procurement Register Excel export ─────────────────────────────────────────

@router.get("/api/procurement/register/export/excel")
def export_procurement_excel(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    entries = (
        db.query(models.ProcurementEntry)
        .filter(models.ProcurementEntry.project_id == user.project_id)
        .order_by(models.ProcurementEntry.package_id, models.ProcurementEntry.id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Procurement Register"

    headers = [
        "Package", "Company", "Current Step", "Status",
        "Technical Compliance", "Technical Note",
        "Commercial Compliance", "Commercial Note",
        "Bid Value", "Bid Currency", "Exclusion Reason",
    ]
    col_widths = [28, 28, 28, 14, 20, 40, 20, 40, 15, 12, 35]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    for ri, e in enumerate(entries, start=2):
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        pkg  = e.package
        row_vals = [
            f"{pkg.tag_number} - {pkg.name}" if pkg else "",
            e.company.name if e.company else "",
            e.current_step.step_id if e.current_step else "",
            e.status or "",
            e.technical_compliance or "-",
            e.technical_compliance_note or "",
            e.commercial_compliance or "-",
            e.commercial_compliance_note or "",
            e.bid_value,
            e.bid_currency or "",
            e.exclusion_reason or "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill
        ws.cell(row=ri, column=9).alignment = Alignment(horizontal="right", vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="procurement_register_{date_str}.xlsx"'},
    )


# ── Scope Changes Excel export ────────────────────────────────────────────────

@router.get("/api/scope-changes/export/excel")
def export_scope_changes_excel(
    status: str = FQuery(None),
    package_id: int = FQuery(None),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(models.ScopeChange).filter(models.ScopeChange.project_id == user.project_id)
    if status:
        q = q.filter(models.ScopeChange.status == status)
    if package_id:
        q = q.filter(models.ScopeChange.package_id == package_id)
    scope_changes = q.order_by(models.ScopeChange.project_seq_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Scope Changes"

    headers = [
        "ID", "Description", "Details", "Package",
        "Cost", "Schedule Impact (months)", "Status",
        "PMC Reviewer", "PMC Approved", "PMC Comment",
        "Client Reviewer", "Client Approved", "Client Comment",
        "Created By", "Submitted At",
    ]
    col_widths = [10, 35, 50, 22, 14, 22, 14, 22, 14, 35, 22, 14, 35, 22, 18]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    STATUS_LABELS = {
        "DRAFT": "Draft", "SUBMITTED": "Submitted",
        "PMC_APPROVED": "PMC Approved", "APPROVED": "Approved",
        "REJECTED": "Rejected", "CANCELLED": "Cancelled",
    }

    for ri, sc in enumerate(scope_changes, start=2):
        pkg = db.query(models.Package).filter(models.Package.id == sc.package_id).first() if sc.package_id else None
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None

        # Reviewer contacts live on the Package (commercial reviewers), not
        # on the scope change itself. Mirrors routers/scope_changes.py:22-23.
        pmc_contact = pkg.pmc_commercial_reviewer if pkg else None
        client_contact = pkg.client_commercial_reviewer if pkg else None
        created_by = db.query(models.User).filter(models.User.id == sc.created_by_id).first() if sc.created_by_id else None

        row_vals = [
            f"SC-{str(sc.project_seq_id or sc.id).zfill(4)}",
            sc.description or "",
            sc.details or "",
            f"{pkg.tag_number} - {pkg.name}" if pkg else "",
            sc.cost,
            sc.schedule_impact_months,
            STATUS_LABELS.get(sc.status, sc.status or ""),
            pmc_contact.name if pmc_contact else "",
            "Yes" if sc.pmc_approved else ("No" if sc.pmc_reviewed else "-"),
            sc.pmc_comment or "",
            client_contact.name if client_contact else "",
            "Yes" if sc.client_approved else ("No" if sc.client_reviewed else "-"),
            sc.client_comment or "",
            created_by.name if created_by else "",
            str(sc.submitted_at)[:16] if sc.submitted_at else "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill
        ws.cell(row=ri, column=5).alignment = Alignment(horizontal="right", vertical="top")
        ws.cell(row=ri, column=6).alignment = Alignment(horizontal="right", vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="scope_changes_{date_str}.xlsx"'},
    )


# ── Punchlist Excel export ────────────────────────────────────────────────────

@router.get("/api/punch-items/export/excel")
def export_punchlist_excel(
    package_id: int = FQuery(None),
    obligation_time_id: int = FQuery(None),
    area_id: int = FQuery(None),
    unit_id: int = FQuery(None),
    status: str = FQuery(None),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(models.PunchItem).filter(models.PunchItem.project_id == user.project_id)
    # DRAFTs are private to the creator — never included in Excel exports.
    q = q.filter(models.PunchItem.status != "DRAFT")
    if package_id:
        q = q.filter(models.PunchItem.package_id == package_id)
    if obligation_time_id:
        q = q.filter(models.PunchItem.obligation_time_id == obligation_time_id)
    if area_id:
        q = q.filter(models.PunchItem.area_id == area_id)
    if unit_id:
        q = q.filter(models.PunchItem.unit_id == unit_id)
    if status:
        q = q.filter(models.PunchItem.status == status)
    punches = q.order_by(models.PunchItem.project_seq_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Punchlist"

    headers    = ["ID", "Package", "Obligation Time", "Area", "Unit", "Topic", "Details", "Response", "Status", "Created By"]
    col_widths = [10,   20,         22,                12,     12,     35,      50,         50,          12,       22]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    STATUS_LABELS = {"OPEN": "Open", "TO_REVIEW": "To Review", "CLOSED": "Closed"}

    for ri, p in enumerate(punches, start=2):
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        row_vals = [
            f"PL-{str(p.project_seq_id or p.id).zfill(4)}",
            f"{p.package.tag_number} - {p.package.name}" if p.package else "",
            f"{p.obligation_time.code} - {p.obligation_time.name}" if p.obligation_time else "",
            p.area.tag if p.area else "",
            p.unit.tag if p.unit else "",
            p.topic or "",
            p.details or "",
            p.response or "",
            STATUS_LABELS.get(p.status, p.status or ""),
            p.created_by.name if p.created_by else "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="punchlist_{date_str}.xlsx"'},
    )


# ── Contacts export ───────────────────────────────────────────────────────────

@router.get("/api/contacts/export/excel")
def export_contacts_excel(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    contacts = (
        db.query(models.Contact)
        .filter(models.Contact.project_id == user.project_id)
        .order_by(models.Contact.name)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Contact List"

    headers    = ["Name", "Company", "Function", "Email", "Phone"]
    col_widths = [30,      30,        30,          35,      18]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    for ri, c in enumerate(contacts, start=2):
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        row_vals = [
            c.name or "",
            c.company or "",
            c.function or "",
            c.email or "",
            c.phone or "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="contacts_{date_str}.xlsx"'},
    )


# ── Document list export ──────────────────────────────────────────────────────

@router.get("/api/documents/export/excel")
def export_documents_excel(
    package_id: int = FQuery(None),
    document_type: str = FQuery(None),
    status: str = FQuery(None),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(models.Document).filter(models.Document.project_id == user.project_id)
    if package_id:
        q = q.filter(models.Document.package_id == package_id)
    if document_type:
        q = q.filter(models.Document.document_type == document_type)
    if status:
        q = q.filter(models.Document.status == status)
    docs = q.order_by(models.Document.project_seq_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Document List"

    headers    = ["ID", "Package", "Subservice", "Type", "Description", "Area", "Unit", "Start Date", "First Issue Date", "Status", "Version"]
    col_widths = [10,   22,         30,            12,     50,            12,     12,     14,            18,                  14,        10]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    for ri, d in enumerate(docs, start=2):
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        row_vals = [
            f"DOC-{str(d.project_seq_id or d.id).zfill(4)}",
            f"{d.package.tag_number} - {d.package.name}" if d.package else "",
            f"{d.subservice.subservice_code} - {d.subservice.subservice_name}" if d.subservice else "",
            d.document_type or "",
            d.description or "",
            d.area.tag if d.area else "",
            d.unit.tag if d.unit else "",
            d.start_date or "",
            d.first_issue_date or "",
            ("In Progress" if d.actual_start_date else "Not Started") if d.status == "DRAFT" else {
                "IN_REVIEW":              "In Review",
                "APPROVED":               "Approved",
                "APPROVED_WITH_COMMENTS": "Approved with Comments",
                "REJECTED":               "Rejected",
            }.get(d.status, d.status or ""),
            d.current_version or 0,
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="documents_{date_str}.xlsx"'},
    )


# ── ITP register export ───────────────────────────────────────────────────────

@router.get("/api/itp/export/excel")
def export_itp_excel(
    package_id: int = FQuery(None),
    test_type_id: int = FQuery(None),
    witness_level_id: int = FQuery(None),
    status: str = FQuery(None),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    q = db.query(models.ITPRecord).filter(models.ITPRecord.project_id == user.project_id)
    if package_id:
        q = q.filter(models.ITPRecord.package_id == package_id)
    if test_type_id:
        q = q.filter(models.ITPRecord.test_type_id == test_type_id)
    if witness_level_id:
        q = q.filter(models.ITPRecord.witness_level_id == witness_level_id)
    if status:
        q = q.filter(models.ITPRecord.status == status)
    records = q.order_by(models.ITPRecord.project_seq_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "ITP Register"

    headers    = ["ID", "Package", "Test Type", "Test", "Details", "Witness Level", "Status", "Approval", "Area", "Unit", "Planned Date", "Executed Date"]
    col_widths = [10,   22,         20,           25,     40,         20,              14,        14,          12,     12,     14,              14]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    APPROVAL_LABELS = {
        "TO_SUBMIT": "To Submit",
        "PENDING":   "Pending",
        "APPROVED":  "Approved",
        "REJECTED":  "Rejected",
    }

    for ri, r in enumerate(records, start=2):
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        row_vals = [
            f"ITP-{str(r.project_seq_id or r.id).zfill(4)}",
            f"{r.package.tag_number} - {r.package.name}" if r.package else "",
            r.test_type.name if r.test_type else "",
            r.test or "",
            r.details or "",
            f"{r.witness_level.code} - {r.witness_level.name}" if r.witness_level else "",
            _STATUS_LABELS.get(r.status, r.status or ""),
            APPROVAL_LABELS.get(r.approval_status, r.approval_status or ""),
            r.area.tag if r.area else "",
            r.unit.tag if r.unit else "",
            r.planned_date or "",
            r.executed_date or "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="itp_register_{date_str}.xlsx"'},
    )


# ── Procurement plan export ───────────────────────────────────────────────────

@router.get("/api/procurement/plan/export/excel")
def export_procurement_plan_excel(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    steps = (
        db.query(models.ProcurementStep)
        .filter(models.ProcurementStep.project_id == user.project_id)
        .order_by(models.ProcurementStep.sort_order)
        .all()
    )
    plans = (
        db.query(models.PackagePlan)
        .filter(models.PackagePlan.project_id == user.project_id)
        .join(models.Package, models.PackagePlan.package_id == models.Package.id)
        .order_by(models.Package.tag_number)
        .all()
    )

    # ── Budget (Forecast) + Progress per package, using the same formulas
    # as /api/procurement/plans so the export matches what the UI displays.
    cum_weights, running = {}, 0.0
    for s in sorted(steps, key=lambda x: x.sort_order):
        cum_weights[s.id] = round(running * 100, 1)
        running += (s.weight or 0.0)

    def _avg_progress(entries):
        active = [e for e in entries if e.status != "EXCLUDED"]
        if not active:
            return 0.0
        total = sum(
            100.0 if e.status == "AWARDED" else cum_weights.get(e.current_step_id, 0.0)
            for e in active
        )
        return round(total / len(active), 1)

    entries_by_pkg = {}
    for e in db.query(models.ProcurementEntry).filter_by(project_id=user.project_id).all():
        entries_by_pkg.setdefault(e.package_id, []).append(e)

    forecasts, currencies = {}, {}
    for pkg_id in {p.package_id for p in plans}:
        bl = db.query(models.BudgetBaseline).filter_by(package_id=pkg_id).first()
        baseline = bl.amount if bl else 0.0
        currency = bl.currency if bl else "EUR"
        tin  = sum(t.amount for t in db.query(models.BudgetTransfer).filter_by(to_package_id=pkg_id).all())
        tout = sum(t.amount for t in db.query(models.BudgetTransfer).filter_by(from_package_id=pkg_id).all())
        forecasts[pkg_id] = baseline + tin - tout
        currencies[pkg_id] = currency

    wb = Workbook()
    ws = wb.active
    ws.title = "Procurement Plan"

    static_headers    = ["Package", "Package Owner", "Budget (Forecast)", "Currency", "Progress (%)", "Contract Type", "Bidders", "Notes"]
    static_col_widths = [25,         22,              18,                   10,         14,             20,              40,        40]
    step_headers    = [s.step_id for s in steps]
    step_col_widths = [16] * len(steps)

    headers    = static_headers + step_headers
    col_widths = static_col_widths + step_col_widths

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    for ri, plan in enumerate(plans, start=2):
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        package = db.query(models.Package).filter(models.Package.id == plan.package_id).first()
        bidders = ", ".join(b.company.name for b in plan.bidders if b.company) if plan.bidders else ""
        step_dates_map = {sd.step_id: sd.due_date or "" for sd in plan.step_dates}
        owner_name = package.package_owner.name if (package and package.package_owner) else ""
        budget = round(forecasts.get(plan.package_id, 0.0), 2)
        currency = currencies.get(plan.package_id, "EUR")
        progress = _avg_progress(entries_by_pkg.get(plan.package_id, []))
        row_vals = [
            f"{package.tag_number} - {package.name}" if package else "",
            owner_name,
            budget,
            currency,
            progress,
            plan.contract_type.name if plan.contract_type else "",
            bidders,
            plan.notes or "",
        ] + [step_dates_map.get(s.id, "") for s in steps]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill
        # Right-align the three numeric/currency columns (Budget, Progress)
        ws.cell(row=ri, column=3).alignment = Alignment(horizontal="right", vertical="top")
        ws.cell(row=ri, column=5).alignment = Alignment(horizontal="right", vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="procurement_plan_{date_str}.xlsx"'},
    )


# ── Tasks export ──────────────────────────────────────────────────────────────

@router.get("/api/tasks/export/excel")
def export_tasks_excel(
    package_id: int = FQuery(None),
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    from datetime import date as _date
    from routers.schedule import _fmt_task as _schedule_fmt_task

    q = db.query(models.Task).filter(models.Task.project_id == user.project_id)
    if package_id:
        q = q.filter(models.Task.package_id == package_id)
    tasks = q.order_by(models.Task.project_seq_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Progress + Status + Late mirror what the schedule UI shows: progress
    # is the latest APPROVED progress-report-entry percentage; status is
    # derived from the percentage; late = today > finish_date AND not yet 100.
    headers    = ["ID", "Package", "Description", "Details", "Start Date", "Finish Date",
                  "Financial Weight (%)", "Progress (%)", "Status", "Late",
                  "Area", "Unit"]
    col_widths = [10,   22,        40,             50,         14,            14,
                  20,                    14,              14,        8,
                  18,     18]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    def _status(pct):
        if pct >= 100: return "Complete"
        if pct > 0:    return "In progress"
        return "Not started"

    for ri, t in enumerate(tasks, start=2):
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        info = _schedule_fmt_task(t, db)
        pct = info.get("current_progress") or 0
        row_vals = [
            f"T-{str(t.project_seq_id or t.id).zfill(6)}",
            f"{t.package.tag_number} - {t.package.name}" if t.package else "",
            t.description or "",
            t.details or "",
            t.start_date or "",
            t.finish_date or "",
            t.financial_weight if t.financial_weight is not None else "",
            round(pct, 1),
            _status(pct),
            "Y" if info.get("is_late") else "N",
            f"{t.area.tag} — {t.area.description}" if t.area else "",
            f"{t.unit.tag} — {t.unit.description}" if t.unit else "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="tasks_{date_str}.xlsx"'},
    )


# ── Budget: Overview Excel export ─────────────────────────────────────────────
# Mirrors the columns shown in budget.js Budget Overview tab.

@router.get("/api/budget/overview/export/excel")
def export_budget_overview_excel(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from routers.budget import get_budget_overview

    rows = get_budget_overview(user=user, db=db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Overview"

    headers = [
        "Package", "Package Name", "Currency",
        "Baseline", "Actual Budget", "Bid Value",
        "Committed", "Remaining", "Pending SC",
        "Remaining incl. pending SC", "Spend",
        "Bid Status", "Awarded Vendor",
    ]
    col_widths = [12, 30, 10, 15, 15, 15, 15, 15, 15, 22, 15, 14, 24]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        hc = ws.cell(row=1, column=col, value=h)
        hc.font = Font(bold=True, color=_WHITE)
        hc.fill = PatternFill("solid", fgColor=_DARK_BLUE)
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    # Build totals row alongside the data rows
    totals = {k: 0.0 for k in (
        "baseline", "forecast", "bid_value", "committed", "remaining",
        "pending_sc_cost", "remaining_incl_pending", "spend",
    )}
    for ri, r in enumerate(rows, start=2):
        fill = PatternFill("solid", fgColor=_LIGHT_BLUE) if ri % 2 == 0 else None
        bid_value = r.get("bid_value")
        bid_value_cell = bid_value if bid_value is not None else ""
        row_vals = [
            r.get("tag_number", ""),
            r.get("name", ""),
            r.get("currency", "EUR"),
            r.get("baseline", 0) or 0,
            r.get("forecast", 0) or 0,
            bid_value_cell,
            r.get("committed", 0) or 0,
            r.get("remaining", 0) or 0,
            r.get("pending_sc_cost", 0) or 0,
            r.get("remaining_incl_pending", 0) or 0,
            r.get("spend", 0) or 0,
            (r.get("bid_status") or "").title(),
            r.get("awarded_company_name") or "",
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

        for k in totals:
            v = r.get(k)
            if isinstance(v, (int, float)):
                totals[k] += v

    # Totals row (skip if no data)
    if rows:
        tr = len(rows) + 2
        bold_fill = PatternFill("solid", fgColor="E5E7EB")
        ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=tr, column=1).fill = bold_fill
        ws.cell(row=tr, column=2).fill = bold_fill
        ws.cell(row=tr, column=3).fill = bold_fill
        amount_cols = {
            4: totals["baseline"], 5: totals["forecast"], 6: None,
            7: totals["committed"], 8: totals["remaining"],
            9: totals["pending_sc_cost"], 10: totals["remaining_incl_pending"],
            11: totals["spend"],
        }
        for col, v in amount_cols.items():
            cell = ws.cell(row=tr, column=col, value=v if v is not None else "")
            cell.font = Font(bold=True)
            cell.fill = bold_fill
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
        ws.cell(row=tr, column=12).fill = bold_fill
        ws.cell(row=tr, column=13).fill = bold_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="budget_overview_{date_str}.xlsx"'},
    )
