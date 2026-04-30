"""
Construction module — setup lists (work-permit types, safety-observation
categories, worker-certificate types) + Workers & Subcontractors register.
"""
from datetime import datetime, date, timedelta
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func, distinct
from sqlalchemy.orm import Session, joinedload, selectinload
from pydantic import BaseModel

from database import get_db
import models
import auth

router = APIRouter(prefix="/api/construction", tags=["construction"])


# ── Permission helpers ───────────────────────────────────────────────────────

def _is_owner_or_admin(user: auth.ProjectContext, db: Session) -> bool:
    return auth.has_owner_or_lead_access(user, "Construction", db)


def _vendor_package_ids(user: auth.ProjectContext, db: Session) -> set:
    """Packages the current vendor (or any user) is linked to via PackageContact."""
    if not user.contact_id:
        return set()
    rows = db.query(models.PackageContact).filter_by(contact_id=user.contact_id).all()
    return {r.package_id for r in rows}


def _user_can_manage_package(user: auth.ProjectContext, package_id: int, db: Session) -> bool:
    """Admin / Project Owner: always. Otherwise user must be linked to the package."""
    if _is_owner_or_admin(user, db):
        return True
    if not user.contact_id:
        return False
    pkg = db.query(models.Package).filter_by(id=package_id, project_id=user.project_id).first()
    if not pkg:
        return False
    if pkg.package_owner_id == user.contact_id:
        return True
    return package_id in _vendor_package_ids(user, db)


def _is_project_site_supervisor(user: auth.ProjectContext, db: Session) -> bool:
    """True iff the current user is declared as a site supervisor on any area
    of the current project (via AreaSiteSupervisor → Area)."""
    if not user.contact_id:
        return False
    row = (db.query(models.AreaSiteSupervisor)
             .join(models.Area, models.Area.id == models.AreaSiteSupervisor.area_id)
             .filter(models.Area.project_id == user.project_id,
                     models.AreaSiteSupervisor.contact_id == user.contact_id)
             .first())
    return row is not None


def _user_can_manage_permit_package(user: auth.ProjectContext,
                                    package_id: int, db: Session) -> bool:
    """Authorisation for work-permit CRUD. Widens _user_can_manage_package by
    also granting rights to any declared site supervisor on the project, so
    supervisors can raise permits for every package."""
    if _user_can_manage_package(user, package_id, db):
        return True
    return _is_project_site_supervisor(user, db)


# ── Schemas ──────────────────────────────────────────────────────────────────

class SetupItemCreate(BaseModel):
    name: str
    description: Optional[str] = None
    sort_order: Optional[int] = None
    # Only used for safety-observation-categories. Ignored elsewhere.
    polarity: Optional[str] = None

class SetupItemUpdate(SetupItemCreate):
    pass


class SubcontractorCreate(BaseModel):
    package_id: int
    company: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    description: Optional[str] = None

class SubcontractorUpdate(SubcontractorCreate):
    pass


class WorkerCreate(BaseModel):
    package_id: int
    name: str
    phone: Optional[str] = None
    is_subcontractor: bool = False
    subcontractor_id: Optional[int] = None
    certificate_type_ids: Optional[List[int]] = None

class WorkerUpdate(WorkerCreate):
    pass


class WorkLogCreate(BaseModel):
    package_id: int
    start_date: str
    end_date: Optional[str] = None
    notes: Optional[str] = None
    ignore_missing_reports: bool = False

class WorkLogUpdate(WorkLogCreate):
    pass


class DailyReportBody(BaseModel):
    package_id: int
    report_date: str
    avg_hours_per_worker: float = 0.0
    description: Optional[str] = None
    worker_ids: List[int] = []
    area_ids: List[int] = []
    no_work: bool = False
    expected_worker_count: Optional[int] = None


class WorkPermitHazardBody(BaseModel):
    hazard_key: str
    preventive_measure: Optional[str] = None

class LOTOItemBody(BaseModel):
    """LOTO entry inside a work permit payload. `id` is present when updating
    an existing LOTO; omitted for new LOTOs.

    `action` lets the vendor drive state transitions from the parent permit:
        • 'resubmit' — REFUSED/CANCELLED → REQUEST (after optional tag/desc edits)
        • 'cancel'   — REFUSED/REQUEST   → CANCELLED
        • None       — edits only, no status change
    """
    id: Optional[int] = None
    tag_number: str
    description: Optional[str] = None
    action: Optional[str] = None

class WorkPermitBody(BaseModel):
    package_id: int
    title: Optional[str] = None
    description: Optional[str] = None
    start_date: str
    end_date: str
    permit_type_ids: List[int] = []
    area_ids: List[int] = []
    hazards: List[WorkPermitHazardBody] = []
    hazards_other: Optional[str] = None
    ppe_keys: List[str] = []
    ppe_other: Optional[str] = None
    lotos: List[LOTOItemBody] = []


# ── Formatters ───────────────────────────────────────────────────────────────

def _fmt_setup_item(r) -> dict:
    out = {
        "id": r.id, "name": r.name,
        "description": r.description or "",
        "sort_order": r.sort_order or 0,
    }
    # Safety-observation categories also carry a polarity (POSITIVE / NEGATIVE).
    if isinstance(r, models.SafetyObservationCategory):
        out["polarity"] = (r.polarity or "NEGATIVE").upper()
    return out


def _fmt_subcontractor(s: models.Subcontractor) -> dict:
    return {
        "id": s.id,
        "seq_id": s.project_seq_id,
        "display_id": f"SU-{(s.project_seq_id or s.id):06d}",
        "package_id": s.package_id,
        "package_tag": s.package.tag_number if s.package else None,
        "package_name": s.package.name if s.package else None,
        "company": s.company,
        "contact_person": s.contact_person or "",
        "phone": s.phone or "",
        "email": s.email or "",
        "description": s.description or "",
        "created_at": s.created_at.isoformat() + 'Z' if s.created_at else None,
        "created_by_name": s.created_by.name if s.created_by else None,
    }


def _fmt_work_log(wl: models.WorkLog) -> dict:
    return {
        "id": wl.id,
        "package_id": wl.package_id,
        "package_tag": wl.package.tag_number if wl.package else None,
        "package_name": wl.package.name if wl.package else None,
        "start_date": wl.start_date,
        "end_date": wl.end_date,
        "notes": wl.notes or "",
        "ignore_missing_reports": bool(wl.ignore_missing_reports),
        "created_at": wl.created_at.isoformat() + 'Z' if wl.created_at else None,
        "created_by_name": wl.created_by.name if wl.created_by else None,
    }


def _fmt_worker(w: models.Worker) -> dict:
    cert_ids = [wc.certificate_type_id for wc in (w.certificates or [])]
    cert_list = []
    for wc in (w.certificates or []):
        t = wc.certificate_type
        if t:
            cert_list.append({"id": t.id, "name": t.name})
    return {
        "id": w.id,
        "seq_id": w.project_seq_id,
        "display_id": f"WK-{(w.project_seq_id or w.id):06d}",
        "package_id": w.package_id,
        "package_tag": w.package.tag_number if w.package else None,
        "package_name": w.package.name if w.package else None,
        "name": w.name,
        "phone": w.phone or "",
        "is_subcontractor": bool(w.is_subcontractor),
        "subcontractor_id": w.subcontractor_id,
        "subcontractor_company": (w.subcontractor.company if w.subcontractor else None),
        "certificate_type_ids": cert_ids,
        "certificates": cert_list,
        "status": w.status or "PENDING",
        "submitted_at": w.submitted_at.isoformat() + 'Z' if w.submitted_at else None,
        "reviewed_at":  w.reviewed_at.isoformat()  + 'Z' if w.reviewed_at  else None,
        "reviewed_by_name": w.reviewed_by.name if w.reviewed_by else None,
        "rejection_comment": w.rejection_comment or "",
        "created_at": w.created_at.isoformat() + 'Z' if w.created_at else None,
        "created_by_name": w.created_by.name if w.created_by else None,
    }


def _log_worker_review(db: Session, worker: models.Worker, event: str,
                       user, approved=None, comment: Optional[str] = None):
    db.add(models.WorkerReview(
        worker_id=worker.id, event=event, approved=approved,
        comment=(comment or None), actor_id=(user.id if user else None),
    ))


def _is_site_supervisor(db: Session, project_id: int, contact_id: Optional[int]) -> bool:
    """A user with `contact_id` is a site supervisor iff they are assigned on
    at least one area of the project (via AreaSiteSupervisor).
    Admins / project owners always pass this check."""
    if not contact_id:
        return False
    row = (db.query(models.AreaSiteSupervisor)
             .join(models.Area, models.Area.id == models.AreaSiteSupervisor.area_id)
             .filter(models.Area.project_id == project_id,
                     models.AreaSiteSupervisor.contact_id == contact_id)
             .first())
    return row is not None


def _can_review_worker(user: auth.ProjectContext, worker: models.Worker, db: Session) -> bool:
    if _is_owner_or_admin(user, db):
        return True
    return _is_site_supervisor(db, worker.project_id, user.contact_id)


def _can_manage_worker(user: auth.ProjectContext, worker: models.Worker, db: Session) -> bool:
    """Vendor-side actions — submit / resubmit / cancel. Admins and package-
    linked contacts always allowed."""
    if _is_owner_or_admin(user, db):
        return True
    return _user_can_manage_package(user, worker.package_id, db)


# ═════════════════════════════════════════════════════════════════════════════
# SETUP — Work-permit types / Safety-observation categories / Worker-cert types
# All setup endpoints: project owners & admins only.
# ═════════════════════════════════════════════════════════════════════════════

_SETUP_MODELS = {
    "work-permit-types":          models.WorkPermitType,
    "safety-observation-categories": models.SafetyObservationCategory,
    "worker-certificate-types":   models.WorkerCertificateType,
}


def _setup_model(kind: str):
    m = _SETUP_MODELS.get(kind)
    if not m:
        raise HTTPException(404, "Unknown setup list")
    return m


# NOTE: literal `/setup/areas-supervisors` route MUST be declared before the
# `/setup/{kind}` dynamic route below — otherwise FastAPI matches "areas-
# supervisors" as the {kind} placeholder and _setup_model() rejects it.
@router.get("/setup/areas-supervisors")
def list_areas_supervisors(db: Session = Depends(get_db),
                           user: auth.ProjectContext = Depends(auth.get_project_user)):
    areas = db.query(models.Area).filter_by(project_id=user.project_id).order_by(models.Area.tag).all()
    out = []
    for a in areas:
        supervisors = []
        for link in (a.site_supervisors or []):
            c = link.contact
            if c:
                supervisors.append({"id": c.id, "name": c.name, "company": c.company})
        out.append({
            "id": a.id, "tag": a.tag, "description": a.description,
            "owner_id": a.owner_id,
            "owner_name": a.owner.name if a.owner else None,
            "site_supervisors": supervisors,
        })
    return out


@router.get("/setup/{kind}")
def list_setup(kind: str, db: Session = Depends(get_db),
               user: auth.ProjectContext = Depends(auth.get_project_user)):
    M = _setup_model(kind)
    rows = db.query(M).filter_by(project_id=user.project_id).order_by(
        M.sort_order, M.id).all()
    return [_fmt_setup_item(r) for r in rows]


@router.post("/setup/{kind}")
def create_setup(kind: str, body: SetupItemCreate,
                 db: Session = Depends(get_db),
                 user: auth.ProjectContext = Depends(auth.get_project_user)):
    if not _is_owner_or_admin(user, db):
        raise HTTPException(403, "Only project owners can edit setup lists")
    M = _setup_model(kind)
    if not body.name.strip():
        raise HTTPException(400, "Name is required")
    next_order = body.sort_order
    if next_order is None:
        existing = db.query(M).filter_by(project_id=user.project_id).all()
        next_order = (max((r.sort_order or 0) for r in existing) + 1) if existing else 0
    row = M(project_id=user.project_id, name=body.name.strip(),
            description=(body.description or "").strip() or None,
            sort_order=next_order)
    if M is models.SafetyObservationCategory:
        p = (body.polarity or "NEGATIVE").upper()
        row.polarity = p if p in ("POSITIVE", "NEGATIVE") else "NEGATIVE"
    db.add(row); db.commit(); db.refresh(row)
    return _fmt_setup_item(row)


@router.put("/setup/{kind}/{item_id}")
def update_setup(kind: str, item_id: int, body: SetupItemUpdate,
                 db: Session = Depends(get_db),
                 user: auth.ProjectContext = Depends(auth.get_project_user)):
    if not _is_owner_or_admin(user, db):
        raise HTTPException(403, "Only project owners can edit setup lists")
    M = _setup_model(kind)
    row = db.query(M).filter_by(id=item_id, project_id=user.project_id).first()
    if not row:
        raise HTTPException(404, "Item not found")
    if not body.name.strip():
        raise HTTPException(400, "Name is required")
    row.name = body.name.strip()
    row.description = (body.description or "").strip() or None
    if body.sort_order is not None:
        row.sort_order = body.sort_order
    if M is models.SafetyObservationCategory and body.polarity is not None:
        p = body.polarity.upper()
        row.polarity = p if p in ("POSITIVE", "NEGATIVE") else "NEGATIVE"
    db.commit(); db.refresh(row)
    return _fmt_setup_item(row)


@router.delete("/setup/{kind}/{item_id}")
def delete_setup(kind: str, item_id: int,
                 db: Session = Depends(get_db),
                 user: auth.ProjectContext = Depends(auth.get_project_user)):
    if not _is_owner_or_admin(user, db):
        raise HTTPException(403, "Only project owners can edit setup lists")
    M = _setup_model(kind)
    row = db.query(M).filter_by(id=item_id, project_id=user.project_id).first()
    if not row:
        raise HTTPException(404, "Item not found")
    db.delete(row); db.commit()
    return {"ok": True}


# ═════════════════════════════════════════════════════════════════════════════
# SUBCONTRACTORS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/subcontractors")
def list_subcontractors(package_id: Optional[int] = None,
                         db: Session = Depends(get_db),
                         user: auth.ProjectContext = Depends(auth.get_project_user)):
    q = db.query(models.Subcontractor).filter_by(project_id=user.project_id)
    if package_id:
        q = q.filter(models.Subcontractor.package_id == package_id)
    # Vendors only see subcontractors for packages they're linked to.
    if user.role == "VENDOR":
        pkg_ids = _vendor_package_ids(user, db)
        if not pkg_ids:
            return []
        q = q.filter(models.Subcontractor.package_id.in_(pkg_ids))
    rows = q.order_by(models.Subcontractor.created_at.desc()).all()
    return [_fmt_subcontractor(s) for s in rows]


@router.post("/subcontractors")
def create_subcontractor(body: SubcontractorCreate,
                         db: Session = Depends(get_db),
                         user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Bidders cannot register subcontractors")
    if not body.company.strip():
        raise HTTPException(400, "Company is required")
    if not _user_can_manage_package(user, body.package_id, db):
        raise HTTPException(403, "You are not linked to that package")
    s = models.Subcontractor(
        project_id=user.project_id,
        project_seq_id=models.next_project_seq(db, models.Subcontractor, user.project_id),
        package_id=body.package_id,
        company=body.company.strip(),
        contact_person=(body.contact_person or "").strip() or None,
        phone=(body.phone or "").strip() or None,
        email=(body.email or "").strip() or None,
        description=(body.description or "").strip() or None,
        created_by_id=user.id,
    )
    db.add(s); db.commit(); db.refresh(s)
    return _fmt_subcontractor(s)


@router.put("/subcontractors/{sub_id}")
def update_subcontractor(sub_id: int, body: SubcontractorUpdate,
                         db: Session = Depends(get_db),
                         user: auth.ProjectContext = Depends(auth.get_project_user)):
    s = db.query(models.Subcontractor).filter_by(id=sub_id, project_id=user.project_id).first()
    if not s:
        raise HTTPException(404, "Subcontractor not found")
    if not _user_can_manage_package(user, s.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    if body.package_id != s.package_id and not _user_can_manage_package(user, body.package_id, db):
        raise HTTPException(403, "Not authorised for the new package")
    s.package_id = body.package_id
    s.company = body.company.strip()
    s.contact_person = (body.contact_person or "").strip() or None
    s.phone = (body.phone or "").strip() or None
    s.email = (body.email or "").strip() or None
    s.description = (body.description or "").strip() or None
    s.updated_by_id = user.id
    db.commit(); db.refresh(s)
    return _fmt_subcontractor(s)


@router.delete("/subcontractors/{sub_id}")
def delete_subcontractor(sub_id: int,
                         db: Session = Depends(get_db),
                         user: auth.ProjectContext = Depends(auth.get_project_user)):
    s = db.query(models.Subcontractor).filter_by(id=sub_id, project_id=user.project_id).first()
    if not s:
        raise HTTPException(404, "Subcontractor not found")
    if not _user_can_manage_package(user, s.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    # Detach any workers from this subcontractor before delete.
    db.query(models.Worker).filter_by(subcontractor_id=sub_id).update(
        {"subcontractor_id": None, "is_subcontractor": False}
    )
    db.delete(s); db.commit()
    return {"ok": True}


# ═════════════════════════════════════════════════════════════════════════════
# WORKERS
# ═════════════════════════════════════════════════════════════════════════════

@router.get("/workers")
def list_workers(package_id: Optional[int] = None,
                 db: Session = Depends(get_db),
                 user: auth.ProjectContext = Depends(auth.get_project_user)):
    q = db.query(models.Worker).filter_by(project_id=user.project_id)
    if package_id:
        q = q.filter(models.Worker.package_id == package_id)
    if user.role == "VENDOR":
        pkg_ids = _vendor_package_ids(user, db)
        if not pkg_ids:
            return []
        q = q.filter(models.Worker.package_id.in_(pkg_ids))
    rows = q.order_by(models.Worker.created_at.desc()).all()
    return [_fmt_worker(w) for w in rows]


def _apply_certificates(db: Session, worker: models.Worker, ids: Optional[List[int]]):
    if ids is None:
        return
    # Validate that each cert type exists for this project
    valid = {t.id for t in db.query(models.WorkerCertificateType).filter_by(
        project_id=worker.project_id).all()}
    clean = [cid for cid in ids if cid in valid]
    db.query(models.WorkerCertificate).filter_by(worker_id=worker.id).delete()
    db.flush()
    for cid in set(clean):
        db.add(models.WorkerCertificate(worker_id=worker.id, certificate_type_id=cid))


@router.post("/workers")
def create_worker(body: WorkerCreate,
                  db: Session = Depends(get_db),
                  user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Bidders cannot register workers")
    if not body.name.strip():
        raise HTTPException(400, "Worker name is required")
    if not _user_can_manage_package(user, body.package_id, db):
        raise HTTPException(403, "You are not linked to that package")
    if body.is_subcontractor:
        if not body.subcontractor_id:
            raise HTTPException(400, "Select a subcontractor or uncheck the box")
        sub = db.query(models.Subcontractor).filter_by(
            id=body.subcontractor_id, project_id=user.project_id).first()
        if not sub:
            raise HTTPException(400, "Subcontractor not found")
        if sub.package_id != body.package_id:
            raise HTTPException(400, "Subcontractor is linked to a different package")
    else:
        body.subcontractor_id = None

    now = datetime.utcnow()
    w = models.Worker(
        project_id=user.project_id,
        project_seq_id=models.next_project_seq(db, models.Worker, user.project_id),
        package_id=body.package_id,
        name=body.name.strip(),
        phone=(body.phone or "").strip() or None,
        is_subcontractor=bool(body.is_subcontractor),
        subcontractor_id=body.subcontractor_id,
        status="PENDING",
        submitted_at=now,
        created_by_id=user.id,
    )
    db.add(w); db.flush()
    _apply_certificates(db, w, body.certificate_type_ids)
    _log_worker_review(db, w, "SUBMIT", user)
    db.commit(); db.refresh(w)
    return _fmt_worker(w)


@router.put("/workers/{worker_id}")
def update_worker(worker_id: int, body: WorkerUpdate,
                  db: Session = Depends(get_db),
                  user: auth.ProjectContext = Depends(auth.get_project_user)):
    w = db.query(models.Worker).filter_by(id=worker_id, project_id=user.project_id).first()
    if not w:
        raise HTTPException(404, "Worker not found")
    if not _user_can_manage_package(user, w.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    if body.package_id != w.package_id and not _user_can_manage_package(user, body.package_id, db):
        raise HTTPException(403, "Not authorised for the new package")

    if body.is_subcontractor:
        if not body.subcontractor_id:
            raise HTTPException(400, "Select a subcontractor or uncheck the box")
        sub = db.query(models.Subcontractor).filter_by(
            id=body.subcontractor_id, project_id=user.project_id).first()
        if not sub:
            raise HTTPException(400, "Subcontractor not found")
        if sub.package_id != body.package_id:
            raise HTTPException(400, "Subcontractor is linked to a different package")
        w.subcontractor_id = body.subcontractor_id
    else:
        w.subcontractor_id = None

    w.package_id = body.package_id
    w.name = body.name.strip()
    w.phone = (body.phone or "").strip() or None
    w.is_subcontractor = bool(body.is_subcontractor)
    w.updated_by_id = user.id
    _apply_certificates(db, w, body.certificate_type_ids)
    db.commit(); db.refresh(w)
    return _fmt_worker(w)


@router.delete("/workers/{worker_id}")
def delete_worker(worker_id: int,
                  db: Session = Depends(get_db),
                  user: auth.ProjectContext = Depends(auth.get_project_user)):
    w = db.query(models.Worker).filter_by(id=worker_id, project_id=user.project_id).first()
    if not w:
        raise HTTPException(404, "Worker not found")
    if not _user_can_manage_package(user, w.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    db.delete(w); db.commit()
    return {"ok": True}


# ═════════════════════════════════════════════════════════════════════════════
# WORKER APPROVAL WORKFLOW
# Mirrors the Invoice approval pattern: an append-only WorkerReview audit log
# tracks every state transition; the worker itself carries a fast-path status.
# ═════════════════════════════════════════════════════════════════════════════

class _ReviewBody(BaseModel):
    comment: Optional[str] = None


class _OverrideBody(BaseModel):
    approved: bool


class _PermitAreaDecisionBody(BaseModel):
    """Per-area decision payload for the work-permit approval workflow.
    `area_ids` restricts the action to a subset of areas the user supervises;
    if omitted, every supervised PENDING area approval is acted on."""
    area_ids: Optional[List[int]] = None
    comment: Optional[str] = None


class _PermitExtensionBody(BaseModel):
    """Request payload for extending an APPROVED work permit.
    `end_date` is the new finish date (must be > current end_date)."""
    end_date: str
    comment: Optional[str] = None
    comment: Optional[str] = None


@router.get("/workers/{worker_id}/history")
def worker_history(worker_id: int,
                   db: Session = Depends(get_db),
                   user: auth.ProjectContext = Depends(auth.get_project_user)):
    w = db.query(models.Worker).filter_by(id=worker_id, project_id=user.project_id).first()
    if not w:
        raise HTTPException(404, "Worker not found")
    rows = (db.query(models.WorkerReview)
              .filter_by(worker_id=worker_id)
              .order_by(models.WorkerReview.created_at.asc())
              .all())
    return [{
        "id": r.id, "event": r.event, "approved": r.approved,
        "comment": r.comment, "actor_id": r.actor_id,
        "actor_name": r.actor.name if r.actor else None,
        "created_at": r.created_at.isoformat() + 'Z' if r.created_at else None,
    } for r in rows]


@router.get("/workers/pending-approval")
def workers_pending_approval(for_action_points: bool = False,
                             db: Session = Depends(get_db),
                             user: auth.ProjectContext = Depends(auth.get_project_user)):
    """PENDING worker queue.
    • Default: visible to ADMIN, PROJECT_OWNER and to any user registered as a
      site supervisor on at least one area of the project. ADMINs and
      PROJECT_OWNERs need to see the list so they can override.
    • `?for_action_points=true` (used by My Action Points): restricted to
      actual site supervisors only — admins/owners who are NOT supervisors
      don't get the notification even though they can still use the tab."""
    is_supervisor = _is_site_supervisor(db, user.project_id, user.contact_id)
    if for_action_points:
        if not is_supervisor:
            return []
    else:
        if not _is_owner_or_admin(user, db) and not is_supervisor:
            return []
    rows = (db.query(models.Worker)
              .filter_by(project_id=user.project_id, status="PENDING")
              .order_by(models.Worker.submitted_at.desc().nullslast(), models.Worker.id.desc())
              .all())
    return [_fmt_worker(w) for w in rows]


@router.get("/workers/my-rejected")
def workers_my_rejected(db: Session = Depends(get_db),
                        user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Rejected workers whose package the current user is linked to via
    PackageContact (so they can re-open and resubmit or cancel). Admins /
    project owners do NOT receive this action point automatically — they
    only see it if they themselves are a PackageContact on the package."""
    if user.role == "BIDDER":
        return []
    if not user.contact_id:
        return []
    pkg_ids = _vendor_package_ids(user, db)
    if not pkg_ids:
        return []
    rows = (db.query(models.Worker)
              .filter(models.Worker.project_id == user.project_id,
                      models.Worker.status == "REJECTED",
                      models.Worker.package_id.in_(pkg_ids))
              .order_by(models.Worker.reviewed_at.desc().nullslast(),
                        models.Worker.id.desc())
              .all())
    return [_fmt_worker(w) for w in rows]


@router.post("/workers/{worker_id}/approve")
def approve_worker(worker_id: int, body: _ReviewBody,
                   db: Session = Depends(get_db),
                   user: auth.ProjectContext = Depends(auth.get_project_user)):
    w = db.query(models.Worker).filter_by(id=worker_id, project_id=user.project_id).first()
    if not w:
        raise HTTPException(404, "Worker not found")
    if w.status != "PENDING":
        raise HTTPException(400, "Worker is not pending approval")
    if not _can_review_worker(user, w, db):
        raise HTTPException(403, "Only a site supervisor can approve this worker")
    now = datetime.utcnow()
    w.status = "APPROVED"
    w.reviewed_at = now
    w.reviewed_by_id = user.id
    w.rejection_comment = None
    _log_worker_review(db, w, "APPROVE", user, approved=True,
                       comment=(body.comment or "").strip() or None)
    db.commit(); db.refresh(w)
    return _fmt_worker(w)


@router.post("/workers/{worker_id}/reject")
def reject_worker(worker_id: int, body: _ReviewBody,
                  db: Session = Depends(get_db),
                  user: auth.ProjectContext = Depends(auth.get_project_user)):
    w = db.query(models.Worker).filter_by(id=worker_id, project_id=user.project_id).first()
    if not w:
        raise HTTPException(404, "Worker not found")
    if w.status != "PENDING":
        raise HTTPException(400, "Worker is not pending approval")
    if not _can_review_worker(user, w, db):
        raise HTTPException(403, "Only a site supervisor can reject this worker")
    comment = (body.comment or "").strip()
    if not comment:
        raise HTTPException(400, "A rejection comment is required")
    now = datetime.utcnow()
    w.status = "REJECTED"
    w.reviewed_at = now
    w.reviewed_by_id = user.id
    w.rejection_comment = comment
    _log_worker_review(db, w, "REJECT", user, approved=False, comment=comment)
    db.commit(); db.refresh(w)
    return _fmt_worker(w)


@router.post("/workers/{worker_id}/override")
def override_worker(worker_id: int, body: _OverrideBody,
                    db: Session = Depends(get_db),
                    user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Admins / project owners / construction module leads / the package owner
    can override a PENDING worker review — either authorise or reject directly
    without being a site supervisor."""
    w = db.query(models.Worker).filter_by(id=worker_id, project_id=user.project_id).first()
    if not w:
        raise HTTPException(404, "Worker not found")
    pkg = db.query(models.Package).filter_by(id=w.package_id).first() if w.package_id else None
    gate = auth.package_access_path(user, "Construction", pkg, db)
    if not gate:
        raise HTTPException(403, "Only Admins, Project Owners, Module Leads or the Package Owner can override")
    if w.status != "PENDING":
        raise HTTPException(400, "Can only override PENDING workers")
    now = datetime.utcnow()
    comment = (body.comment or "").strip() or auth.override_default_comment(user.name, gate)
    w.status = "APPROVED" if body.approved else "REJECTED"
    w.reviewed_at = now
    w.reviewed_by_id = user.id
    w.rejection_comment = None if body.approved else comment
    _log_worker_review(db, w, "OVERRIDE", user,
                       approved=bool(body.approved), comment=comment)
    db.commit(); db.refresh(w)
    return _fmt_worker(w)


@router.post("/workers/{worker_id}/resubmit")
def resubmit_worker(worker_id: int, body: _ReviewBody,
                    db: Session = Depends(get_db),
                    user: auth.ProjectContext = Depends(auth.get_project_user)):
    w = db.query(models.Worker).filter_by(id=worker_id, project_id=user.project_id).first()
    if not w:
        raise HTTPException(404, "Worker not found")
    if w.status not in ("REJECTED", "CANCELLED"):
        raise HTTPException(400, "Only a rejected or cancelled worker can be resubmitted")
    if not _can_manage_worker(user, w, db):
        raise HTTPException(403, "Not authorised for this package")
    w.status = "PENDING"
    w.submitted_at = datetime.utcnow()
    w.reviewed_at = None
    w.reviewed_by_id = None
    _log_worker_review(db, w, "RESUBMIT", user,
                       comment=(body.comment or "").strip() or None)
    db.commit(); db.refresh(w)
    return _fmt_worker(w)


@router.post("/workers/{worker_id}/cancel")
def cancel_worker(worker_id: int, body: _ReviewBody,
                  db: Session = Depends(get_db),
                  user: auth.ProjectContext = Depends(auth.get_project_user)):
    w = db.query(models.Worker).filter_by(id=worker_id, project_id=user.project_id).first()
    if not w:
        raise HTTPException(404, "Worker not found")
    if w.status == "CANCELLED":
        raise HTTPException(400, "Worker already cancelled")
    if not _can_manage_worker(user, w, db):
        raise HTTPException(403, "Not authorised for this package")
    w.status = "CANCELLED"
    _log_worker_review(db, w, "CANCEL", user,
                       comment=(body.comment or "").strip() or None)
    db.commit(); db.refresh(w)
    return _fmt_worker(w)


# ═════════════════════════════════════════════════════════════════════════════
# WORK LOGS — declare work periods per package (owners + team members only)
# ═════════════════════════════════════════════════════════════════════════════

def _can_manage_work_logs(user: auth.ProjectContext) -> bool:
    return user.role in ("ADMIN", "PROJECT_OWNER", "PROJECT_TEAM")


@router.get("/work-logs")
def list_work_logs(package_id: Optional[int] = None,
                   db: Session = Depends(get_db),
                   user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    q = db.query(models.WorkLog).filter_by(project_id=user.project_id)
    if package_id:
        q = q.filter(models.WorkLog.package_id == package_id)
    rows = q.order_by(models.WorkLog.start_date.desc(), models.WorkLog.id.desc()).all()
    return [_fmt_work_log(r) for r in rows]


@router.post("/work-logs")
def create_work_log(body: WorkLogCreate,
                    db: Session = Depends(get_db),
                    user: auth.ProjectContext = Depends(auth.get_project_user)):
    if not _can_manage_work_logs(user):
        raise HTTPException(403, "Only project owners and team members can add work logs")
    if not body.start_date:
        raise HTTPException(400, "Start date is required")
    pkg = db.query(models.Package).filter_by(id=body.package_id, project_id=user.project_id).first()
    if not pkg:
        raise HTTPException(404, "Package not found")
    if body.end_date and body.end_date < body.start_date:
        raise HTTPException(400, "End date cannot be before start date")
    wl = models.WorkLog(
        project_id=user.project_id,
        package_id=body.package_id,
        start_date=body.start_date,
        end_date=body.end_date or None,
        notes=(body.notes or "").strip() or None,
        ignore_missing_reports=bool(body.ignore_missing_reports),
        created_by_id=user.id,
    )
    db.add(wl); db.commit(); db.refresh(wl)
    return _fmt_work_log(wl)


@router.put("/work-logs/{log_id}")
def update_work_log(log_id: int, body: WorkLogUpdate,
                    db: Session = Depends(get_db),
                    user: auth.ProjectContext = Depends(auth.get_project_user)):
    if not _can_manage_work_logs(user):
        raise HTTPException(403, "Only project owners and team members can edit work logs")
    wl = db.query(models.WorkLog).filter_by(id=log_id, project_id=user.project_id).first()
    if not wl:
        raise HTTPException(404, "Work log not found")
    pkg = db.query(models.Package).filter_by(id=body.package_id, project_id=user.project_id).first()
    if not pkg:
        raise HTTPException(404, "Package not found")
    if body.end_date and body.start_date and body.end_date < body.start_date:
        raise HTTPException(400, "End date cannot be before start date")
    wl.package_id = body.package_id
    wl.start_date = body.start_date
    wl.end_date = body.end_date or None
    wl.notes = (body.notes or "").strip() or None
    wl.ignore_missing_reports = bool(body.ignore_missing_reports)
    wl.updated_by_id = user.id
    db.commit(); db.refresh(wl)
    return _fmt_work_log(wl)


@router.delete("/work-logs/{log_id}")
def delete_work_log(log_id: int,
                    db: Session = Depends(get_db),
                    user: auth.ProjectContext = Depends(auth.get_project_user)):
    if not _can_manage_work_logs(user):
        raise HTTPException(403, "Only project owners and team members can delete work logs")
    wl = db.query(models.WorkLog).filter_by(id=log_id, project_id=user.project_id).first()
    if not wl:
        raise HTTPException(404, "Work log not found")
    db.delete(wl); db.commit()
    return {"ok": True}


# ═════════════════════════════════════════════════════════════════════════════
# DAILY REPORTS — vendors report per-package daily works. Pending-report
# action points are computed from WorkLog windows up to today.
# ═════════════════════════════════════════════════════════════════════════════

def _fmt_daily_report(r: models.DailyReport, db: Session) -> dict:
    workers = []
    for link in (r.workers or []):
        w = link.worker
        if w:
            workers.append({"id": w.id, "name": w.name,
                            "display_id": f"WK-{(w.project_seq_id or w.id):06d}"})
    areas = []
    for link in (r.areas or []):
        a = link.area
        if a:
            areas.append({"id": a.id, "tag": a.tag, "description": a.description})
    return {
        "id": r.id,
        "project_id": r.project_id,
        "package_id": r.package_id,
        "package_tag": r.package.tag_number if r.package else None,
        "package_name": r.package.name if r.package else None,
        "report_date": r.report_date,
        "description": r.description or "",
        "avg_hours_per_worker": r.avg_hours_per_worker or 0.0,
        "expected_worker_count": r.expected_worker_count,
        "no_work": bool(r.no_work),
        "workers": workers,
        "worker_ids": [w["id"] for w in workers],
        "areas": areas,
        "area_ids": [a["id"] for a in areas],
        "total_hours": round((r.avg_hours_per_worker or 0.0) * len(workers), 2),
        "created_at": r.created_at.isoformat() + 'Z' if r.created_at else None,
        "created_by_name": r.created_by.name if r.created_by else None,
        "locked": bool(r.locked) if r.locked is not None else True,
        "locked_at": r.locked_at.isoformat() + 'Z' if r.locked_at else None,
        "unlocked_at": r.unlocked_at.isoformat() + 'Z' if r.unlocked_at else None,
        "unlocked_by_name": r.unlocked_by.name if r.unlocked_by else None,
        "unlock_comment": r.unlock_comment or "",
    }


def _expected_report_dates(db: Session, project_id: int, package_id: int) -> set:
    """All YYYY-MM-DD dates on which a report is expected for the package,
    derived from its WorkLog windows (future dates excluded). Work periods
    flagged `ignore_missing_reports` are skipped — those days never count
    as missing."""
    today = date.today()
    out = set()
    logs = db.query(models.WorkLog).filter_by(project_id=project_id, package_id=package_id).all()
    for wl in logs:
        if getattr(wl, "ignore_missing_reports", False):
            continue
        try:
            start = date.fromisoformat(wl.start_date) if wl.start_date else None
        except Exception:
            start = None
        if not start:
            continue
        try:
            end = date.fromisoformat(wl.end_date) if wl.end_date else today
        except Exception:
            end = today
        # Clamp to today
        if end > today: end = today
        if start > today: continue
        cur = start
        while cur <= end:
            out.add(cur.isoformat())
            cur += timedelta(days=1)
    return out


@router.get("/dashboard/active-workers")
def dashboard_active_workers(
    package_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    """Lightweight series for the Construction dashboard chart:
    [{date, count}] of unique workers declared per day. Replaces the
    previous client-side aggregation over the full daily-reports payload."""
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")

    # Reports in scope (project + optional package + vendor restriction)
    reports_q = db.query(models.DailyReport.id, models.DailyReport.report_date).filter(
        models.DailyReport.project_id == user.project_id
    )
    if package_id:
        reports_q = reports_q.filter(models.DailyReport.package_id == package_id)
    if user.role == "VENDOR" and not _is_owner_or_admin(user, db):
        pkg_ids = _vendor_package_ids(user, db)
        if not pkg_ids:
            return []
        reports_q = reports_q.filter(models.DailyReport.package_id.in_(pkg_ids))

    # JOIN to workers and group by date in SQL — one query, ~N-days rows back.
    rows = (
        db.query(
            models.DailyReport.report_date,
            func.count(distinct(models.DailyReportWorker.worker_id)),
        )
          .outerjoin(models.DailyReportWorker,
                     models.DailyReportWorker.daily_report_id == models.DailyReport.id)
          .filter(models.DailyReport.id.in_(reports_q.with_entities(models.DailyReport.id)))
          .group_by(models.DailyReport.report_date)
          .order_by(models.DailyReport.report_date)
          .all()
    )
    return [{"date": r[0], "count": int(r[1] or 0)} for r in rows]


@router.get("/daily-reports")
def list_daily_reports(package_id: Optional[int] = None,
                       area_id: Optional[int] = None,
                       db: Session = Depends(get_db),
                       user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    q = (
        db.query(models.DailyReport)
          .options(
              joinedload(models.DailyReport.package),
              joinedload(models.DailyReport.created_by),
              joinedload(models.DailyReport.unlocked_by),
              selectinload(models.DailyReport.workers).joinedload(models.DailyReportWorker.worker),
              selectinload(models.DailyReport.areas).joinedload(models.DailyReportArea.area),
          )
          .filter_by(project_id=user.project_id)
    )
    if package_id:
        q = q.filter(models.DailyReport.package_id == package_id)
    if area_id:
        q = q.join(models.DailyReportArea,
                   models.DailyReportArea.daily_report_id == models.DailyReport.id)\
             .filter(models.DailyReportArea.area_id == area_id)
    # Vendors only see their linked packages
    if user.role == "VENDOR" and not _is_owner_or_admin(user, db):
        pkg_ids = _vendor_package_ids(user, db)
        if not pkg_ids:
            return []
        q = q.filter(models.DailyReport.package_id.in_(pkg_ids))
    rows = q.order_by(models.DailyReport.report_date.desc(),
                      models.DailyReport.id.desc()).all()
    return [_fmt_daily_report(r, db) for r in rows]


def _validate_report_body(body: DailyReportBody):
    if not body.report_date:
        raise HTTPException(400, "Report date is required")
    # no_work / all-zero shortcut: no mandatory fields beyond date + package
    if body.no_work or (body.avg_hours_per_worker == 0
                        and not body.worker_ids and not body.area_ids
                        and not (body.description or "").strip()):
        return
    # Otherwise everything is mandatory
    if body.avg_hours_per_worker is None or body.avg_hours_per_worker <= 0:
        raise HTTPException(400, "Average hours per worker must be greater than 0")
    if not body.worker_ids:
        raise HTTPException(400, "Select at least one worker")
    if not body.area_ids:
        raise HTTPException(400, "Select at least one area")
    if not (body.description or "").strip():
        raise HTTPException(400, "Description is required")


def _apply_daily_report_links(db: Session, report: models.DailyReport,
                              worker_ids: List[int], area_ids: List[int]):
    # Filter provided IDs to those belonging to the correct project & package
    if worker_ids:
        valid_ws = {w.id for w in db.query(models.Worker).filter(
            models.Worker.project_id == report.project_id,
            models.Worker.id.in_(worker_ids)).all()}
    else:
        valid_ws = set()
    if area_ids:
        valid_as = {a.id for a in db.query(models.Area).filter(
            models.Area.project_id == report.project_id,
            models.Area.id.in_(area_ids)).all()}
    else:
        valid_as = set()
    db.query(models.DailyReportWorker).filter_by(daily_report_id=report.id).delete()
    db.query(models.DailyReportArea).filter_by(daily_report_id=report.id).delete()
    db.flush()
    for wid in valid_ws:
        db.add(models.DailyReportWorker(daily_report_id=report.id, worker_id=wid))
    for aid in valid_as:
        db.add(models.DailyReportArea(daily_report_id=report.id, area_id=aid))


@router.post("/daily-reports")
def create_daily_report(body: DailyReportBody,
                        db: Session = Depends(get_db),
                        user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    if not _user_can_manage_package(user, body.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    _validate_report_body(body)
    existing = db.query(models.DailyReport).filter_by(
        project_id=user.project_id, package_id=body.package_id,
        report_date=body.report_date,
    ).first()
    if existing:
        raise HTTPException(400, "A report already exists for that package and date")
    r = models.DailyReport(
        project_id=user.project_id,
        package_id=body.package_id,
        report_date=body.report_date,
        description=(body.description or "").strip() or None,
        avg_hours_per_worker=float(body.avg_hours_per_worker or 0.0),
        expected_worker_count=(int(body.expected_worker_count)
                               if body.expected_worker_count is not None
                               else None),
        no_work=bool(body.no_work) or (
            body.avg_hours_per_worker == 0 and not body.worker_ids and not body.area_ids
            and not (body.description or "").strip()
        ),
        created_by_id=user.id,
        locked=True,
        locked_at=datetime.utcnow(),
    )
    db.add(r); db.flush()
    _apply_daily_report_links(db, r, body.worker_ids, body.area_ids)
    db.commit(); db.refresh(r)
    return _fmt_daily_report(r, db)


@router.put("/daily-reports/{report_id}")
def update_daily_report(report_id: int, body: DailyReportBody,
                        db: Session = Depends(get_db),
                        user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    r = db.query(models.DailyReport).filter_by(id=report_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Daily report not found")
    if not _user_can_manage_package(user, r.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    if r.locked:
        raise HTTPException(403, "Daily report is locked — ask a project owner, admin or site supervisor to re-open it")
    _validate_report_body(body)
    # Date + package cannot change after creation (simpler for uniqueness)
    r.description = (body.description or "").strip() or None
    r.avg_hours_per_worker = float(body.avg_hours_per_worker or 0.0)
    r.expected_worker_count = (int(body.expected_worker_count)
                               if body.expected_worker_count is not None
                               else None)
    r.no_work = bool(body.no_work) or (
        body.avg_hours_per_worker == 0 and not body.worker_ids and not body.area_ids
        and not (body.description or "").strip()
    )
    r.updated_by_id = user.id
    # Re-lock immediately on the vendor's save.
    r.locked = True
    r.locked_at = datetime.utcnow()
    _apply_daily_report_links(db, r, body.worker_ids, body.area_ids)
    db.commit(); db.refresh(r)
    return _fmt_daily_report(r, db)


@router.delete("/daily-reports/{report_id}")
def delete_daily_report(report_id: int,
                        db: Session = Depends(get_db),
                        user: auth.ProjectContext = Depends(auth.get_project_user)):
    r = db.query(models.DailyReport).filter_by(id=report_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Daily report not found")
    if not _user_can_manage_package(user, r.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    # While locked, only owners/admins/supervisors can delete a submitted report.
    if r.locked and not (_is_owner_or_admin(user, db)
                         or _is_site_supervisor(db, user.project_id, user.contact_id)):
        raise HTTPException(403, "Daily report is locked — ask a project owner, admin or site supervisor to re-open it first")
    db.delete(r); db.commit()
    return {"ok": True}


class _UnlockBody(BaseModel):
    comment: Optional[str] = None


@router.post("/daily-reports/{report_id}/unlock")
def unlock_daily_report(report_id: int, body: _UnlockBody,
                        db: Session = Depends(get_db),
                        user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Re-open a locked daily report so the vendor can edit & resubmit.
    Allowed for project owners, admins and site supervisors only."""
    r = db.query(models.DailyReport).filter_by(id=report_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Daily report not found")
    if not (_is_owner_or_admin(user, db)
            or _is_site_supervisor(db, user.project_id, user.contact_id)):
        raise HTTPException(403, "Only project owners, admins or site supervisors can re-open a daily report")
    if not r.locked:
        raise HTTPException(400, "Daily report is already unlocked")
    r.locked = False
    r.unlocked_at = datetime.utcnow()
    r.unlocked_by_id = user.id
    r.unlock_comment = (body.comment or "").strip() or None
    db.commit(); db.refresh(r)
    return _fmt_daily_report(r, db)


@router.get("/daily-reports/pending")
def pending_daily_reports(db: Session = Depends(get_db),
                          user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Per-package pending dates that still need a daily report. Visible to
    package-linked contacts (via PackageContact) or package owners. Admins
    only see entries for packages where they themselves are a contact/owner."""
    if user.role == "BIDDER" or not user.contact_id:
        return []
    cid = user.contact_id
    # Packages the user can manage (linked contact or owner)
    linked = {r.package_id for r in
              db.query(models.PackageContact).filter_by(contact_id=cid).all()}
    owned = {p.id for p in
             db.query(models.Package).filter(
                 models.Package.project_id == user.project_id,
                 models.Package.package_owner_id == cid,
             ).all()}
    pkg_ids = list(linked | owned)
    if not pkg_ids:
        return []
    # Map package -> existing report dates
    existing_rows = db.query(models.DailyReport.package_id, models.DailyReport.report_date).filter(
        models.DailyReport.project_id == user.project_id,
        models.DailyReport.package_id.in_(pkg_ids),
    ).all()
    existing = {}
    for pid, d in existing_rows:
        existing.setdefault(pid, set()).add(d)

    results = []
    packages_by_id = {p.id: p for p in
                      db.query(models.Package).filter(models.Package.id.in_(pkg_ids)).all()}
    for pid in pkg_ids:
        pkg = packages_by_id.get(pid)
        if not pkg:
            continue
        expected = _expected_report_dates(db, user.project_id, pid)
        missing = sorted(expected - existing.get(pid, set()))
        for d in missing:
            results.append({
                "package_id": pid,
                "package_tag": pkg.tag_number,
                "package_name": pkg.name,
                "report_date": d,
            })
    results.sort(key=lambda x: (x["report_date"], x["package_tag"] or ""))
    return results


# ═════════════════════════ WORK PERMITS ═════════════════════════════════════

def _fmt_loto(l: models.LOTO) -> dict:
    return {
        "id": l.id,
        "project_seq_id": l.project_seq_id,
        "display_id": f"LT-{(l.project_seq_id or l.id):06d}",
        "work_permit_id": l.work_permit_id,
        "work_permit_display_id": (f"WP-{(l.work_permit.project_seq_id or l.work_permit.id):06d}"
                                   if l.work_permit else None),
        "work_permit_title": l.work_permit.title if l.work_permit else None,
        "package_id": l.work_permit.package_id if l.work_permit else None,
        "package_tag": (l.work_permit.package.tag_number
                        if l.work_permit and l.work_permit.package else None),
        "package_name": (l.work_permit.package.name
                         if l.work_permit and l.work_permit.package else None),
        "tag_number": l.tag_number,
        "description": l.description or "",
        "status": l.status,
        "locked_state": bool(l.locked_state),
        "submitted_at": l.submitted_at.isoformat() + 'Z' if l.submitted_at else None,
        "reviewed_at":  l.reviewed_at.isoformat()  + 'Z' if l.reviewed_at  else None,
        "reviewed_by_name": l.reviewed_by.name if l.reviewed_by else None,
        "refusal_comment": l.refusal_comment or "",
        "created_at": l.created_at.isoformat() + 'Z' if l.created_at else None,
        "created_by_name": l.created_by.name if l.created_by else None,
    }


def _log_loto_review(db: Session, loto: models.LOTO, event: str,
                     actor: Optional[auth.ProjectContext] = None,
                     confirmed: Optional[bool] = None,
                     comment: Optional[str] = None):
    db.add(models.LOTOReview(
        loto_id=loto.id, event=event, confirmed=confirmed,
        comment=(comment or None), actor_id=(actor.id if actor else None),
    ))


def _permit_area_ids(permit: models.WorkPermit) -> List[int]:
    return [l.area_id for l in (permit.areas or [])]


def _is_supervisor_on_permit_areas(db: Session,
                                    user: auth.ProjectContext,
                                    permit: models.WorkPermit) -> bool:
    if not user.contact_id:
        return False
    area_ids = _permit_area_ids(permit)
    if not area_ids:
        return False
    row = (db.query(models.AreaSiteSupervisor)
             .filter(models.AreaSiteSupervisor.area_id.in_(area_ids),
                     models.AreaSiteSupervisor.contact_id == user.contact_id)
             .first())
    return row is not None


def _can_review_loto(user: auth.ProjectContext, loto: models.LOTO,
                     db: Session) -> bool:
    """Any declared site supervisor on the project may confirm/refuse any
    LOTO, regardless of the permit's areas. This is intentionally looser
    than work-permit approval (which is strictly area-scoped)."""
    if _is_owner_or_admin(user, db):
        return True
    return _is_project_site_supervisor(user, db)


def _can_manage_loto(user: auth.ProjectContext, loto: models.LOTO,
                     db: Session) -> bool:
    if _is_owner_or_admin(user, db):
        return True
    permit = loto.work_permit
    if not permit:
        return False
    return _user_can_manage_package(user, permit.package_id, db)


def _permit_loto_rollup(permit: models.WorkPermit) -> str:
    """Match the frontend rollup: 'NA' | 'REFUSED' | 'DONE' | 'IN PROGRESS'."""
    arr = list(permit.lotos or [])
    if not arr:
        return "NA"
    if any(l.status == "REFUSED" for l in arr):
        return "REFUSED"
    if all(l.status in ("LOCKED", "CANCELLED") for l in arr):
        return "DONE"
    return "IN PROGRESS"


def _fmt_work_permit_area_approval(ap: models.WorkPermitAreaApproval) -> dict:
    return {
        "id": ap.id,
        "area_id": ap.area_id,
        "area_tag": ap.area.tag if ap.area else None,
        "area_description": ap.area.description if ap.area else None,
        "status": ap.status,
        "reviewed_at": ap.reviewed_at.isoformat() + 'Z' if ap.reviewed_at else None,
        "reviewed_by_id": ap.reviewed_by_id,
        "reviewed_by_name": ap.reviewed_by.name if ap.reviewed_by else None,
        "rejection_comment": ap.rejection_comment or "",
    }


def _fmt_work_permit(r: models.WorkPermit, db: Session) -> dict:
    permit_types = []
    for link in (r.permit_types or []):
        pt = link.permit_type
        if pt:
            permit_types.append({"id": pt.id, "name": pt.name})
    areas = []
    for link in (r.areas or []):
        a = link.area
        if a:
            areas.append({"id": a.id, "tag": a.tag, "description": a.description})
    hazards = [{"hazard_key": h.hazard_key,
                "preventive_measure": h.preventive_measure or ""}
               for h in (r.hazards or [])]
    ppes = [p.ppe_key for p in (r.ppes or [])]
    approvals = [_fmt_work_permit_area_approval(ap)
                 for ap in sorted((r.area_approvals or []),
                                  key=lambda x: x.area_id)]
    # `pending_kind` tells the UI whether the PENDING state represents a
    # first submission / resubmission after rejection, or an extension
    # request. Derived from the most recent submission-family event.
    pending_kind = None
    for h in sorted((r.review_history or []),
                    key=lambda x: (x.created_at or datetime.min, x.id or 0),
                    reverse=True):
        if h.event in ("SUBMIT", "RESUBMIT", "EXTEND"):
            pending_kind = h.event
            break
    return {
        "id": r.id,
        "project_seq_id": r.project_seq_id,
        "display_id": f"WP-{(r.project_seq_id or r.id):06d}",
        "project_id": r.project_id,
        "package_id": r.package_id,
        "package_tag": r.package.tag_number if r.package else None,
        "package_name": r.package.name if r.package else None,
        "title": r.title or "",
        "description": r.description or "",
        "start_date": r.start_date,
        "end_date": r.end_date,
        "permit_types": permit_types,
        "permit_type_ids": [pt["id"] for pt in permit_types],
        "areas": areas,
        "area_ids": [a["id"] for a in areas],
        "hazards": hazards,
        "hazard_keys": [h["hazard_key"] for h in hazards],
        "hazards_other": r.hazards_other or "",
        "ppe_keys": ppes,
        "ppe_other": r.ppe_other or "",
        "lotos": [_fmt_loto(l) for l in sorted(
            (r.lotos or []), key=lambda x: (x.project_seq_id or x.id)
        )],
        "loto_rollup": _permit_loto_rollup(r),
        "status": r.status or "DRAFT",
        "pending_kind": pending_kind,
        "area_approvals": approvals,
        "submitted_at": r.submitted_at.isoformat() + 'Z' if r.submitted_at else None,
        "submitted_by_id": r.submitted_by_id,
        "submitted_by_name": r.submitted_by.name if r.submitted_by else None,
        "created_at": r.created_at.isoformat() + 'Z' if r.created_at else None,
        "created_by_name": r.created_by.name if r.created_by else None,
    }


def _validate_work_permit_body(body: WorkPermitBody):
    if not (body.title or "").strip():
        raise HTTPException(400, "Title is required")
    if not body.start_date or not body.end_date:
        raise HTTPException(400, "Start and finish dates are required")
    if body.end_date < body.start_date:
        raise HTTPException(400, "Finish date must be on or after the start date")
    if not (body.description or "").strip():
        raise HTTPException(400, "Description of the work is required")
    if not body.permit_type_ids:
        raise HTTPException(400, "Select at least one permit type")
    if not body.area_ids:
        raise HTTPException(400, "Select at least one area")
    # Every checked hazard must have a preventive measure
    for h in (body.hazards or []):
        if not (h.preventive_measure or "").strip():
            raise HTTPException(400, f"A preventive measure is required for hazard '{h.hazard_key}'")


def _apply_work_permit_links(db: Session, permit: models.WorkPermit,
                             body: WorkPermitBody):
    # Scope all links to the correct project
    valid_pts = {pt.id for pt in db.query(models.WorkPermitType).filter(
        models.WorkPermitType.project_id == permit.project_id,
        models.WorkPermitType.id.in_(body.permit_type_ids or []),
    ).all()} if body.permit_type_ids else set()
    valid_as = {a.id for a in db.query(models.Area).filter(
        models.Area.project_id == permit.project_id,
        models.Area.id.in_(body.area_ids or []),
    ).all()} if body.area_ids else set()

    db.query(models.WorkPermitPermitType).filter_by(work_permit_id=permit.id).delete()
    db.query(models.WorkPermitArea).filter_by(work_permit_id=permit.id).delete()
    db.query(models.WorkPermitHazard).filter_by(work_permit_id=permit.id).delete()
    db.query(models.WorkPermitPPE).filter_by(work_permit_id=permit.id).delete()
    db.flush()

    for pid in valid_pts:
        db.add(models.WorkPermitPermitType(work_permit_id=permit.id, permit_type_id=pid))
    for aid in valid_as:
        db.add(models.WorkPermitArea(work_permit_id=permit.id, area_id=aid))

    seen = set()
    for h in (body.hazards or []):
        key = (h.hazard_key or "").strip()
        if not key or key in seen: continue
        seen.add(key)
        db.add(models.WorkPermitHazard(
            work_permit_id=permit.id, hazard_key=key,
            preventive_measure=(h.preventive_measure or "").strip() or None,
        ))

    seen = set()
    for k in (body.ppe_keys or []):
        key = (k or "").strip()
        if not key or key in seen: continue
        seen.add(key)
        db.add(models.WorkPermitPPE(work_permit_id=permit.id, ppe_key=key))


def _apply_work_permit_lotos(db: Session, permit: models.WorkPermit,
                             body: WorkPermitBody,
                             user: auth.ProjectContext):
    """Sync LOTO rows on a permit against the list in the body.
      • new items (no id) → create in REQUEST, log SUBMIT
      • existing items    → update tag/description unless LOCKED, and apply
                            any per-row `action` ('resubmit' | 'cancel') to
                            drive state transitions from the parent permit
      • omitted items     → delete unless LOCKED (LOCKED LOTOs cannot be
                            detached)
    """
    current = {l.id: l for l in (permit.lotos or [])}
    sent_ids = set()
    for item in (body.lotos or []):
        tag = (item.tag_number or "").strip()
        desc = (item.description or "").strip() or None
        action = (item.action or "").strip().lower() or None
        if not tag:
            raise HTTPException(400, "LOTO tag number is required")
        if item.id and item.id in current:
            sent_ids.add(item.id)
            existing = current[item.id]
            # LOCKED LOTOs are read-only — ignore edits + actions silently
            if existing.status == "LOCKED":
                continue
            existing.tag_number = tag
            existing.description = desc
            if action == "resubmit":
                if existing.status not in ("REFUSED", "CANCELLED"):
                    raise HTTPException(
                        400,
                        "Only a refused or cancelled LOTO can be resubmitted",
                    )
                existing.status = "REQUEST"
                existing.locked_state = False
                existing.submitted_at = datetime.utcnow()
                existing.reviewed_at = None
                existing.reviewed_by_id = None
                existing.refusal_comment = None
                _log_loto_review(db, existing, "RESUBMIT", user, comment=None)
            elif action == "cancel":
                if existing.status == "CANCELLED":
                    continue   # idempotent — nothing to do
                if existing.status not in ("REFUSED", "REQUEST"):
                    raise HTTPException(
                        400,
                        "Only a REQUEST or REFUSED LOTO can be cancelled",
                    )
                existing.status = "CANCELLED"
                existing.locked_state = False
                _log_loto_review(db, existing, "CANCEL", user, comment=None)
        else:
            new_loto = models.LOTO(
                project_seq_id=models.next_project_seq(db, models.LOTO, permit.project_id),
                project_id=permit.project_id,
                work_permit_id=permit.id,
                tag_number=tag,
                description=desc,
                status="REQUEST",
                locked_state=False,
                submitted_at=datetime.utcnow(),
                created_by_id=user.id,
            )
            db.add(new_loto); db.flush()
            _log_loto_review(db, new_loto, "SUBMIT", user, comment=None)
    # Delete omitted non-locked LOTOs
    for lid, existing in current.items():
        if lid not in sent_ids and existing.status != "LOCKED":
            db.delete(existing)


@router.get("/work-permits")
def list_work_permits(package_id: Optional[int] = None,
                      area_id: Optional[int] = None,
                      db: Session = Depends(get_db),
                      user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    q = db.query(models.WorkPermit).filter_by(project_id=user.project_id)
    if package_id:
        q = q.filter(models.WorkPermit.package_id == package_id)
    if area_id:
        q = q.join(models.WorkPermitArea,
                   models.WorkPermitArea.work_permit_id == models.WorkPermit.id)\
             .filter(models.WorkPermitArea.area_id == area_id)
    if (user.role == "VENDOR" and not _is_owner_or_admin(user, db)
            and not _is_project_site_supervisor(user, db)):
        pkg_ids = _vendor_package_ids(user, db)
        if not pkg_ids:
            return []
        q = q.filter(models.WorkPermit.package_id.in_(pkg_ids))
    rows = q.order_by(models.WorkPermit.start_date.desc(),
                      models.WorkPermit.id.desc()).all()
    return [_fmt_work_permit(r, db) for r in rows]


@router.get("/work-permits/pending-approval")
def work_permits_pending_approval(for_action_points: bool = False,
                                  db: Session = Depends(get_db),
                                  user: auth.ProjectContext = Depends(auth.get_project_user)):
    """PENDING work-permits awaiting review.
    • Default: visible to ADMIN, PROJECT_OWNER and any site supervisor.
    • `?for_action_points=true` restricts to permits where at least one of the
      PENDING area approvals is on an area the current contact supervises.
    Registered before GET /work-permits/{permit_id} so the static path wins
    over the int-typed permit_id route matcher."""
    if user.role == "BIDDER":
        return []
    is_supervisor = _is_project_site_supervisor(user, db)
    if not (is_supervisor or _is_owner_or_admin(user, db)):
        return []
    rows = (db.query(models.WorkPermit)
              .filter(models.WorkPermit.project_id == user.project_id,
                      models.WorkPermit.status == "PENDING")
              .order_by(models.WorkPermit.submitted_at.desc().nullslast(),
                        models.WorkPermit.id.desc())
              .all())
    if for_action_points:
        if not user.contact_id:
            return []
        my_areas = _my_supervised_area_ids(user, db)
        if not my_areas:
            return []
        filtered = []
        for r in rows:
            pending_here = [ap for ap in (r.area_approvals or [])
                            if ap.status == "PENDING" and ap.area_id in my_areas]
            if pending_here:
                filtered.append(r)
        rows = filtered
    return [_fmt_work_permit(r, db) for r in rows]


@router.get("/work-permits/my-rejected")
def work_permits_my_rejected(db: Session = Depends(get_db),
                             user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Rejected work-permits whose package the current user is linked to via
    PackageContact — shown as action points so they can edit and resubmit."""
    if user.role == "BIDDER" or not user.contact_id:
        return []
    pkg_ids = _vendor_package_ids(user, db)
    if not pkg_ids:
        return []
    rows = (db.query(models.WorkPermit)
              .filter(models.WorkPermit.project_id == user.project_id,
                      models.WorkPermit.status == "REJECTED",
                      models.WorkPermit.package_id.in_(pkg_ids))
              .order_by(models.WorkPermit.submitted_at.desc().nullslast(),
                        models.WorkPermit.id.desc())
              .all())
    return [_fmt_work_permit(r, db) for r in rows]


@router.get("/work-permits/approved-due")
def work_permits_approved_due(db: Session = Depends(get_db),
                              user: auth.ProjectContext = Depends(auth.get_project_user)):
    """APPROVED permits whose finish date is today or earlier, filtered to
    the current user's packages. Drives the vendor's "Close or Extend"
    action point. Dates are stored as ISO strings so a string compare is
    sufficient for the `<=` check."""
    if user.role == "BIDDER" or not user.contact_id:
        return []
    pkg_ids = _vendor_package_ids(user, db)
    if not pkg_ids:
        return []
    today = date.today().isoformat()
    rows = (db.query(models.WorkPermit)
              .filter(models.WorkPermit.project_id == user.project_id,
                      models.WorkPermit.status == "APPROVED",
                      models.WorkPermit.package_id.in_(pkg_ids),
                      models.WorkPermit.end_date <= today)
              .order_by(models.WorkPermit.end_date.asc(),
                        models.WorkPermit.id.asc())
              .all())
    return [_fmt_work_permit(r, db) for r in rows]


@router.get("/work-permits/{permit_id}")
def get_work_permit(permit_id: int,
                    db: Session = Depends(get_db),
                    user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    if (user.role == "VENDOR" and not _is_owner_or_admin(user, db)
            and not _is_project_site_supervisor(user, db)):
        if r.package_id not in _vendor_package_ids(user, db):
            raise HTTPException(403, "Not authorised for this package")
    return _fmt_work_permit(r, db)


@router.post("/work-permits")
def create_work_permit(body: WorkPermitBody,
                       db: Session = Depends(get_db),
                       user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    if not _user_can_manage_permit_package(user, body.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    _validate_work_permit_body(body)
    r = models.WorkPermit(
        project_seq_id=models.next_project_seq(db, models.WorkPermit, user.project_id),
        project_id=user.project_id,
        package_id=body.package_id,
        title=(body.title or "").strip() or None,
        description=(body.description or "").strip() or None,
        start_date=body.start_date,
        end_date=body.end_date,
        hazards_other=(body.hazards_other or "").strip() or None,
        ppe_other=(body.ppe_other or "").strip() or None,
        created_by_id=user.id,
    )
    db.add(r); db.flush()
    _apply_work_permit_links(db, r, body)
    _apply_work_permit_lotos(db, r, body, user)
    db.commit(); db.refresh(r)
    return _fmt_work_permit(r, db)


@router.put("/work-permits/{permit_id}")
def update_work_permit(permit_id: int, body: WorkPermitBody,
                       db: Session = Depends(get_db),
                       user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    if not _user_can_manage_permit_package(user, r.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    # Lock the permit while under review — except for the deadlock case
    # where at least one LOTO is REFUSED: supervisors cannot approve until
    # LOTOs are NA/DONE, so the vendor needs a way to fix and resubmit.
    if r.status == "PENDING" and not _is_owner_or_admin(user, db):
        if _permit_loto_rollup(r) != "REFUSED":
            raise HTTPException(
                400,
                "Permit is pending approval — cannot be modified until supervisors review",
            )
    if r.status == "APPROVED" and not _is_owner_or_admin(user, db):
        raise HTTPException(400, "Approved permits are read-only")
    _validate_work_permit_body(body)
    # Package cannot change after creation
    r.title = (body.title or "").strip() or None
    r.description = (body.description or "").strip() or None
    r.start_date = body.start_date
    r.end_date = body.end_date
    r.hazards_other = (body.hazards_other or "").strip() or None
    r.ppe_other = (body.ppe_other or "").strip() or None
    r.updated_by_id = user.id
    _apply_work_permit_links(db, r, body)
    _apply_work_permit_lotos(db, r, body, user)
    db.commit(); db.refresh(r)
    return _fmt_work_permit(r, db)


@router.delete("/work-permits/{permit_id}")
def delete_work_permit(permit_id: int,
                       db: Session = Depends(get_db),
                       user: auth.ProjectContext = Depends(auth.get_project_user)):
    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    if not _user_can_manage_permit_package(user, r.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    db.delete(r); db.commit()
    return {"ok": True}


# ── Work-permit approval workflow ────────────────────────────────────────────

def _log_permit_review(db: Session, permit: models.WorkPermit, event: str,
                       actor: Optional[auth.ProjectContext] = None,
                       area_id: Optional[int] = None,
                       approved: Optional[bool] = None,
                       comment: Optional[str] = None):
    db.add(models.WorkPermitReview(
        work_permit_id=permit.id, event=event, area_id=area_id,
        approved=approved, comment=(comment or None),
        actor_id=(actor.id if actor else None),
    ))


def _my_supervised_area_ids(user: auth.ProjectContext, db: Session) -> set:
    if not user.contact_id:
        return set()
    rows = (db.query(models.AreaSiteSupervisor)
              .join(models.Area, models.Area.id == models.AreaSiteSupervisor.area_id)
              .filter(models.Area.project_id == user.project_id,
                      models.AreaSiteSupervisor.contact_id == user.contact_id)
              .all())
    return {r.area_id for r in rows}


def _reset_permit_area_approvals(db: Session, permit: models.WorkPermit):
    """Rebuild the WorkPermitAreaApproval rows to match the permit's current
    areas exactly — one PENDING row per area, wiping any previous decisions.
    Used on submit/resubmit so supervisors start fresh each cycle."""
    for ap in list(permit.area_approvals or []):
        db.delete(ap)
    db.flush()
    area_ids = {l.area_id for l in (permit.areas or [])}
    for aid in sorted(area_ids):
        db.add(models.WorkPermitAreaApproval(
            work_permit_id=permit.id, area_id=aid, status="PENDING",
        ))
    db.flush()


def _roll_up_permit_status(permit: models.WorkPermit) -> str:
    """Derive the permit status from its per-area approvals. Called after
    each approve/reject to update permit.status without touching DRAFT."""
    approvals = list(permit.area_approvals or [])
    if not approvals:
        return permit.status   # no approvals yet → keep current
    statuses = {a.status for a in approvals}
    if "REJECTED" in statuses:
        return "REJECTED"
    if statuses == {"APPROVED"}:
        return "APPROVED"
    return "PENDING"


@router.post("/work-permits/{permit_id}/submit")
def submit_work_permit(permit_id: int, body: _ReviewBody,
                       db: Session = Depends(get_db),
                       user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Submit a DRAFT (or previously REJECTED) permit for site-supervisor
    review. Allowed for anyone who can manage the permit's package (i.e.,
    vendor contact linked to the package, site supervisor, owner, admin).
    Resets per-area approvals to PENDING and logs SUBMIT or RESUBMIT."""
    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    if not _user_can_manage_permit_package(user, r.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    # Normally only DRAFT or REJECTED permits submit. The exception is the
    # LOTO-deadlock recovery path: if the permit is already PENDING and at
    # least one LOTO is REFUSED, the supervisors can't approve until the
    # vendor fixes the LOTOs — so we allow a RESUBMIT from PENDING too,
    # which resets the area approvals and restarts the review cycle.
    loto_refused_deadlock = (
        r.status == "PENDING" and _permit_loto_rollup(r) == "REFUSED"
    )
    if r.status not in ("DRAFT", "REJECTED") and not loto_refused_deadlock:
        raise HTTPException(400, "Only DRAFT or REJECTED permits can be submitted")
    if not (r.areas or []):
        raise HTTPException(400, "The permit must list at least one area before submission")
    is_resubmission = (r.status == "REJECTED" or loto_refused_deadlock)
    _reset_permit_area_approvals(db, r)
    r.status = "PENDING"
    r.submitted_at = datetime.utcnow()
    r.submitted_by_id = user.id
    _log_permit_review(db, r, "RESUBMIT" if is_resubmission else "SUBMIT",
                       user, comment=(body.comment or "").strip() or None)
    db.commit(); db.refresh(r)
    return _fmt_work_permit(r, db)


@router.post("/work-permits/{permit_id}/approve")
def approve_work_permit(permit_id: int, body: _PermitAreaDecisionBody,
                        db: Session = Depends(get_db),
                        user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Approve one or more areas of a PENDING permit — limited to areas the
    current user supervises. Permit transitions to APPROVED when every area
    approval becomes APPROVED. Refused if the permit's LOTO rollup is not
    in {NA, DONE} — vendor must finish LOTO review first."""
    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    if r.status != "PENDING":
        raise HTTPException(400, "Permit is not pending approval")
    rollup = _permit_loto_rollup(r)
    if rollup not in ("NA", "DONE"):
        raise HTTPException(400, "LOTO to be executed before release")
    permit_pkg = db.query(models.Package).filter_by(id=r.package_id).first() if r.package_id else None
    override_gate = auth.package_access_path(user, "Construction", permit_pkg, db)
    is_override = override_gate is not None
    my_areas = _my_supervised_area_ids(user, db)
    if not is_override and not my_areas:
        raise HTTPException(403, "Only a site supervisor can review this permit")
    requested = set(body.area_ids or [])
    acted = 0
    for ap in (r.area_approvals or []):
        if ap.status != "PENDING":
            continue
        if requested and ap.area_id not in requested:
            continue
        if not is_override and ap.area_id not in my_areas:
            continue
        ap.status = "APPROVED"
        ap.reviewed_at = datetime.utcnow()
        ap.reviewed_by_id = user.id
        ap.rejection_comment = None
        log_comment = (body.comment or "").strip() or (
            auth.override_default_comment(user.name, override_gate) if is_override else None
        )
        _log_permit_review(db, r, "OVERRIDE" if is_override else "APPROVE",
                           user, area_id=ap.area_id, approved=True,
                           comment=log_comment)
        acted += 1
    if acted == 0:
        raise HTTPException(400,
            "No pending areas available for this user — either already decided "
            "or the user does not supervise any of the requested areas")
    r.status = _roll_up_permit_status(r)
    db.commit(); db.refresh(r)
    return _fmt_work_permit(r, db)


@router.post("/work-permits/{permit_id}/reject")
def reject_work_permit(permit_id: int, body: _PermitAreaDecisionBody,
                       db: Session = Depends(get_db),
                       user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Reject one or more areas of a PENDING permit. A comment is required.
    The permit transitions to REJECTED immediately — vendor must edit and
    resubmit. Rejecting later areas after the first rejection is idempotent."""
    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    if r.status != "PENDING":
        raise HTTPException(400, "Permit is not pending approval")
    comment = (body.comment or "").strip()
    if not comment:
        raise HTTPException(400, "A rejection comment is required")
    permit_pkg = db.query(models.Package).filter_by(id=r.package_id).first() if r.package_id else None
    override_gate = auth.package_access_path(user, "Construction", permit_pkg, db)
    is_override = override_gate is not None
    my_areas = _my_supervised_area_ids(user, db)
    if not is_override and not my_areas:
        raise HTTPException(403, "Only a site supervisor can review this permit")
    requested = set(body.area_ids or [])
    acted = 0
    for ap in (r.area_approvals or []):
        if ap.status != "PENDING":
            continue
        if requested and ap.area_id not in requested:
            continue
        if not is_override and ap.area_id not in my_areas:
            continue
        ap.status = "REJECTED"
        ap.reviewed_at = datetime.utcnow()
        ap.reviewed_by_id = user.id
        ap.rejection_comment = comment
        _log_permit_review(db, r, "OVERRIDE" if is_override else "REJECT",
                           user, area_id=ap.area_id, approved=False,
                           comment=comment)
        acted += 1
    if acted == 0:
        raise HTTPException(400,
            "No pending areas available for this user — either already decided "
            "or the user does not supervise any of the requested areas")
    r.status = _roll_up_permit_status(r)
    db.commit(); db.refresh(r)
    return _fmt_work_permit(r, db)


@router.post("/work-permits/{permit_id}/close")
def close_work_permit(permit_id: int, body: _ReviewBody,
                      db: Session = Depends(get_db),
                      user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Close an APPROVED permit. All LOCKED LOTOs cascade to TO_BE_RELEASED
    so site supervisors can formally release them. Other LOTO states
    (REFUSED / CANCELLED / RELEASED) are left alone — they are already
    terminal for the active permit window."""
    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    if r.status != "APPROVED":
        raise HTTPException(400, "Only an APPROVED permit can be closed")
    if not _user_can_manage_permit_package(user, r.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    now = datetime.utcnow()
    r.status = "CLOSED"
    r.updated_by_id = user.id
    _log_permit_review(db, r, "CLOSE", user,
                       comment=(body.comment or "").strip() or None)
    for l in (r.lotos or []):
        if l.status == "LOCKED":
            l.status = "TO_BE_RELEASED"
            l.locked_state = False
            l.submitted_at = now
            l.reviewed_at = None
            l.reviewed_by_id = None
            _log_loto_review(db, l, "RELEASE_REQUEST", user,
                             comment="Permit closed — release requested")
    db.commit(); db.refresh(r)
    return _fmt_work_permit(r, db)


@router.post("/work-permits/{permit_id}/request-extension")
def request_permit_extension(permit_id: int, body: _PermitExtensionBody,
                             db: Session = Depends(get_db),
                             user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Extend an APPROVED permit's end date. Flips status back to PENDING
    and resets every area approval so every supervisor re-approves for the
    new window. The history carries an EXTEND event which `pending_kind`
    picks up so the UI can badge the approval as an extension request."""
    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    if r.status != "APPROVED":
        raise HTTPException(400, "Only an APPROVED permit can be extended")
    if not _user_can_manage_permit_package(user, r.package_id, db):
        raise HTTPException(403, "Not authorised for this package")
    new_end = (body.end_date or "").strip()
    if not new_end:
        raise HTTPException(400, "New finish date is required")
    if new_end <= (r.end_date or ""):
        raise HTTPException(400, "New finish date must be after the current finish date")
    old_end = r.end_date
    r.end_date = new_end
    _reset_permit_area_approvals(db, r)
    r.status = "PENDING"
    r.submitted_at = datetime.utcnow()
    r.submitted_by_id = user.id
    comment = (body.comment or "").strip()
    ext_note = f"Extension requested: {old_end} → {new_end}"
    _log_permit_review(db, r, "EXTEND", user,
                       comment=(comment + " | " + ext_note) if comment else ext_note)
    db.commit(); db.refresh(r)
    return _fmt_work_permit(r, db)


@router.get("/work-permits/{permit_id}/history")
def work_permit_history(permit_id: int,
                        db: Session = Depends(get_db),
                        user: auth.ProjectContext = Depends(auth.get_project_user)):
    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    # Tighter access: same check as GET /work-permits/{id}
    if (user.role == "VENDOR" and not _is_owner_or_admin(user, db)
            and not _is_project_site_supervisor(user, db)):
        if r.package_id not in _vendor_package_ids(user, db):
            raise HTTPException(403, "Not authorised for this package")
    rows = (db.query(models.WorkPermitReview)
              .filter_by(work_permit_id=permit_id)
              .order_by(models.WorkPermitReview.created_at.asc(),
                        models.WorkPermitReview.id.asc())
              .all())
    out = []
    for h in rows:
        out.append({
            "id": h.id,
            "event": h.event,
            "area_id": h.area_id,
            "area_tag": h.area.tag if h.area else None,
            "approved": h.approved,
            "comment": h.comment or "",
            "actor_id": h.actor_id,
            "actor_name": h.actor.name if h.actor else None,
            "created_at": h.created_at.isoformat() + 'Z' if h.created_at else None,
        })
    return out


# ── Work-permit PDF export ────────────────────────────────────────────────────

_PDF_LOGO_CANDIDATES = [
    "static/assets/impulse-logo-dark@2x.png",
    "static/assets/impulse-logo-dark.png",
    "static/assets/impulse-logo-dark.svg",
    "static/assets/ips-logo.png",
]

# Keep this in lock-step with the front-end hazard / PPE catalogues
# (construction.js data()). Each entry is (key, label, icon_path). Icons
# come from the "_Supporting files/Construction Management/" folder; they
# are rendered on the PDF for every entry — unselected ones appear greyed out.
_PDF_HAZARD_CATALOG = [
    ("General danger",          "General danger",          "_Supporting files/Construction Management/Hazard symbols/General danger.svg"),
    ("Fire Hazard",             "Fire hazard",             "_Supporting files/Construction Management/Hazard symbols/Fire Hazard.svg"),
    ("Electrical Danger",       "Electrical danger",       "_Supporting files/Construction Management/Hazard symbols/Electrical Danger.svg"),
    ("Corrosive substances",    "Corrosive substances",    "_Supporting files/Construction Management/Hazard symbols/Corrosive substances.svg"),
    ("Toxic substances",        "Toxic substances",        "_Supporting files/Construction Management/Hazard symbols/Toxic substances.svg"),
    ("Hot surface",             "Hot surface",             "_Supporting files/Construction Management/Hazard symbols/Hot surface.svg"),
    ("Crusshing hazard",        "Crushing hazard",         "_Supporting files/Construction Management/Hazard symbols/Crusshing hazard.svg"),
    ("Lifting operations",      "Lifting operations",      "_Supporting files/Construction Management/Hazard symbols/Lifting operations.svg"),
    ("Pressurized cilinders",   "Pressurised cylinders",   "_Supporting files/Construction Management/Hazard symbols/Pressurized cilinders.svg"),
    ("Risk of Falling",         "Risk of falling",         "_Supporting files/Construction Management/Hazard symbols/Risk of Falling.svg"),
    ("Slippery surface",        "Slippery surface",        "_Supporting files/Construction Management/Hazard symbols/Slippery surface.svg"),
]

_PDF_PPE_CATALOG = [
    ("Safety goggles",          "Safety goggles",          "_Supporting files/Construction Management/Protection Symbols/Safety goggles.svg"),
    ("Safety helmet",           "Safety helmet",           "_Supporting files/Construction Management/Protection Symbols/Safety helmet.svg"),
    ("Safety shoes",            "Safety shoes",            "_Supporting files/Construction Management/Protection Symbols/Safety shoes.svg"),
    ("Protective gloves",       "Protective gloves",       "_Supporting files/Construction Management/Protection Symbols/Protective gloves.svg"),
    ("Safety clothing",         "Safety clothing",         "_Supporting files/Construction Management/Protection Symbols/Safety clothing.svg"),
    ("Ear protection",          "Ear protection",          "_Supporting files/Construction Management/Protection Symbols/Ear protection.svg"),
    ("Mask",                    "Dust mask",               "_Supporting files/Construction Management/Protection Symbols/Mask.svg"),
    ("respiratory protection",  "Respiratory protection",  "_Supporting files/Construction Management/Protection Symbols/respiratory protection.svg"),
    ("Harness",                 "Harness",                 "_Supporting files/Construction Management/Protection Symbols/Harness.svg"),
]

_PDF_STATUS_COLORS = {
    "DRAFT":    (100, 116, 139),
    "PENDING":  (180, 120, 0),
    "APPROVED": (30,  130, 60),
    "REJECTED": (180, 30,  30),
    "CLOSED":   (71,  85,  105),
}

_PDF_LOTO_STATUS_COLORS = {
    "REQUEST":        (180, 120, 0),
    "LOCKED":         (30,  130, 60),
    "REFUSED":        (180, 30,  30),
    "CANCELLED":      (100, 116, 139),
    "TO_BE_RELEASED": (234, 88,  12),
    "RELEASED":       (71,  85,  105),
}


def _pdf_status_label(s: str) -> str:
    return {
        "DRAFT": "Draft", "PENDING": "Pending approval",
        "APPROVED": "Approved", "REJECTED": "Rejected", "CLOSED": "Closed",
    }.get(s, s or "—")


def _pdf_loto_status_label(s: str) -> str:
    if s == "TO_BE_RELEASED":
        return "TO BE RELEASED"
    return s or "—"


@router.get("/work-permits/{permit_id}/export-pdf")
def export_work_permit_pdf(
        permit_id: int,
        db: Session = Depends(get_db),
        user: auth.ProjectContext = Depends(auth.get_project_user)):
    """One-page A4 PDF of an approved (or closed) work permit.

    Contents:
      • Platform logo, client name, project code + description, printing date
      • Permit identifier + status pill
      • Package / areas / dates / creator
      • Description of works
      • Permit types
      • Hazards + preventive measures (+ other)
      • PPE icons/keys (+ other)
      • LOTO table
      • Per-area approvers (name + date)
    """
    import io, os
    from fpdf import FPDF, XPos, YPos

    r = db.query(models.WorkPermit).filter_by(id=permit_id, project_id=user.project_id).first()
    if not r:
        raise HTTPException(404, "Work permit not found")
    if r.status not in ("APPROVED", "CLOSED"):
        raise HTTPException(400, "Only approved or closed permits can be exported")
    # Access: same rule as GET /work-permits/{id}
    if (user.role == "VENDOR" and not _is_owner_or_admin(user, db)
            and not _is_project_site_supervisor(user, db)):
        if r.package_id not in _vendor_package_ids(user, db):
            raise HTTPException(403, "Not authorised for this package")

    project = db.query(models.Project).filter_by(id=r.project_id).first()

    # Logo — pick the first one that exists on disk.
    logo_path = None
    for cand in _PDF_LOGO_CANDIDATES:
        if os.path.exists(cand):
            logo_path = cand
            break

    # Colour palette
    IPS_BLUE    = (27,  79,  140)
    ACCENT_BLUE = (0,   174, 239)
    LIGHT_GRAY  = (241, 245, 249)
    BORDER_GRAY = (203, 213, 225)
    TEXT_DARK   = (30,  41,  59)
    TEXT_MUTED  = (100, 116, 139)
    WHITE       = (255, 255, 255)

    NL = dict(new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    ST = dict(new_x=XPos.RIGHT,   new_y=YPos.TOP)

    # Printing time in the project's configured timezone, falling back to
    # the server's local timezone when IANA zone data (tzdata) is not
    # installed — Windows in particular ships without it. The final
    # fallback is the naive server time with no suffix, which still
    # matches the user's wall clock as long as the server runs locally.
    tz_name = "Europe/Brussels"
    try:
        row = (db.query(models.Setting)
                 .filter_by(project_id=user.project_id, key="timezone").first())
        if row and (row.value or "").strip():
            tz_name = row.value.strip()
    except Exception:
        pass
    try:
        from datetime import timezone as _tz
        from zoneinfo import ZoneInfo
        now_local = datetime.now(_tz.utc).astimezone(ZoneInfo(tz_name))
        printing_ts = now_local.strftime("%Y-%m-%d %H:%M ") + (now_local.tzname() or "")
    except Exception:
        # Fall back to the OS-local timezone via naive datetime.now().
        # Works on Windows without tzdata and gives the user's wall time.
        printing_ts = datetime.now().strftime("%Y-%m-%d %H:%M (server time)")

    # Helvetica (the core fpdf font) is Latin-1 only, so any em-dash /
    # bullet / ellipsis from the UI has to be transliterated before it
    # lands in a cell. Anything outside Latin-1 is dropped (rare — we
    # don't author content with exotic characters).
    _REPL = {
        "\u2014": "-", "\u2013": "-",
        "\u2022": "*",
        "\u2026": "...",
        "\u201c": '"', "\u201d": '"', "\u2018": "'", "\u2019": "'",
        "\u00a0": " ",
    }

    def _s(txt):
        if txt is None:
            return ""
        s = str(txt)
        for k, v in _REPL.items():
            if k in s:
                s = s.replace(k, v)
        return s.encode("latin-1", "replace").decode("latin-1")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False, margin=10)   # one-page constraint
    pdf.set_margins(10, 10, 10)
    pdf.add_page()

    W = pdf.w - pdf.l_margin - pdf.r_margin   # 190 mm usable

    def tc(*rgb): pdf.set_text_color(*rgb)
    def fc(*rgb): pdf.set_fill_color(*rgb)
    def dc(*rgb): pdf.set_draw_color(*rgb)

    # ── Header band ──────────────────────────────────────────────────────
    header_h = 18
    fc(*IPS_BLUE); pdf.rect(pdf.l_margin, pdf.t_margin, W, header_h, "F")
    logo_w = 22    # small enough to leave room for "WORK PERMIT"
    if logo_path:
        try:
            pdf.image(logo_path,
                      x=pdf.l_margin + 3, y=pdf.t_margin + 3,
                      w=logo_w, keep_aspect_ratio=True)
        except Exception:
            pass
    tc(*WHITE)
    # Title shifts to clear the smaller logo
    pdf.set_xy(pdf.l_margin + logo_w + 8, pdf.t_margin + 2)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(W - (logo_w + 8) - 60, 6, _s("WORK PERMIT"), **NL)
    pdf.set_x(pdf.l_margin + logo_w + 8)
    pdf.set_font("Helvetica", "", 9)
    client_line = (project.client if project and project.client else "-")
    pdf.cell(W - (logo_w + 8) - 60, 4.5, _s(f"Client: {client_line}"), **NL)
    pdf.set_x(pdf.l_margin + logo_w + 8)
    pcode = (project.project_number if project else "") or "-"
    pdesc = ((project.description if project else "") or "").strip()
    if pdesc and len(pdesc) > 70:
        pdesc = pdesc[:67] + "..."
    pdf.cell(W - (logo_w + 8) - 60, 4.5,
             _s(f"Project: {pcode}" + (f" - {pdesc}" if pdesc else "")), **NL)

    # Printing date + requester (right-aligned block)
    pdf.set_xy(pdf.l_margin + W - 60, pdf.t_margin + 2)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(60, 4.5, _s(f"Printed: {printing_ts}"), align="R", **NL)
    pdf.set_x(pdf.l_margin + W - 60)
    actor_name = user.name if hasattr(user, "name") and user.name else "-"
    pdf.cell(60, 4.5, _s(f"Exported by: {actor_name}"), align="R", **NL)

    pdf.set_xy(pdf.l_margin, pdf.t_margin + header_h + 3)
    tc(*TEXT_DARK)

    # ── Permit id + title banner ─────────────────────────────────────────
    display_id = f"WP-{(r.project_seq_id or r.id):06d}"
    status = r.status or "DRAFT"
    banner_h = 9
    fc(*LIGHT_GRAY); pdf.rect(pdf.l_margin, pdf.get_y(), W, banner_h, "F")
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_xy(pdf.l_margin + 3, pdf.get_y() + 1.5)
    pdf.cell(W - 60, 6, _s(f"{display_id}   {r.title or ''}"), **ST)
    # Status pill on the right
    pill_w, pill_h = 40, 6
    pill_x = pdf.l_margin + W - pill_w - 3
    pill_y = pdf.get_y() - 0.5
    fc(*_PDF_STATUS_COLORS.get(status, (128, 128, 128)))
    pdf.rect(pill_x, pill_y, pill_w, pill_h, "F")
    pdf.set_xy(pill_x, pill_y)
    tc(*WHITE); pdf.set_font("Helvetica", "B", 8)
    pdf.cell(pill_w, pill_h, _s(_pdf_status_label(status).upper()), align="C", **NL)
    pdf.set_xy(pdf.l_margin, pill_y + banner_h + 2)
    tc(*TEXT_DARK)

    # ── Meta grid (two columns) ──────────────────────────────────────────
    pkg_label = f"{r.package.tag_number} - {r.package.name}" if r.package else "-"
    area_labels = ", ".join(f"{link.area.tag}" for link in (r.areas or []) if link.area) or "-"
    type_labels = ", ".join(link.permit_type.name for link in (r.permit_types or []) if link.permit_type) or "-"
    requester_name = r.created_by.name if r.created_by else "-"
    requested_at = r.created_at.strftime("%Y-%m-%d") if r.created_at else "-"
    submitter_name = r.submitted_by.name if r.submitted_by else "-"
    submitted_at = r.submitted_at.strftime("%Y-%m-%d") if r.submitted_at else "-"

    def meta_pair(lbl_a, val_a, lbl_b, val_b):
        half = W / 2
        pdf.set_font("Helvetica", "B", 8); tc(*IPS_BLUE)
        pdf.cell(32, 5, _s(lbl_a + ":"), **ST)
        pdf.set_font("Helvetica", "", 9); tc(*TEXT_DARK)
        pdf.cell(half - 32, 5, _s(val_a), **ST)
        pdf.set_font("Helvetica", "B", 8); tc(*IPS_BLUE)
        pdf.cell(32, 5, _s(lbl_b + ":"), **ST)
        pdf.set_font("Helvetica", "", 9); tc(*TEXT_DARK)
        pdf.cell(half - 32, 5, _s(val_b), **NL)

    meta_pair("Package", pkg_label, "Areas", area_labels)
    meta_pair("Start date", r.start_date or "-", "Finish date", r.end_date or "-")
    # Requested by + Submitted for approval share a single row.
    submitted_val = f"{submitter_name} ({submitted_at})" if r.submitted_by else "-"
    meta_pair("Requested by", f"{requester_name} ({requested_at})",
              "Submitted for approval", submitted_val)
    # Permit types gets its own full-width line so long lists don't wrap.
    pdf.set_font("Helvetica", "B", 8); tc(*IPS_BLUE)
    pdf.cell(32, 5, _s("Permit types:"), **ST)
    pdf.set_font("Helvetica", "", 9); tc(*TEXT_DARK)
    pdf.multi_cell(W - 32, 5, _s(type_labels), **NL)
    pdf.ln(1)

    # ── Description of works ─────────────────────────────────────────────
    def section_title(t):
        pdf.set_font("Helvetica", "B", 10); tc(*IPS_BLUE)
        fc(*LIGHT_GRAY)
        pdf.cell(W, 6, _s(" " + t), fill=True, **NL)
        pdf.ln(1); tc(*TEXT_DARK); pdf.set_font("Helvetica", "", 9)

    section_title("Description of the work")
    desc = (r.description or "-").strip()
    pdf.multi_cell(W, 4.3, _s(desc), **NL)
    pdf.ln(1)

    # ── Icon-grid helper (shared by hazards + PPE) ───────────────────────
    # Renders every catalog entry as a small card: icon on the left,
    # label on the right. Selected entries get a highlighted border and
    # a small "(selected)" tag; unselected ones are drawn with reduced
    # opacity via a light grey overlay so they read as context.
    def _icon_grid(catalog, selected_keys, cols=4, cell_w=None,
                   cell_h=14, selected_color=(30, 130, 60),
                   unselected_color=(203, 213, 225)):
        if cell_w is None:
            cell_w = W / cols
        x0 = pdf.l_margin
        y0 = pdf.get_y()
        for i, (key, label, icon_path) in enumerate(catalog):
            col = i % cols
            row = i // cols
            cx = x0 + col * cell_w
            cy = y0 + row * cell_h
            is_sel = key in selected_keys
            # Card background
            if is_sel:
                fc(255, 255, 255); dc(*selected_color); pdf.set_line_width(0.6)
            else:
                fc(248, 250, 252); dc(*unselected_color); pdf.set_line_width(0.2)
            pdf.rect(cx + 0.3, cy + 0.3, cell_w - 0.6, cell_h - 0.6, "DF")
            # Icon (only if file exists and is renderable)
            icon_size = cell_h - 3
            icon_x = cx + 1.5
            icon_y = cy + 1.5
            if os.path.exists(icon_path):
                try:
                    pdf.image(icon_path,
                              x=icon_x, y=icon_y,
                              w=icon_size, h=icon_size,
                              keep_aspect_ratio=True)
                except Exception:
                    pass
            # Not-selected marker: a red diagonal strike-through over the
            # icon so it reads as "crossed out / not applicable".
            if not is_sel:
                pdf.set_draw_color(200, 40, 40)
                pdf.set_line_width(0.7)
                pdf.line(icon_x, icon_y,
                         icon_x + icon_size, icon_y + icon_size)
                pdf.set_line_width(0.2); dc(*BORDER_GRAY)
            # Label
            label_x = cx + cell_h - 1
            label_w = cell_w - (cell_h - 1) - 1
            pdf.set_font("Helvetica", "B" if is_sel else "", 7)
            tc(*(TEXT_DARK if is_sel else TEXT_MUTED))
            pdf.set_xy(label_x, cy + 2.5)
            pdf.cell(label_w, 4, _s(label), **NL)
            if is_sel:
                pdf.set_font("Helvetica", "B", 6); tc(*selected_color)
                pdf.set_xy(label_x, cy + 7.5)
                pdf.cell(label_w, 3, _s("SELECTED"), **NL)
            tc(*TEXT_DARK)
            pdf.set_line_width(0.2); dc(*BORDER_GRAY)
        # Leave the cursor below the grid
        total_rows = (len(catalog) + cols - 1) // cols
        pdf.set_xy(pdf.l_margin, y0 + total_rows * cell_h + 1)

    # ── Hazards: full grid + mitigations for selected ───────────────────
    section_title("Hazards")
    selected_hazards = {h.hazard_key: (h.preventive_measure or "").strip()
                        for h in (r.hazards or [])}
    _icon_grid(_PDF_HAZARD_CATALOG, set(selected_hazards.keys()),
               cols=4, cell_h=14,
               selected_color=(180, 30, 30))
    # Preventive measures list for the hazards the user picked
    if selected_hazards:
        pdf.set_font("Helvetica", "B", 8); tc(*IPS_BLUE)
        pdf.cell(W, 4.5, _s("Preventive measures for selected hazards"), **NL)
        pdf.set_font("Helvetica", "", 8); tc(*TEXT_DARK)
        for key, measure in selected_hazards.items():
            label = next((l for (k, l, _i) in _PDF_HAZARD_CATALOG if k == key), key)
            pdf.multi_cell(W, 3.8, _s(f"- {label}: {measure or '-'}"), **NL)
    if (r.hazards_other or "").strip():
        pdf.set_font("Helvetica", "B", 8); tc(*IPS_BLUE)
        pdf.cell(W, 4.5, _s("Other hazards & measures"), **NL)
        pdf.set_font("Helvetica", "", 8); tc(*TEXT_DARK)
        pdf.multi_cell(W, 3.8, _s(r.hazards_other), **NL)
    pdf.ln(0.5)

    # ── PPE: full grid + other ──────────────────────────────────────────
    section_title("Required personal protective equipment")
    selected_ppe = {p.ppe_key for p in (r.ppes or [])}
    _icon_grid(_PDF_PPE_CATALOG, selected_ppe,
               cols=5, cell_h=13,
               selected_color=(0, 120, 180))
    if (r.ppe_other or "").strip():
        pdf.set_font("Helvetica", "B", 8); tc(*IPS_BLUE)
        pdf.cell(W, 4.5, _s("Other PPE"), **NL)
        pdf.set_font("Helvetica", "", 8); tc(*TEXT_DARK)
        pdf.multi_cell(W, 3.8, _s(r.ppe_other), **NL)
    pdf.ln(0.5)

    # ── LOTO table ───────────────────────────────────────────────────────
    section_title("Lock-Out / Tag-Out (LOTO)")
    lotos = sorted((r.lotos or []), key=lambda x: (x.project_seq_id or x.id))
    if lotos:
        col_id, col_tag, col_status, col_desc = 22, 32, 26, W - 22 - 32 - 26
        fc(*LIGHT_GRAY); dc(*BORDER_GRAY)
        pdf.set_font("Helvetica", "B", 8); tc(*IPS_BLUE)
        pdf.cell(col_id,     5, _s("LOTO"),        border=1, fill=True, **ST)
        pdf.cell(col_tag,    5, _s("Tag #"),       border=1, fill=True, **ST)
        pdf.cell(col_status, 5, _s("Status"),      border=1, fill=True, **ST)
        pdf.cell(col_desc,   5, _s("Description"), border=1, fill=True, **NL)
        pdf.set_font("Helvetica", "", 8); tc(*TEXT_DARK)
        for l in lotos:
            lid = f"LT-{(l.project_seq_id or l.id):06d}"
            pdf.cell(col_id,     5, _s(lid), border=1, **ST)
            pdf.cell(col_tag,    5, _s((l.tag_number or "")[:18]), border=1, **ST)
            tc(*_PDF_LOTO_STATUS_COLORS.get(l.status, (100, 100, 100)))
            pdf.set_font("Helvetica", "B", 7)
            pdf.cell(col_status, 5, _s(_pdf_loto_status_label(l.status)),
                     border=1, align="C", **ST)
            pdf.set_font("Helvetica", "", 8); tc(*TEXT_DARK)
            pdf.cell(col_desc, 5, _s((l.description or "")[:90]), border=1, **NL)
    else:
        tc(*TEXT_MUTED); pdf.set_font("Helvetica", "I", 8)
        pdf.cell(W, 4.3, _s("No LOTOs attached (NA)."), **NL)
        tc(*TEXT_DARK)
    pdf.ln(1)

    # ── Area approvers ───────────────────────────────────────────────────
    section_title("Area approvals")
    approvals = sorted((r.area_approvals or []), key=lambda x: x.area_id)
    if approvals:
        col_a, col_desc, col_name, col_date = 18, W - 18 - 55 - 32, 55, 32
        fc(*LIGHT_GRAY)
        pdf.set_font("Helvetica", "B", 8); tc(*IPS_BLUE)
        pdf.cell(col_a,    5, _s("Area"),        border=1, fill=True, **ST)
        pdf.cell(col_desc, 5, _s("Description"), border=1, fill=True, **ST)
        pdf.cell(col_name, 5, _s("Approved by"), border=1, fill=True, **ST)
        pdf.cell(col_date, 5, _s("On"),          border=1, fill=True, **NL)
        pdf.set_font("Helvetica", "", 8); tc(*TEXT_DARK)
        for ap in approvals:
            tag = ap.area.tag if ap.area else "-"
            desc = (ap.area.description if ap.area else "") or ""
            name = ap.reviewed_by.name if ap.reviewed_by else "-"
            on = ap.reviewed_at.strftime("%Y-%m-%d %H:%M") if ap.reviewed_at else "-"
            pdf.cell(col_a,    5, _s(tag[:8]),   border=1, **ST)
            pdf.cell(col_desc, 5, _s(desc[:60]), border=1, **ST)
            pdf.cell(col_name, 5, _s(name[:36]), border=1, **ST)
            pdf.cell(col_date, 5, _s(on),        border=1, **NL)
    else:
        tc(*TEXT_MUTED); pdf.set_font("Helvetica", "I", 8)
        pdf.cell(W, 4.3, _s("No area approvals recorded."), **NL)
        tc(*TEXT_DARK)

    # ── Footer strip ─────────────────────────────────────────────────────
    footer_y = pdf.h - 8
    pdf.set_xy(pdf.l_margin, footer_y)
    tc(*TEXT_MUTED); pdf.set_font("Helvetica", "I", 7)
    pdf.cell(W, 4,
             _s(f"{display_id} - {pcode} - Printed {printing_ts} - Single-page controlled export"),
             align="C", **NL)

    # ── Emit ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    # fpdf2.output() returns a bytearray
    data = pdf.output(dest="S")
    if isinstance(data, str):           # older fpdf versions
        data = data.encode("latin-1")
    buf.write(bytes(data)); buf.seek(0)
    safe_title = "".join(c if c.isalnum() or c in "-_" else "_" for c in (r.title or ""))[:30]
    filename = f"{display_id}_{safe_title or 'permit'}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ═════════════════════════ LOTO ═════════════════════════════════════════════

@router.get("/lotos")
def list_lotos(package_id: Optional[int] = None,
               work_permit_id: Optional[int] = None,
               status: Optional[str] = None,
               db: Session = Depends(get_db),
               user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    q = (db.query(models.LOTO)
           .join(models.WorkPermit, models.WorkPermit.id == models.LOTO.work_permit_id)
           .filter(models.LOTO.project_id == user.project_id))
    if work_permit_id:
        q = q.filter(models.LOTO.work_permit_id == work_permit_id)
    if package_id:
        q = q.filter(models.WorkPermit.package_id == package_id)
    if status:
        q = q.filter(models.LOTO.status == status)
    if user.role == "VENDOR" and not _is_owner_or_admin(user, db):
        pkg_ids = _vendor_package_ids(user, db)
        if not pkg_ids:
            return []
        q = q.filter(models.WorkPermit.package_id.in_(pkg_ids))
    rows = q.order_by(models.LOTO.submitted_at.desc().nullslast(),
                      models.LOTO.id.desc()).all()
    return [_fmt_loto(l) for l in rows]


@router.get("/lotos/pending-approval")
def lotos_pending_approval(for_action_points: bool = False,
                           db: Session = Depends(get_db),
                           user: auth.ProjectContext = Depends(auth.get_project_user)):
    """REQUEST LOTO queue.
    • Visible to ADMIN, PROJECT_OWNER and to any user registered as a site
      supervisor on at least one area of the project.
    • `?for_action_points=true` (used by My Action Points) restricts to actual
      site supervisors only — admins/owners who are not also supervisors
      don't get the notification even though they can still use the tab.

    LOTO review is project-scoped, not area-scoped: any site supervisor can
    confirm / refuse any LOTO. This is deliberately looser than the
    work-permit approval flow, which is strictly per area."""
    if user.role == "BIDDER":
        return []
    is_supervisor = _is_site_supervisor(db, user.project_id, user.contact_id)
    if for_action_points:
        if not is_supervisor:
            return []
    else:
        if not (is_supervisor or _is_owner_or_admin(user, db)):
            return []
    rows = (db.query(models.LOTO)
              .filter(models.LOTO.project_id == user.project_id,
                      models.LOTO.status == "REQUEST")
              .order_by(models.LOTO.submitted_at.desc().nullslast(),
                        models.LOTO.id.desc())
              .all())
    return [_fmt_loto(l) for l in rows]


@router.get("/lotos/my-refused")
def lotos_my_refused(db: Session = Depends(get_db),
                     user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Refused LOTOs whose permit's package the current user is linked to via
    PackageContact — shown as action points so they can resubmit or cancel."""
    if user.role == "BIDDER" or not user.contact_id:
        return []
    pkg_ids = _vendor_package_ids(user, db)
    if not pkg_ids:
        return []
    rows = (db.query(models.LOTO)
              .join(models.WorkPermit, models.WorkPermit.id == models.LOTO.work_permit_id)
              .filter(models.LOTO.project_id == user.project_id,
                      models.LOTO.status == "REFUSED",
                      models.WorkPermit.package_id.in_(pkg_ids))
              .order_by(models.LOTO.reviewed_at.desc().nullslast(),
                        models.LOTO.id.desc())
              .all())
    return [_fmt_loto(l) for l in rows]


@router.get("/lotos/pending-release")
def lotos_pending_release(for_action_points: bool = False,
                          db: Session = Depends(get_db),
                          user: auth.ProjectContext = Depends(auth.get_project_user)):
    """LOTOs in TO_BE_RELEASED state — cascaded from a permit close, waiting
    for any site supervisor to confirm physical release on site. Project-
    scoped like LOTO review: any declared site supervisor can release any
    LOTO regardless of the permit's areas."""
    if user.role == "BIDDER":
        return []
    is_supervisor = _is_site_supervisor(db, user.project_id, user.contact_id)
    if for_action_points:
        if not is_supervisor:
            return []
    else:
        if not (is_supervisor or _is_owner_or_admin(user, db)):
            return []
    rows = (db.query(models.LOTO)
              .filter(models.LOTO.project_id == user.project_id,
                      models.LOTO.status == "TO_BE_RELEASED")
              .order_by(models.LOTO.submitted_at.desc().nullslast(),
                        models.LOTO.id.desc())
              .all())
    return [_fmt_loto(l) for l in rows]


@router.get("/lotos/{loto_id}")
def get_loto(loto_id: int,
             db: Session = Depends(get_db),
             user: auth.ProjectContext = Depends(auth.get_project_user)):
    if user.role == "BIDDER":
        raise HTTPException(403, "Access denied")
    l = db.query(models.LOTO).filter_by(id=loto_id, project_id=user.project_id).first()
    if not l:
        raise HTTPException(404, "LOTO not found")
    if user.role == "VENDOR" and not _is_owner_or_admin(user, db):
        if l.work_permit and l.work_permit.package_id not in _vendor_package_ids(user, db):
            raise HTTPException(403, "Not authorised for this package")
    return _fmt_loto(l)


@router.get("/lotos/{loto_id}/history")
def loto_history(loto_id: int,
                 db: Session = Depends(get_db),
                 user: auth.ProjectContext = Depends(auth.get_project_user)):
    l = db.query(models.LOTO).filter_by(id=loto_id, project_id=user.project_id).first()
    if not l:
        raise HTTPException(404, "LOTO not found")
    rows = (db.query(models.LOTOReview)
              .filter_by(loto_id=loto_id)
              .order_by(models.LOTOReview.created_at.asc())
              .all())
    return [{
        "id": r.id, "event": r.event, "confirmed": r.confirmed,
        "comment": r.comment, "actor_id": r.actor_id,
        "actor_name": r.actor.name if r.actor else None,
        "created_at": r.created_at.isoformat() + 'Z' if r.created_at else None,
    } for r in rows]


@router.post("/lotos/{loto_id}/confirm")
def confirm_loto(loto_id: int, body: _ReviewBody,
                 db: Session = Depends(get_db),
                 user: auth.ProjectContext = Depends(auth.get_project_user)):
    l = db.query(models.LOTO).filter_by(id=loto_id, project_id=user.project_id).first()
    if not l:
        raise HTTPException(404, "LOTO not found")
    if l.status != "REQUEST":
        raise HTTPException(400, "LOTO is not awaiting confirmation")
    if not _can_review_loto(user, l, db):
        raise HTTPException(403, "Only a site supervisor on one of the permit's areas can confirm this LOTO")
    now = datetime.utcnow()
    l.status = "LOCKED"
    l.locked_state = True
    l.reviewed_at = now
    l.reviewed_by_id = user.id
    l.refusal_comment = None
    _log_loto_review(db, l, "CONFIRM", user, confirmed=True,
                     comment=(body.comment or "").strip() or None)
    db.commit(); db.refresh(l)
    return _fmt_loto(l)


@router.post("/lotos/{loto_id}/refuse")
def refuse_loto(loto_id: int, body: _ReviewBody,
                db: Session = Depends(get_db),
                user: auth.ProjectContext = Depends(auth.get_project_user)):
    l = db.query(models.LOTO).filter_by(id=loto_id, project_id=user.project_id).first()
    if not l:
        raise HTTPException(404, "LOTO not found")
    if l.status != "REQUEST":
        raise HTTPException(400, "LOTO is not awaiting confirmation")
    if not _can_review_loto(user, l, db):
        raise HTTPException(403, "Only a site supervisor on one of the permit's areas can refuse this LOTO")
    comment = (body.comment or "").strip()
    if not comment:
        raise HTTPException(400, "A refusal comment is required")
    now = datetime.utcnow()
    l.status = "REFUSED"
    l.locked_state = False
    l.reviewed_at = now
    l.reviewed_by_id = user.id
    l.refusal_comment = comment
    _log_loto_review(db, l, "REFUSE", user, confirmed=False, comment=comment)
    db.commit(); db.refresh(l)
    return _fmt_loto(l)


@router.post("/lotos/{loto_id}/override")
def override_loto(loto_id: int, body: _OverrideBody,
                  db: Session = Depends(get_db),
                  user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Admins / project owners can override a REQUEST LOTO — confirm or refuse
    directly without being a site supervisor on the permit's areas."""
    if not auth.has_owner_or_lead_access(user, "Construction", db):
        raise HTTPException(403, "Only Admins and Project Owners can override")
    l = db.query(models.LOTO).filter_by(id=loto_id, project_id=user.project_id).first()
    if not l:
        raise HTTPException(404, "LOTO not found")
    if l.status != "REQUEST":
        raise HTTPException(400, "Can only override LOTOs awaiting confirmation")
    now = datetime.utcnow()
    comment = (body.comment or "").strip() or f"Decision overridden by {user.name}"
    if body.approved:
        l.status = "LOCKED"
        l.locked_state = True
        l.refusal_comment = None
    else:
        l.status = "REFUSED"
        l.locked_state = False
        l.refusal_comment = comment
    l.reviewed_at = now
    l.reviewed_by_id = user.id
    _log_loto_review(db, l, "OVERRIDE", user,
                     confirmed=bool(body.approved), comment=comment)
    db.commit(); db.refresh(l)
    return _fmt_loto(l)


@router.post("/lotos/{loto_id}/resubmit")
def resubmit_loto(loto_id: int, body: _ReviewBody,
                  db: Session = Depends(get_db),
                  user: auth.ProjectContext = Depends(auth.get_project_user)):
    l = db.query(models.LOTO).filter_by(id=loto_id, project_id=user.project_id).first()
    if not l:
        raise HTTPException(404, "LOTO not found")
    if l.status not in ("REFUSED", "CANCELLED"):
        raise HTTPException(400, "Only a refused or cancelled LOTO can be resubmitted")
    if not _can_manage_loto(user, l, db):
        raise HTTPException(403, "Not authorised for this permit")
    l.status = "REQUEST"
    l.locked_state = False
    l.submitted_at = datetime.utcnow()
    l.reviewed_at = None
    l.reviewed_by_id = None
    _log_loto_review(db, l, "RESUBMIT", user,
                     comment=(body.comment or "").strip() or None)
    db.commit(); db.refresh(l)
    return _fmt_loto(l)


@router.post("/lotos/{loto_id}/cancel")
def cancel_loto(loto_id: int, body: _ReviewBody,
                db: Session = Depends(get_db),
                user: auth.ProjectContext = Depends(auth.get_project_user)):
    l = db.query(models.LOTO).filter_by(id=loto_id, project_id=user.project_id).first()
    if not l:
        raise HTTPException(404, "LOTO not found")
    if l.status == "CANCELLED":
        raise HTTPException(400, "LOTO already cancelled")
    if l.status == "LOCKED":
        raise HTTPException(400, "A LOCKED LOTO cannot be cancelled")
    if not _can_manage_loto(user, l, db):
        raise HTTPException(403, "Not authorised for this permit")
    l.status = "CANCELLED"
    l.locked_state = False
    _log_loto_review(db, l, "CANCEL", user,
                     comment=(body.comment or "").strip() or None)
    db.commit(); db.refresh(l)
    return _fmt_loto(l)


@router.post("/lotos/{loto_id}/release")
def release_loto(loto_id: int, body: _ReviewBody,
                 db: Session = Depends(get_db),
                 user: auth.ProjectContext = Depends(auth.get_project_user)):
    """Confirm physical release of a LOTO that was cascaded to
    TO_BE_RELEASED when its parent permit was closed. Any declared site
    supervisor on the project may release any LOTO (same rule as LOTO
    confirm/refuse)."""
    l = db.query(models.LOTO).filter_by(id=loto_id, project_id=user.project_id).first()
    if not l:
        raise HTTPException(404, "LOTO not found")
    if l.status != "TO_BE_RELEASED":
        raise HTTPException(400, "Only a LOTO awaiting release can be released")
    if not _can_review_loto(user, l, db):
        raise HTTPException(403, "Only a site supervisor can confirm release")
    l.status = "RELEASED"
    l.locked_state = False
    l.reviewed_at = datetime.utcnow()
    l.reviewed_by_id = user.id
    _log_loto_review(db, l, "RELEASE", user,
                     comment=(body.comment or "").strip() or None)
    db.commit(); db.refresh(l)
    return _fmt_loto(l)


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT (audit-style, with created_by / created_at / status / reviewer)
#
# Read-only snapshots of each construction list. Frontend is the green
# "Export Excel" button — mirrors the Risk Register download pattern.
# ═════════════════════════════════════════════════════════════════════════════

try:
    from openpyxl import Workbook as _XlsxWorkbook
    from openpyxl.styles import Font as _XlsxFont, PatternFill as _XlsxPatternFill
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False


def _xlsx_check():
    if not _XLSX_OK:
        raise HTTPException(500, "openpyxl is not installed. Run: pip install openpyxl")


def _xlsx_header(ws):
    fill = _XlsxPatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        if cell.value is not None:
            cell.font = _XlsxFont(bold=True, color="FFFFFF")
            cell.fill = fill


def _xlsx_autowidth(ws, max_width: int = 50):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_fmt_ts(dt) -> str:
    if not dt:
        return ""
    try:
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(dt)


def _xlsx_user_name(u) -> str:
    if not u:
        return ""
    return getattr(u, "name", None) or getattr(u, "email", "") or ""


# ── Daily Reports ────────────────────────────────────────────────────────────
@router.get("/daily-reports/export/excel")
def export_daily_reports_xlsx(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _xlsx_check()
    rows = (
        db.query(models.DailyReport)
          .filter(models.DailyReport.project_id == user.project_id)
          .order_by(models.DailyReport.report_date.desc(), models.DailyReport.id.desc())
          .all()
    )
    wb = _XlsxWorkbook()
    ws = wb.active
    ws.title = "Daily Reports"
    ws.append([
        "Package", "Package Name", "Report Date", "No Work",
        "Avg Hours/Worker", "Workers (count)", "Total Hours",
        "Workers", "Areas", "Description",
        "Locked", "Locked At", "Unlocked At", "Unlocked By", "Unlock Comment",
        "Created By", "Created At", "Updated By", "Updated At",
    ])
    _xlsx_header(ws)
    for r in rows:
        workers = list(r.workers) if r.workers else []
        worker_names = ", ".join(drw.worker.name for drw in workers if drw.worker)
        worker_count = len(workers)
        total_hours = 0 if r.no_work else round((r.avg_hours_per_worker or 0) * worker_count, 2)
        areas = ", ".join(dra.area.tag for dra in (r.areas or []) if dra.area)
        ws.append([
            r.package.tag_number if r.package else "",
            r.package.name if r.package else "",
            r.report_date or "",
            "Y" if r.no_work else "N",
            r.avg_hours_per_worker or 0,
            worker_count,
            total_hours,
            worker_names,
            areas,
            r.description or "",
            "Y" if r.locked else "N",
            _xlsx_fmt_ts(r.locked_at),
            _xlsx_fmt_ts(r.unlocked_at),
            _xlsx_user_name(r.unlocked_by),
            r.unlock_comment or "",
            _xlsx_user_name(r.created_by),
            _xlsx_fmt_ts(r.created_at),
            _xlsx_user_name(r.updated_by),
            _xlsx_fmt_ts(r.updated_at),
        ])
    _xlsx_autowidth(ws)
    return _xlsx_response(wb, "daily_reports.xlsx")


# ── Work Logs ────────────────────────────────────────────────────────────────
@router.get("/work-logs/export/excel")
def export_work_logs_xlsx(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _xlsx_check()
    rows = (
        db.query(models.WorkLog)
          .filter(models.WorkLog.project_id == user.project_id)
          .order_by(models.WorkLog.start_date.desc(), models.WorkLog.id.desc())
          .all()
    )
    wb = _XlsxWorkbook()
    ws = wb.active
    ws.title = "Work Logs"
    ws.append([
        "Package", "Package Name", "Start Date", "End Date", "Status", "Notes",
        "Created By", "Created At", "Updated By", "Updated At",
    ])
    _xlsx_header(ws)
    for l in rows:
        ws.append([
            l.package.tag_number if l.package else "",
            l.package.name if l.package else "",
            l.start_date or "",
            l.end_date or "",
            "ONGOING" if not l.end_date else "ENDED",
            l.notes or "",
            _xlsx_user_name(l.created_by),
            _xlsx_fmt_ts(l.created_at),
            _xlsx_user_name(l.updated_by),
            _xlsx_fmt_ts(l.updated_at),
        ])
    _xlsx_autowidth(ws)
    return _xlsx_response(wb, "work_logs.xlsx")


# ── Workers & Subcontractors (two sheets in one workbook) ────────────────────
@router.get("/workers-subs/export/excel")
def export_workers_subs_xlsx(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _xlsx_check()
    workers = (
        db.query(models.Worker)
          .filter(models.Worker.project_id == user.project_id)
          .order_by(models.Worker.id.asc())
          .all()
    )
    subs = (
        db.query(models.Subcontractor)
          .filter(models.Subcontractor.project_id == user.project_id)
          .order_by(models.Subcontractor.id.asc())
          .all()
    )

    wb = _XlsxWorkbook()

    # Workers sheet
    ws = wb.active
    ws.title = "Workers"
    ws.append([
        "ID", "Package", "Package Name", "Name", "Phone",
        "Employed By Subcontractor", "Subcontractor Company",
        "Status", "Submitted At", "Reviewed At", "Reviewed By",
        "Rejection Comment",
        "Created By", "Created At", "Updated By", "Updated At",
    ])
    _xlsx_header(ws)
    for w in workers:
        ws.append([
            f"WK-{(w.project_seq_id or w.id):06d}",
            w.package.tag_number if w.package else "",
            w.package.name if w.package else "",
            w.name,
            w.phone or "",
            "Y" if w.is_subcontractor else "N",
            w.subcontractor.company if w.subcontractor else "",
            w.status or "",
            _xlsx_fmt_ts(w.submitted_at),
            _xlsx_fmt_ts(w.reviewed_at),
            _xlsx_user_name(w.reviewed_by),
            w.rejection_comment or "",
            _xlsx_user_name(w.created_by),
            _xlsx_fmt_ts(w.created_at),
            _xlsx_user_name(w.updated_by),
            _xlsx_fmt_ts(w.updated_at),
        ])
    _xlsx_autowidth(ws)

    # Subcontractors sheet
    sws = wb.create_sheet("Subcontractors")
    sws.append([
        "ID", "Package", "Package Name", "Company", "Contact Person",
        "Phone", "Email", "Scope / Description",
        "Created By", "Created At", "Updated By", "Updated At",
    ])
    _xlsx_header(sws)
    for s in subs:
        sws.append([
            f"SU-{(s.project_seq_id or s.id):06d}",
            s.package.tag_number if s.package else "",
            s.package.name if s.package else "",
            s.company,
            s.contact_person or "",
            s.phone or "",
            s.email or "",
            s.description or "",
            _xlsx_user_name(s.created_by),
            _xlsx_fmt_ts(s.created_at),
            _xlsx_user_name(s.updated_by),
            _xlsx_fmt_ts(s.updated_at),
        ])
    _xlsx_autowidth(sws)
    return _xlsx_response(wb, "workers_subcontractors.xlsx")


# ── Work Permits ─────────────────────────────────────────────────────────────
@router.get("/work-permits/export/excel")
def export_work_permits_xlsx(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _xlsx_check()
    permits = (
        db.query(models.WorkPermit)
          .filter(models.WorkPermit.project_id == user.project_id)
          .order_by(models.WorkPermit.id.desc())
          .all()
    )
    wb = _XlsxWorkbook()
    ws = wb.active
    ws.title = "Work Permits"
    ws.append([
        "ID", "Package", "Title", "Description",
        "Start Date", "End Date",
        "Permit Types", "Areas", "Hazards", "PPE",
        "Status", "Approved Areas", "Rejected Areas",
        "Area Approvals (detail)",
        "Submitted By", "Submitted At",
        "Created By", "Created At", "Updated By", "Updated At",
    ])
    _xlsx_header(ws)
    for p in permits:
        permit_types = ", ".join(
            (pt.permit_type.name for pt in (p.permit_types or []) if pt.permit_type)
        )
        areas = ", ".join(wpa.area.tag for wpa in (p.areas or []) if wpa.area)
        hazards = ", ".join((h.hazard_key or "") for h in (p.hazards or []))
        ppe = ", ".join((pp.ppe_key or "") for pp in (p.ppes or []))
        approvals = list(p.area_approvals or [])
        approved_areas = ", ".join(a.area.tag for a in approvals if a.status == "APPROVED" and a.area)
        rejected_areas = ", ".join(a.area.tag for a in approvals if a.status == "REJECTED" and a.area)
        detail_parts = []
        for a in approvals:
            tag = a.area.tag if a.area else "?"
            part = f"{tag}: {a.status}"
            if a.status in ("APPROVED", "REJECTED"):
                part += f" by {_xlsx_user_name(a.reviewed_by)} on {_xlsx_fmt_ts(a.reviewed_at)}"
            if a.rejection_comment:
                part += f" — {a.rejection_comment}"
            detail_parts.append(part)
        detail = " | ".join(detail_parts)
        ws.append([
            f"WP-{(p.project_seq_id or p.id):06d}",
            p.package.tag_number if p.package else "",
            p.title or "",
            p.description or "",
            p.start_date or "",
            p.end_date or "",
            permit_types,
            areas,
            hazards,
            ppe,
            p.status or "",
            approved_areas,
            rejected_areas,
            detail,
            _xlsx_user_name(p.submitted_by),
            _xlsx_fmt_ts(p.submitted_at),
            _xlsx_user_name(p.created_by),
            _xlsx_fmt_ts(p.created_at),
            _xlsx_user_name(p.updated_by),
            _xlsx_fmt_ts(p.updated_at),
        ])
    _xlsx_autowidth(ws)
    return _xlsx_response(wb, "work_permits.xlsx")


# ── LOTO ─────────────────────────────────────────────────────────────────────
@router.get("/lotos/export/excel")
def export_lotos_xlsx(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    _xlsx_check()
    lotos = (
        db.query(models.LOTO)
          .filter(models.LOTO.project_id == user.project_id)
          .order_by(models.LOTO.id.desc())
          .all()
    )
    wb = _XlsxWorkbook()
    ws = wb.active
    ws.title = "LOTO"
    ws.append([
        "ID", "Work Permit", "Package", "Tag Number", "Description",
        "Status", "Locked State",
        "Submitted At", "Reviewed At", "Reviewed By", "Refusal Comment",
        "Created By", "Created At", "Updated By", "Updated At",
    ])
    _xlsx_header(ws)
    for l in lotos:
        permit = l.work_permit
        permit_disp = f"WP-{(permit.project_seq_id or permit.id):06d}" if permit else ""
        pkg_tag = permit.package.tag_number if permit and permit.package else ""
        ws.append([
            f"LO-{(l.project_seq_id or l.id):06d}",
            permit_disp,
            pkg_tag,
            l.tag_number or "",
            l.description or "",
            l.status or "",
            "Y" if l.locked_state else "N",
            _xlsx_fmt_ts(l.submitted_at),
            _xlsx_fmt_ts(l.reviewed_at),
            _xlsx_user_name(l.reviewed_by),
            l.refusal_comment or "",
            _xlsx_user_name(l.created_by),
            _xlsx_fmt_ts(l.created_at),
            _xlsx_user_name(l.updated_by),
            _xlsx_fmt_ts(l.updated_at),
        ])
    _xlsx_autowidth(ws)
    return _xlsx_response(wb, "lotos.xlsx")
