"""
Safety Observations PDF export.

Generates a multi-page A4 report with:
  • Cover page  — big Impulse logo, project info, applied filters, and a
                  tabular index of all included observations.
  • Item pages  — A4 portrait, one observation per page. Data top-left,
                  floorplan-with-pin bottom-left, photos stacked on the right.
  • Floorplan summary pages — landscape, one per applicable floorplan with
                              every applicable pin and its sequence number.

Reports are generated in a background thread and stored under
`uploads/{PROJECT}/Safety Reports/`. The endpoint enqueues a Report row and
returns immediately; the frontend polls for status and downloads when ready.

Every page (except cover) shows a small Impulse logo in the header band.
"""
import json
import re
import threading
import traceback
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db, SessionLocal
import models
import auth


export_router = APIRouter(prefix="/api/safety", tags=["safety-export"])


# ── Constants ────────────────────────────────────────────────────────────────

LOGO_PATH = Path("Logos/impulse_logo_for light background.png")
UPLOAD_ROOT = Path("uploads")

IPS_BLUE     = (27,  79,  140)
ACCENT_BLUE  = (0,   174, 239)
LIGHT_GRAY   = (241, 245, 249)
BORDER_GRAY  = (203, 213, 225)
TEXT_DARK    = (30,  41,  59)
TEXT_MUTED   = (100, 116, 139)
WHITE        = (255, 255, 255)
GREEN        = (16,  185, 129)
RED          = (220, 38,  38)
AMBER        = (245, 158, 11)
GRAY_400     = (156, 163, 175)
GRAY_100     = (243, 244, 246)


# Latin-1 sanitiser (Helvetica, the core fpdf font, is Latin-1 only).
_REPL = {
    "—": "-", "–": "-",
    "•": "*",
    "…": "...",
    "“": '"', "”": '"', "‘": "'", "’": "'",
    " ": " ",
}


def _s(text) -> str:
    if text is None:
        return ""
    s = str(text)
    for k, v in _REPL.items():
        if k in s:
            s = s.replace(k, v)
    return s.encode("latin-1", "replace").decode("latin-1")


def _status_color(status: str):
    return {
        "DRAFT":     GRAY_400,
        "SUBMITTED": ACCENT_BLUE,
        "RECEIVED":  AMBER,
        "CLOSED":    GREEN,
    }.get(status, GRAY_400)


def _polarity_color(polarity: str):
    return GREEN if (polarity or "").upper() == "POSITIVE" else RED


# ── Filters / grouping ───────────────────────────────────────────────────────

class ExportFilters(BaseModel):
    package_ids: List[int] = []
    area_ids:    List[int] = []
    statuses:    List[str] = []
    group_by:    str = "package_area"   # package_area | area_package | package | area | status | none
    per_package_plans: bool = False     # one summary page per (floorplan, package) instead of per floorplan


_GROUP_LABELS = {
    "package_area": "Package then Area",
    "area_package": "Area then Package",
    "package":      "Package",
    "area":         "Area",
    "status":       "Status",
    "none":         "None (chronological)",
}


def _grouping_keys(o: models.SafetyObservation, mode: str):
    """Return (level1_key, level1_label, level2_key, level2_label).
    level2_* may be None when single-level grouping is requested."""
    pkg = (o.package.tag_number if o.package else "—")
    pkg_full = f"{pkg} - {o.package.name}" if o.package else "(no package)"
    area = (o.area.tag if o.area else "—")
    area_full = f"{area} - {o.area.description}" if o.area else "(no area)"
    if mode == "package_area":
        return (pkg, pkg_full, area, area_full)
    if mode == "area_package":
        return (area, area_full, pkg, pkg_full)
    if mode == "package":
        return (pkg, pkg_full, None, None)
    if mode == "area":
        return (area, area_full, None, None)
    if mode == "status":
        return (o.status, o.status, None, None)
    return ("", "All observations", None, None)


def _group_observations(obs_list, mode: str):
    """Returns a list of (level1_label, [(level2_label_or_None, [obs, ...]), ...])."""
    by_l1 = {}
    for o in obs_list:
        k1, l1, k2, l2 = _grouping_keys(o, mode)
        by_l1.setdefault(k1, {"label": l1, "by_l2": {}})
        sub = by_l1[k1]["by_l2"].setdefault(k2 if k2 is not None else "_", {"label": l2, "items": []})
        sub["items"].append(o)
    out = []
    for k1 in sorted(by_l1.keys()):
        sub_items = []
        for k2 in sorted(by_l1[k1]["by_l2"].keys()):
            d = by_l1[k1]["by_l2"][k2]
            sub_items.append((d["label"], d["items"]))
        out.append((by_l1[k1]["label"], sub_items))
    return out


def _filter_summary(filters: ExportFilters, db: Session, project_id: int) -> dict:
    """Human-readable summary of the applied filters."""
    if filters.package_ids:
        rows = db.query(models.Package).filter(
            models.Package.project_id == project_id,
            models.Package.id.in_(filters.package_ids),
        ).all()
        pkg_text = ", ".join(r.tag_number for r in rows) or "(none)"
    else:
        pkg_text = "All"
    if filters.area_ids:
        rows = db.query(models.Area).filter(
            models.Area.project_id == project_id,
            models.Area.id.in_(filters.area_ids),
        ).all()
        area_text = ", ".join(r.tag for r in rows) or "(none)"
    else:
        area_text = "All"
    status_text = ", ".join(filters.statuses) if filters.statuses else "All"
    return {
        "Packages": pkg_text,
        "Areas":    area_text,
        "Statuses": status_text,
        "Grouped by": _GROUP_LABELS.get(filters.group_by, filters.group_by),
    }


# ── Image helpers ────────────────────────────────────────────────────────────

def _floorplan_with_pins(fp_row: models.Floorplan, pins: list, label_pins: bool = True) -> Optional[BytesIO]:
    """Open the floorplan image, draw a red marker (and optional number) at
    each (x_norm, y_norm), and return a PNG BytesIO. Returns None if the
    file can't be loaded."""
    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception:
        return None
    if not fp_row or not fp_row.stored_path:
        return None
    abs_path = (UPLOAD_ROOT / fp_row.stored_path).resolve()
    try:
        abs_path.relative_to(UPLOAD_ROOT.resolve())
    except ValueError:
        return None
    if not abs_path.exists():
        return None
    try:
        img = Image.open(abs_path).convert("RGB")
    except Exception:
        return None
    draw = ImageDraw.Draw(img)
    w, h = img.size
    diag = (w * w + h * h) ** 0.5
    # Marker sizing — tuned between the over-large original and the too-tight
    # second pass. Halves the dot relative to the image but keeps numbers
    # readable on a landscape A4 print.
    r = max(6, int(diag / 350))
    ring = max(2, r // 3)
    font_size = max(15, int(diag / 145))
    font = None
    try:
        font = ImageFont.truetype("arialbd.ttf", font_size)
    except Exception:
        try:
            font = ImageFont.truetype("arial.ttf", font_size)
        except Exception:
            try:
                font = ImageFont.load_default()
            except Exception:
                font = None
    for pin in pins:
        x_norm = pin.get("x")
        y_norm = pin.get("y")
        if x_norm is None or y_norm is None:
            continue
        cx = int(x_norm * w)
        # Saved coordinate represents the *tip* of the web-UI teardrop pin
        # (where the user tapped). Anchor the dot's bottom at that point so
        # the visual "pointer" matches between web view and printed PDF.
        cy_tip = int(y_norm * h)
        center_y = cy_tip - r
        # White ring
        draw.ellipse([cx - r - ring, center_y - r - ring, cx + r + ring, center_y + r + ring], fill="white")
        # Red dot
        color = pin.get("color") or "#dc2626"
        draw.ellipse([cx - r, center_y - r, cx + r, center_y + r], fill=color)
        # Number label
        if label_pins and pin.get("label") and font is not None:
            label = str(pin["label"])
            try:
                # Position label to the upper-right of the marker. Stroke
                # adds a subtle white halo for legibility on busy plans.
                draw.text(
                    (cx + r + 3, center_y - r - int(font_size * 0.85)),
                    label,
                    fill="black", font=font,
                    stroke_width=2, stroke_fill="white",
                )
            except Exception:
                pass
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def _open_attachment_buf(stored_path: str) -> Optional[BytesIO]:
    """Open an image attachment from disk, normalise EXIF rotation, and
    return a JPEG-encoded BytesIO. Returns None on any failure."""
    try:
        from PIL import Image, ImageOps
    except Exception:
        return None
    if not stored_path:
        return None
    abs_path = (UPLOAD_ROOT / stored_path).resolve()
    try:
        abs_path.relative_to(UPLOAD_ROOT.resolve())
    except ValueError:
        return None
    if not abs_path.exists():
        return None
    try:
        img = Image.open(abs_path)
        img = ImageOps.exif_transpose(img)
        if img.mode != "RGB":
            img = img.convert("RGB")
    except Exception:
        return None
    buf = BytesIO()
    try:
        img.save(buf, format="JPEG", quality=80)
    except Exception:
        return None
    buf.seek(0)
    return buf


# ── PDF building ─────────────────────────────────────────────────────────────

def _build_safety_pdf(obs_list, filters: ExportFilters, project, db: Session) -> bytes:
    from fpdf import FPDF, XPos, YPos

    NL = dict(new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False, margin=10)
    pdf.set_margins(10, 10, 10)

    def tc(*rgb): pdf.set_text_color(*rgb)
    def fc(*rgb): pdf.set_fill_color(*rgb)
    def dc(*rgb): pdf.set_draw_color(*rgb)

    def small_logo_header():
        """Tiny logo top-left + project info top-right on every non-cover page."""
        if LOGO_PATH.exists():
            try:
                pdf.image(str(LOGO_PATH), x=10, y=8, w=22, keep_aspect_ratio=True)
            except Exception:
                pass
        tc(*TEXT_MUTED)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_xy(pdf.w - 100, 10)
        proj_num = (project.project_number if project else "") or "-"
        pdf.cell(90, 4, _s(f"Safety Observations Report  •  {proj_num}"), align="R", **NL)
        pdf.set_xy(pdf.w - 100, 14)
        pdf.cell(90, 4, _s(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"), align="R", **NL)
        # Thin separator line
        dc(*BORDER_GRAY)
        pdf.set_line_width(0.2)
        pdf.line(10, 22, pdf.w - 10, 22)

    # ─── COVER PAGE ──────────────────────────────────────────────────────────
    pdf.add_page()
    if LOGO_PATH.exists():
        # Big centered logo
        pdf.image(str(LOGO_PATH), x=(pdf.w - 90) / 2, y=22, w=90, keep_aspect_ratio=True)

    pdf.set_y(80)
    pdf.set_font("Helvetica", "B", 24)
    tc(*IPS_BLUE)
    pdf.cell(0, 12, _s("Safety Observations Report"), align="C", **NL)

    pdf.ln(2)
    pdf.set_font("Helvetica", "", 12)
    tc(*TEXT_DARK)
    proj_num = (project.project_number if project else "") or "-"
    proj_desc = ((project.description if project else "") or "").strip() or "-"
    pdf.cell(0, 7, _s(f"{proj_num}  -  {proj_desc}"), align="C", **NL)

    pdf.ln(4)
    pdf.set_font("Helvetica", "", 9)
    tc(*TEXT_MUTED)
    pdf.cell(0, 5, _s(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"), align="C", **NL)

    # Filter summary card
    pdf.ln(8)
    box_x = 20
    box_w = pdf.w - 40
    fc(*LIGHT_GRAY)
    dc(*BORDER_GRAY)
    pdf.rect(box_x, pdf.get_y(), box_w, 32, "DF")
    pdf.set_xy(box_x + 4, pdf.get_y() + 3)
    pdf.set_font("Helvetica", "B", 10)
    tc(*IPS_BLUE)
    pdf.cell(0, 6, _s("Applied filters"), **NL)
    summary = _filter_summary(filters, db, project.id)
    pdf.set_font("Helvetica", "", 9)
    tc(*TEXT_DARK)
    for k, v in summary.items():
        pdf.set_x(box_x + 4)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(28, 5, _s(f"{k}:"))
        pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(box_w - 32, 5, _s(v))

    pdf.ln(4)

    # Index table
    pdf.set_font("Helvetica", "B", 11)
    tc(*IPS_BLUE)
    pdf.cell(0, 6, _s(f"Included observations  ({len(obs_list)})"), **NL)
    pdf.ln(1)

    def _index_table_header():
        pdf.set_font("Helvetica", "B", 9)
        fc(*IPS_BLUE)
        tc(*WHITE)
        pdf.cell(28, 6, _s("ID"), border=1, fill=True, align="C")
        pdf.cell(30, 6, _s("Package"), border=1, fill=True, align="C")
        pdf.cell(94, 6, _s("Category"), border=1, fill=True, align="C")
        pdf.cell(38, 6, _s("Status"), border=1, fill=True, align="C", **NL)

    _index_table_header()
    pdf.set_font("Helvetica", "", 9)
    tc(*TEXT_DARK)
    fill_alt = False
    for o in obs_list:
        if pdf.get_y() > 275:
            pdf.add_page()
            small_logo_header()
            pdf.set_y(28)
            _index_table_header()
            pdf.set_font("Helvetica", "", 9)
            tc(*TEXT_DARK)
            fill_alt = False
        seq = f"SO-{(o.project_seq_id or o.id):06d}"
        pkg = (o.package.tag_number if o.package else "-")
        cat = (o.category.name if o.category else "-")
        if len(cat) > 60:
            cat = cat[:57] + "..."
        status = o.status or "-"
        if fill_alt:
            fc(*GRAY_100)
        else:
            fc(*WHITE)
        pdf.cell(28, 5, _s(seq), border=1, fill=True)
        pdf.cell(30, 5, _s(pkg), border=1, fill=True)
        pdf.cell(94, 5, _s(cat), border=1, fill=True)
        # Coloured status pill
        sc = _status_color(status)
        fc(*sc)
        tc(*WHITE)
        pdf.cell(38, 5, _s(status), border=1, fill=True, align="C", **NL)
        tc(*TEXT_DARK)
        fill_alt = not fill_alt

    # ─── ITEM PAGES ──────────────────────────────────────────────────────────
    grouped = _group_observations(obs_list, filters.group_by)

    # Layout (portrait, one observation per page):
    # margins 10mm, header band 22mm at top
    HEADER_BOTTOM = 24    # below the small header line
    LEFT_W = 110
    GAP = 5
    RIGHT_W = pdf.w - 20 - LEFT_W - GAP   # ≈ 75mm

    def _draw_status_pill(x, y, w, h, label, color):
        fc(*color)
        pdf.rect(x, y, w, h, "F")
        tc(*WHITE)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_xy(x, y)
        pdf.cell(w, h, _s(label), align="C")
        tc(*TEXT_DARK)

    def _draw_observation(x0, y0, o):
        """Render one observation block starting at (x0, y0), using the rest
        of the page height down to the bottom margin."""
        block_h = pdf.h - 12 - y0     # remaining content height
        # Floorplan thumbnail height — generous now that each item gets a page.
        fp_h = 80
        # Data area takes everything above the floorplan (with a small gap).
        data_h = block_h - fp_h - 4
        # Title row
        pdf.set_xy(x0, y0)
        pdf.set_font("Helvetica", "B", 11)
        tc(*IPS_BLUE)
        seq = f"SO-{(o.project_seq_id or o.id):06d}"
        pdf.cell(40, 5, _s(seq))
        polarity = (o.category.polarity if o.category else "") or "NEGATIVE"
        pol_label = "Positive" if polarity == "POSITIVE" else "Negative"
        _draw_status_pill(x0 + 40, y0, 18, 5, pol_label, _polarity_color(polarity))
        sc = _status_color(o.status or "")
        _draw_status_pill(x0 + 60, y0, 22, 5, o.status or "-", sc)

        # Data block
        pdf.set_xy(x0, y0 + 7)
        pdf.set_font("Helvetica", "", 8)
        tc(*TEXT_MUTED)

        def _row(label, value, lbl_w=22, val_w=LEFT_W - 22):
            pdf.set_x(x0)
            pdf.set_font("Helvetica", "B", 8)
            tc(*TEXT_MUTED)
            pdf.cell(lbl_w, 4.5, _s(label))
            pdf.set_font("Helvetica", "", 8)
            tc(*TEXT_DARK)
            pdf.multi_cell(val_w, 4.5, _s(value or "-"))

        _row("Category:", o.category.name if o.category else "-")
        _row("Package:",  f"{o.package.tag_number} - {o.package.name}" if o.package else "-")
        _row("Area:",     f"{o.area.tag} - {o.area.description}" if o.area else "-")
        if o.subcontractor:
            _row("Subcontractor:", o.subcontractor.company)
        if o.worker:
            _row("Worker:", o.worker.name)
        if o.created_by:
            ts = o.created_at.strftime("%Y-%m-%d") if o.created_at else "-"
            _row("Reported:", f"{ts} by {o.created_by.name}")
        if o.acknowledged_at and o.acknowledged_by:
            ts = o.acknowledged_at.strftime("%Y-%m-%d")
            _row("Acked:", f"{ts} by {o.acknowledged_by.name}")
        if o.closed_at and o.closed_by:
            ts = o.closed_at.strftime("%Y-%m-%d")
            _row("Closed:", f"{ts} by {o.closed_by.name}")

        # Details / remediation paragraphs — generous wrap with no aggressive
        # truncation now that each item has a full page.
        pdf.ln(1)
        pdf.set_x(x0)
        pdf.set_font("Helvetica", "B", 8)
        tc(*TEXT_MUTED)
        pdf.cell(LEFT_W, 4.5, _s("Details:"), **NL)
        pdf.set_font("Helvetica", "", 8)
        tc(*TEXT_DARK)
        pdf.set_x(x0)
        pdf.multi_cell(LEFT_W, 4, _s((o.details or "-").strip()))
        if (o.remediation_request or "").strip():
            pdf.ln(0.5)
            pdf.set_x(x0)
            pdf.set_font("Helvetica", "B", 8)
            tc(*TEXT_MUTED)
            pdf.cell(LEFT_W, 4.5, _s("Remediation:"), **NL)
            pdf.set_font("Helvetica", "", 8)
            tc(*TEXT_DARK)
            pdf.set_x(x0)
            pdf.multi_cell(LEFT_W, 4, _s(o.remediation_request))
        # Workflow history with comments — covers ACKNOWLEDGED comments,
        # CLOSED comments, and REOPENED reasons. The acknowledge_comment
        # column on the row itself is cleared on reopen, but the history
        # log retains it, so this is the canonical view.
        history_with_comments = [
            h for h in (o.history or [])
            if (h.comment or "").strip()
        ]
        if history_with_comments:
            pdf.ln(0.5)
            pdf.set_x(x0)
            pdf.set_font("Helvetica", "B", 8)
            tc(*TEXT_MUTED)
            pdf.cell(LEFT_W, 4.5, _s("History:"), **NL)
            for h in history_with_comments:
                date_str = h.created_at.strftime("%Y-%m-%d") if h.created_at else "-"
                actor = (h.actor.name if h.actor else "-")
                event_label = (h.event or "").replace("_", " ").title()
                pdf.set_x(x0)
                pdf.set_font("Helvetica", "B", 8)
                tc(*IPS_BLUE)
                pdf.cell(LEFT_W, 4, _s(f"  {event_label}  -  {date_str}  -  {actor}"), **NL)
                pdf.set_x(x0)
                pdf.set_font("Helvetica", "", 8)
                tc(*TEXT_DARK)
                pdf.multi_cell(LEFT_W, 4, _s(f"     {h.comment}"))

        # Floorplan (bottom-left)
        fp_y = y0 + block_h - fp_h
        fp_w = LEFT_W
        dc(*BORDER_GRAY)
        pdf.set_line_width(0.2)
        pdf.rect(x0, fp_y, fp_w, fp_h)
        if o.floorplan_id and o.floorplan_x is not None and o.floorplan_y is not None and o.floorplan:
            buf = _floorplan_with_pins(o.floorplan, [{
                "x": o.floorplan_x,
                "y": o.floorplan_y,
                "label": str(o.project_seq_id or o.id),
                "color": "#" + "%02x%02x%02x" % _polarity_color(polarity),
            }], label_pins=False)  # label not needed on a single-pin thumb
            if buf is not None:
                try:
                    pdf.image(buf, x=x0 + 0.5, y=fp_y + 0.5, w=fp_w - 1, h=fp_h - 1, keep_aspect_ratio=True)
                except Exception:
                    pass
        else:
            pdf.set_xy(x0, fp_y + fp_h / 2 - 2)
            tc(*TEXT_MUTED)
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(fp_w, 4, _s("No floorplan pinned"), align="C")

        # Photos (right column, 2x2 grid of up to 4) — full block height.
        rx = x0 + LEFT_W + GAP
        ry = y0
        rh = block_h
        atts = (
            db.query(models.FileAttachment)
              .filter(models.FileAttachment.record_type == "safety_observation")
              .filter(models.FileAttachment.record_id == o.id)
              .filter(models.FileAttachment.content_type.like("image/%"))
              .order_by(models.FileAttachment.uploaded_at, models.FileAttachment.id)
              .limit(5)
              .all()
        )
        # All photos stack in a single column. Each cell uses the full
        # column width — no side-by-side grid — so portraits and landscapes
        # both have room to breathe.
        n = len(atts)
        cell_w = RIGHT_W
        if n <= 1:
            cell_h = rh
            positions = [(rx, ry)] if n == 1 else []
        else:
            cell_h = (rh - 2 * (n - 1)) / n
            positions = [(rx, ry + i * (cell_h + 2)) for i in range(n)]

        for i, att in enumerate(atts):
            if i >= len(positions):
                break
            cx, cy = positions[i]
            dc(*BORDER_GRAY)
            pdf.rect(cx, cy, cell_w, cell_h)
            buf = _open_attachment_buf(att.stored_path)
            if buf is not None:
                try:
                    pdf.image(buf, x=cx + 0.5, y=cy + 0.5,
                              w=cell_w - 1, h=cell_h - 1,
                              keep_aspect_ratio=True)
                except Exception:
                    pass
        # Empty-state placeholder when zero photos
        if not atts:
            pdf.set_xy(rx, ry + rh / 2 - 4)
            tc(*TEXT_MUTED)
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(RIGHT_W, 4, _s("No photos attached"), align="C")
            dc(*BORDER_GRAY)
            pdf.rect(rx, ry, RIGHT_W, rh)

        # Outer block frame
        dc(*BORDER_GRAY)
        pdf.set_line_width(0.3)
        pdf.rect(x0 - 1, y0 - 1, pdf.w - 18, block_h + 2)

    # One observation per A4 portrait page. Group separator bands are
    # rendered at the top of the page when the grouping changes.
    last_l1 = None
    last_l2 = None
    for l1_label, l2_groups in grouped:
        for l2_label, items in l2_groups:
            for o in items:
                pdf.add_page()
                small_logo_header()
                pdf.set_y(HEADER_BOTTOM + 2)
                if l1_label != last_l1:
                    _group_band(pdf, l1_label, level=1)
                    last_l1 = l1_label
                    last_l2 = None
                if l2_label is not None and l2_label != last_l2:
                    _group_band(pdf, l2_label, level=2)
                    last_l2 = l2_label
                _draw_observation(11, pdf.get_y(), o)

    # ─── FLOORPLAN SUMMARY PAGES ─────────────────────────────────────────────
    # One landscape page per floorplan, or — when per_package_plans is on —
    # one landscape page per (floorplan, package) so dense plans split into
    # several less-crowded pages.
    plan_groups = {}
    for o in obs_list:
        if (o.floorplan_id is None or o.floorplan_x is None
                or o.floorplan_y is None or not o.floorplan):
            continue
        if filters.per_package_plans:
            key = (o.floorplan_id, o.package_id)
        else:
            key = (o.floorplan_id, None)
        bucket = plan_groups.setdefault(key, {
            "fp": o.floorplan,
            "package": o.package if filters.per_package_plans else None,
            "pins": [],
        })
        polarity = (o.category.polarity if o.category else "") or "NEGATIVE"
        col = _polarity_color(polarity) if (o.status or "") != "CLOSED" else GRAY_400
        bucket["pins"].append({
            "x": o.floorplan_x,
            "y": o.floorplan_y,
            "label": str(o.project_seq_id or o.id),
            "color": "#%02x%02x%02x" % col,
        })

    # Stable ordering: by floorplan id, then by package tag (for the per-package mode)
    def _plan_key_sort(k):
        fp_id, pkg_id = k
        pkg_tag = ""
        if pkg_id is not None and plan_groups[k]["package"]:
            pkg_tag = plan_groups[k]["package"].tag_number or ""
        return (fp_id, pkg_tag)

    for key in sorted(plan_groups.keys(), key=_plan_key_sort):
        blob = plan_groups[key]
        fp = blob["fp"]
        pkg = blob["package"]
        pins = blob["pins"]
        if not pins:
            continue
        pdf.add_page(orientation="L")
        small_logo_header()
        # Title
        pdf.set_xy(10, HEADER_BOTTOM + 2)
        pdf.set_font("Helvetica", "B", 14)
        tc(*IPS_BLUE)
        if pkg:
            title = f"Floorplan: {fp.name}  -  Package: {pkg.tag_number} - {pkg.name}"
        else:
            title = f"Floorplan: {fp.name}"
        pdf.cell(0, 6, _s(title), **NL)
        pdf.set_font("Helvetica", "", 9)
        tc(*TEXT_MUTED)
        pdf.set_x(10)
        pdf.cell(0, 5, _s(f"{len(pins)} pinned observation(s)"), **NL)
        # Image
        avail_w = pdf.w - 20
        avail_h = pdf.h - HEADER_BOTTOM - 20
        img_y = HEADER_BOTTOM + 16
        buf = _floorplan_with_pins(fp, pins, label_pins=True)
        if buf is not None:
            try:
                pdf.image(buf, x=10, y=img_y, w=avail_w, h=avail_h, keep_aspect_ratio=True)
            except Exception:
                tc(*TEXT_MUTED)
                pdf.set_xy(10, img_y + 10)
                pdf.cell(avail_w, 6, _s("(Floorplan image could not be loaded.)"), align="C")
        else:
            tc(*TEXT_MUTED)
            pdf.set_xy(10, img_y + 10)
            pdf.cell(avail_w, 6, _s("(Floorplan image not available.)"), align="C")

    # Output
    data = pdf.output()
    if isinstance(data, str):
        return data.encode("latin-1")
    return bytes(data)


def _group_band(pdf, label: str, level: int = 1):
    """Render a group separator band at the current Y."""
    from fpdf import XPos, YPos
    NL = dict(new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    h = 7 if level == 1 else 5
    if level == 1:
        pdf.set_fill_color(*IPS_BLUE)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 10)
    else:
        pdf.set_fill_color(*LIGHT_GRAY)
        pdf.set_text_color(*IPS_BLUE)
        pdf.set_font("Helvetica", "B", 9)
    pdf.set_x(10)
    pdf.cell(pdf.w - 20, h, _s("  " + label), border=0, fill=True, **NL)
    pdf.set_text_color(*TEXT_DARK)
    pdf.ln(1)


# ── Background generation ────────────────────────────────────────────────────

def _sanitize_folder(name: str) -> str:
    """Sanitize a string for use as a folder/file name. Mirrors the helper
    in routers/areas_units.py so the project-folder convention stays
    consistent across modules."""
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', str(name)).strip("_. ") or "unknown"


def _filters_short_summary(filters: ExportFilters, item_count: int) -> str:
    """Compact one-liner stored on the Report row for the UI."""
    parts = []
    if filters.package_ids:
        parts.append(f"{len(filters.package_ids)} package(s)")
    else:
        parts.append("all packages")
    if filters.area_ids:
        parts.append(f"{len(filters.area_ids)} area(s)")
    else:
        parts.append("all areas")
    if filters.statuses:
        parts.append("status: " + ", ".join(filters.statuses))
    parts.append("grouped by " + _GROUP_LABELS.get(filters.group_by, filters.group_by))
    if filters.per_package_plans:
        parts.append("plans split per package")
    return f"{item_count} record(s) · " + " · ".join(parts)


def _run_safety_pdf_job(report_id: int) -> None:
    """Background worker. Opens its own DB session, regenerates the PDF, and
    writes status updates onto the Report row. Catches exceptions so a
    crash leaves a FAILED row instead of a stuck PENDING one."""
    db = SessionLocal()
    try:
        report = db.query(models.Report).filter_by(id=report_id).first()
        if not report:
            return
        report.status = "GENERATING"
        report.started_at = datetime.utcnow()
        db.commit()

        try:
            filters = ExportFilters(**json.loads(report.filters_json or "{}"))
        except Exception:
            filters = ExportFilters()

        project = db.query(models.Project).filter_by(id=report.project_id).first()
        if not project:
            raise RuntimeError("Project not found")

        q = db.query(models.SafetyObservation).filter(
            models.SafetyObservation.project_id == report.project_id
        )
        if filters.package_ids:
            q = q.filter(models.SafetyObservation.package_id.in_(filters.package_ids))
        if filters.area_ids:
            q = q.filter(models.SafetyObservation.area_id.in_(filters.area_ids))
        if filters.statuses:
            q = q.filter(models.SafetyObservation.status.in_(filters.statuses))
        obs_list = q.order_by(models.SafetyObservation.id).all()

        pdf_bytes = _build_safety_pdf(obs_list, filters, project, db)

        # Persist to uploads/{project}/Safety Reports/safety_observations_<ts>.pdf
        folder = (
            UPLOAD_ROOT
            / _sanitize_folder(project.project_number or f"project_{project.id}")
            / "Safety Reports"
        )
        folder.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"safety_observations_{ts}.pdf"
        out_path = folder / filename
        out_path.write_bytes(pdf_bytes)

        report.stored_path = str(out_path.relative_to(UPLOAD_ROOT))
        report.file_size = len(pdf_bytes)
        report.item_count = len(obs_list)
        report.filter_summary = _filters_short_summary(filters, len(obs_list))
        report.status = "READY"
        report.completed_at = datetime.utcnow()
        report.error_message = None
        db.commit()
    except Exception as e:
        try:
            report = db.query(models.Report).filter_by(id=report_id).first()
            if report:
                report.status = "FAILED"
                report.completed_at = datetime.utcnow()
                report.error_message = (str(e) or "Unknown error")[:1000]
                db.commit()
        except Exception:
            traceback.print_exc()
    finally:
        db.close()


# ── Endpoint ─────────────────────────────────────────────────────────────────

@export_router.post("/observations/export-pdf")
def enqueue_observations_pdf(
    body: ExportFilters,
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    """Enqueue a background PDF export. Returns the Report row immediately;
    the frontend polls /api/reports/{id} for status and downloads when ready."""
    if user.role == "BIDDER":
        raise HTTPException(403, "Bidders cannot export safety observations")

    project = db.query(models.Project).filter_by(id=user.project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    # Vendors can only export observations on packages they are linked to —
    # clamp the package filter at enqueue time so the background job never
    # widens scope.
    if user.role == "VENDOR":
        from routers.safety import _vendor_visible_package_ids
        visible = set(_vendor_visible_package_ids(user, db))
        if body.package_ids:
            body.package_ids = [pid for pid in body.package_ids if pid in visible]
        else:
            body.package_ids = list(visible)
        if not body.package_ids:
            raise HTTPException(403, "No packages linked to your account")

    title = "Safety Observations Report"
    report = models.Report(
        project_id=user.project_id,
        kind="safety",
        status="PENDING",
        title=title,
        filters_json=body.model_dump_json() if hasattr(body, "model_dump_json") else json.dumps(body.dict()),
        filter_summary=_filters_short_summary(body, 0),
        requested_by_id=user.id,
        requested_at=datetime.utcnow(),
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    threading.Thread(
        target=_run_safety_pdf_job,
        args=(report.id,),
        daemon=True,
    ).start()

    return {
        "id": report.id,
        "status": report.status,
        "kind": report.kind,
        "title": report.title,
        "filter_summary": report.filter_summary,
        "requested_at": report.requested_at.isoformat() + "Z",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel exports — observations / incidents / toolboxes
# Same pattern as construction's daily_reports xlsx export.
# ─────────────────────────────────────────────────────────────────────────────

_XLSX_DARK_BLUE  = "1E3A5F"
_XLSX_LIGHT_BLUE = "D6E4F0"
_XLSX_WHITE      = "FFFFFF"


def _xlsx_fmt_ts(dt) -> str:
    if not dt:
        return ""
    try:
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(dt)


def _xlsx_user_name(u) -> str:
    if not u:
        return ""
    return getattr(u, "name", None) or getattr(u, "email", "") or ""


def _xlsx_finalise(wb, ws, filename: str):
    """Style header row + auto-width + return StreamingResponse."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    fill = PatternFill("solid", fgColor=_XLSX_DARK_BLUE)
    for cell in ws[1]:
        if cell.value is not None:
            cell.font = Font(bold=True, color=_XLSX_WHITE)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _safety_status_label(s: str) -> str:
    return {
        "DRAFT":               "Draft",
        "SUBMITTED":           "Submitted",
        "RECEIVED":            "Received",
        "CLOSED":              "Closed",
        "UNDER_INVESTIGATION": "Under investigation",
        "ACTION_IN_PROGRESS":  "Action in progress",
        "PENDING_REVIEW":      "Pending review",
    }.get(s, s or "")


@export_router.get("/observations/export/excel")
def export_safety_observations_excel(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    rows = (
        db.query(models.SafetyObservation)
          .filter(models.SafetyObservation.project_id == user.project_id)
          .order_by(models.SafetyObservation.submitted_at.desc().nullslast(),
                    models.SafetyObservation.id.desc())
          .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Observations"
    ws.append([
        "Seq #", "Status", "Category", "Polarity",
        "Package", "Area",
        "Subcontractor", "Worker",
        "Details", "Remediation Request",
        "Submitted At", "Submitted By",
        "Acknowledged At", "Acknowledged By", "Acknowledge Comment",
        "Closed At", "Closed By",
        "Created By", "Created At", "Updated At",
    ])
    for r in rows:
        cat = r.category
        ws.append([
            r.project_seq_id or r.id,
            _safety_status_label(r.status),
            cat.name if cat else "",
            cat.polarity if cat else "",
            r.package.tag_number if r.package else "",
            r.area.tag if r.area else "",
            r.subcontractor.name if r.subcontractor else "",
            r.worker.name if r.worker else "",
            r.details or "",
            r.remediation_request or "",
            _xlsx_fmt_ts(r.submitted_at),
            _xlsx_user_name(r.created_by),  # creator is the submitter
            _xlsx_fmt_ts(r.acknowledged_at),
            _xlsx_user_name(r.acknowledged_by),
            r.acknowledge_comment or "",
            _xlsx_fmt_ts(r.closed_at),
            _xlsx_user_name(r.closed_by),
            _xlsx_user_name(r.created_by),
            _xlsx_fmt_ts(r.created_at),
            _xlsx_fmt_ts(r.updated_at),
        ])
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return _xlsx_finalise(wb, ws, f"safety_observations_{date_str}.xlsx")


@export_router.get("/incidents/export/excel")
def export_safety_incidents_excel(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    rows = (
        db.query(models.SafetyIncident)
          .filter(models.SafetyIncident.project_id == user.project_id)
          .order_by(models.SafetyIncident.incident_date.desc().nullslast(),
                    models.SafetyIncident.id.desc())
          .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidents"
    ws.append([
        "Seq #", "Status",
        "Severity (level)", "Severity Class",
        "Cause", "Other Cause",
        "Package", "Area", "Incident Date",
        "Workers Involved",
        "Details", "Action",
        "Submitted At", "Submitted By",
        "Investigated At", "Investigated By", "Investigation Comment",
        "Action Completed At", "Action Completed By", "Completion Comment",
        "Closed At", "Closed By",
        "Created By", "Created At", "Updated At",
    ])
    for r in rows:
        sev = r.severity_class
        cause = r.incident_cause
        worker_names = ", ".join(
            (siw.worker.name if siw.worker else "")
            for siw in (r.workers or [])
        )
        ws.append([
            r.project_seq_id or r.id,
            _safety_status_label(r.status),
            sev.level if sev else "",
            sev.name if sev else "",
            cause.name if cause else "",
            r.other_cause_text or "",
            r.package.tag_number if r.package else "",
            r.area.tag if r.area else "",
            r.incident_date or "",
            worker_names,
            r.details or "",
            r.action or "",
            _xlsx_fmt_ts(r.submitted_at),
            _xlsx_user_name(r.submitted_by),
            _xlsx_fmt_ts(r.investigated_at),
            _xlsx_user_name(r.investigated_by),
            r.investigation_comment or "",
            _xlsx_fmt_ts(r.action_completed_at),
            _xlsx_user_name(r.action_completed_by),
            r.action_completion_comment or "",
            _xlsx_fmt_ts(r.closed_at),
            _xlsx_user_name(r.closed_by),
            _xlsx_user_name(r.created_by),
            _xlsx_fmt_ts(r.created_at),
            _xlsx_fmt_ts(r.updated_at),
        ])
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return _xlsx_finalise(wb, ws, f"safety_incidents_{date_str}.xlsx")


@export_router.get("/toolboxes/export/excel")
def export_safety_toolboxes_excel(
    db: Session = Depends(get_db),
    user: auth.ProjectContext = Depends(auth.get_project_user),
):
    from openpyxl import Workbook
    rows = (
        db.query(models.SafetyToolbox)
          .filter(models.SafetyToolbox.project_id == user.project_id)
          .order_by(models.SafetyToolbox.talk_date.desc().nullslast(),
                    models.SafetyToolbox.id.desc())
          .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Toolbox Talks"
    ws.append([
        "Seq #", "Status",
        "Category", "Other Category",
        "Talk Date", "Given By",
        "Packages", "Attendees (count)", "Attendees",
        "Linked Observations", "Linked Incidents",
        "Details",
        "Submitted At", "Submitted By",
        "Acknowledged At", "Acknowledged By", "Acknowledge Comment",
        "Reopened At", "Reopened By",
        "Created By", "Created At", "Updated At",
    ])
    for r in rows:
        given_by = (r.given_by_user.name if r.given_by_user else
                    r.given_by_worker.name if r.given_by_worker else "")
        packages = ", ".join(
            (tp.package.tag_number if tp.package else "") for tp in (r.packages or [])
        )
        attendees = list(r.workers or [])
        attendee_names = ", ".join(
            (tw.worker.name if tw.worker else "") for tw in attendees
        )
        obs_seqs = ", ".join(
            f"#{(to.observation.project_seq_id or to.observation.id)}"
            for to in (r.observations or []) if to.observation
        )
        inc_seqs = ", ".join(
            f"#{(ti.incident.project_seq_id or ti.incident.id)}"
            for ti in (r.incidents or []) if ti.incident
        )
        ws.append([
            r.project_seq_id or r.id,
            _safety_status_label(r.status),
            r.category.name if r.category else "",
            r.other_category_text or "",
            r.talk_date or "",
            given_by,
            packages,
            len(attendees),
            attendee_names,
            obs_seqs,
            inc_seqs,
            r.details or "",
            _xlsx_fmt_ts(r.submitted_at),
            _xlsx_user_name(r.submitted_by),
            _xlsx_fmt_ts(r.acknowledged_at),
            _xlsx_user_name(r.acknowledged_by),
            r.acknowledge_comment or "",
            _xlsx_fmt_ts(r.reopened_at),
            _xlsx_user_name(r.reopened_by),
            _xlsx_user_name(r.created_by),
            _xlsx_fmt_ts(r.created_at),
            _xlsx_fmt_ts(r.updated_at),
        ])
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    return _xlsx_finalise(wb, ws, f"safety_toolboxes_{date_str}.xlsx")
